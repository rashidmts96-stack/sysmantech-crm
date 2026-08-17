from flask import Flask, render_template, request, redirect, flash, session, jsonify, Response, send_from_directory
from flask.sessions import SecureCookieSessionInterface
import mysql.connector
from mysql.connector import Error, pooling
from mysql.connector.errors import PoolError
import os
import re
import csv
import io
import secrets
import time
from datetime import datetime, timedelta
from threading import BoundedSemaphore, Lock
from urllib.parse import quote, urlencode, urlparse
from uuid import uuid4
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from dashboard_settings_routes import register_dashboard_settings_routes
from job_transfer_rules import ACTIVE_TRANSFER_STATUSES, COMPLETED_TRANSFER_STATUSES, TRANSFER_STATUSES, allowed_next_transfer_statuses, can_transition_transfer_status, compute_job_transfer_split, normalize_transfer_status, summarize_job_transfers
from onsite_calls_blueprint import create_onsite_calls_blueprint
from syscare_routes import register_syscare_routes
from time_utils import business_now_naive, format_date_display as format_date_display_value, format_datetime_display as format_datetime_display_value, mysql_session_timezone_value, normalize_display_datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


APP_DEBUG = os.getenv("APP_DEBUG", "0") == "1"
MAX_UPLOAD_REQUEST_MB = max(10, int(os.getenv("MAX_UPLOAD_REQUEST_MB", "50") or "50"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")


class RequestAwareSessionInterface(SecureCookieSessionInterface):
    def get_cookie_secure(self, app):
        configured_value = os.getenv("SESSION_COOKIE_SECURE", "").strip().lower()
        if configured_value in {"1", "true", "yes", "on"}:
            return True
        if configured_value in {"0", "false", "no", "off"}:
            return False
        return request.is_secure


def _load_config_value_from_file(name):
    configured_path = os.getenv(f"{name}_FILE", "").strip()
    candidate_paths = []
    if configured_path:
        candidate_paths.append(configured_path)
    candidate_paths.append(os.path.join(INSTANCE_DIR, name.lower()))

    for candidate_path in candidate_paths:
        if not candidate_path or not os.path.exists(candidate_path):
            continue
        with open(candidate_path, "r", encoding="utf-8") as config_handle:
            configured_value = config_handle.read().strip()
        if configured_value:
            return configured_value

    return ""


def _require_config_value(name, allow_debug_fallback=False, fallback_value=""):
    value = os.getenv(name, "").strip()
    if value:
        return value
    file_value = _load_config_value_from_file(name)
    if file_value:
        return file_value
    if allow_debug_fallback and APP_DEBUG:
        return fallback_value
    raise RuntimeError(f"{name} environment variable, {name}_FILE, or instance/{name.lower()} is required")


def _load_secret_key():
    configured_secret = os.getenv("SECRET_KEY", "").strip()
    if configured_secret:
        return configured_secret
    secret_file = os.getenv("SECRET_KEY_FILE", "").strip() or os.path.join(INSTANCE_DIR, "secret_key")
    if os.path.exists(secret_file):
        with open(secret_file, "r", encoding="utf-8") as secret_handle:
            stored_secret = secret_handle.read().strip()
        if stored_secret:
            return stored_secret

    secret_dir = os.path.dirname(secret_file)
    if secret_dir:
        os.makedirs(secret_dir, exist_ok=True)

    generated_secret = secrets.token_hex(32)
    with open(secret_file, "w", encoding="utf-8") as secret_handle:
        secret_handle.write(generated_secret)
    return generated_secret

# Flask app initialization (single instance, at the top)
app = Flask(__name__, instance_path=INSTANCE_DIR)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
app.session_interface = RequestAwareSessionInterface()
app.secret_key = _load_secret_key()
app.config.update(
    JOB_PHOTO_FOLDER=os.path.join(app.instance_path, "job_photos"),
    PROFILE_PICTURE_FOLDER=os.path.join(app.instance_path, "profile_pictures"),
    MAX_CONTENT_LENGTH=MAX_UPLOAD_REQUEST_MB * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    CSRF_ENFORCE_ALL_POSTS=os.getenv("CSRF_ENFORCE_ALL_POSTS", "1").strip().lower() in {"1", "true", "yes", "on"},
    CSRF_PROTECTED_ENDPOINTS=(),
    CSRF_EXEMPT_ENDPOINTS=("static",),
)
for _upload_dir in [app.config["JOB_PHOTO_FOLDER"], app.config["PROFILE_PICTURE_FOLDER"]]:
    os.makedirs(_upload_dir, exist_ok=True)


## (Removed duplicate Flask app initialization)


@app.route("/manifest.webmanifest")
def manifest_webmanifest():
    return send_from_directory(
        app.static_folder,
        "manifest.webmanifest",
        mimetype="application/manifest+json",
        max_age=300,
    )


@app.route("/service-worker.js")
def service_worker():
    response = send_from_directory(
        app.static_folder,
        "service-worker.js",
        mimetype="application/javascript",
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Service-Worker-Allowed"] = "/"
    return response

DB_NAME = os.getenv("DB_NAME", "crm_system")
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_USER = os.getenv("DB_USER", "root")
DB_PASSWORD = _require_config_value("DB_PASSWORD", allow_debug_fallback=True)
DB_POOL_NAME = os.getenv("DB_POOL_NAME", "crm_pool")
DB_POOL_CAP = max(1, int(getattr(pooling, "CNX_POOL_MAXSIZE", 32) or 32))
DB_POOL_SIZE = max(1, min(int(os.getenv("DB_POOL_SIZE", "24") or "24"), DB_POOL_CAP))
DB_POOL_ACQUIRE_TIMEOUT_SECONDS = max(1.0, float(os.getenv("DB_POOL_ACQUIRE_TIMEOUT_SECONDS", "20") or "20"))
DEFAULT_WAITRESS_THREADS = max(4, min(DB_POOL_SIZE, 24))
APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("APP_PORT", "5000"))
_db_pool = None
_db_pool_semaphore = None
_db_pool_lock = Lock()
LOGIN_ATTEMPT_WINDOW_SECONDS = 900
LOGIN_ATTEMPT_LIMIT = 5
_login_attempts = {}
PASSWORD_HASH_PREFIXES = ("pbkdf2:", "scrypt:")
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
MAX_JOB_ATTACHMENTS = 10
ALLOWED_SPARES_BILLING_STATUSES = ("Pending", "Billed", "Not Required")
JOB_ATTACHMENT_RETENTION_DAYS = max(30, int(os.getenv("JOB_ATTACHMENT_RETENTION_DAYS", "90") or "90"))
JOB_ATTACHMENT_CLEANUP_INTERVAL_SECONDS = max(300, int(os.getenv("JOB_ATTACHMENT_CLEANUP_INTERVAL_SECONDS", "3600") or "3600"))
_job_attachment_cleanup_state = {"last_run": 0.0}
CSRF_SESSION_KEY = "_csrf_token"
CSRF_FORM_FIELD = "csrf_token"
CSRF_HEADER_NAME = "X-CSRF-Token"
CSRF_HEADER_NAMES = (CSRF_HEADER_NAME, "X-CSRFToken")


def _request_origin_is_allowed():
    fetch_site = (request.headers.get("Sec-Fetch-Site") or "").strip().lower()
    if fetch_site in {"same-origin", "none"}:
        return True

    allowed_origin = request.host_url.rstrip("/")
    candidates = [request.headers.get("Origin", ""), request.headers.get("Referer", "")]
    saw_origin_signal = False

    for raw_value in candidates:
        if not raw_value:
            continue
        saw_origin_signal = True
        parsed = urlparse(raw_value)
        candidate_origin = f"{parsed.scheme}://{parsed.netloc}".rstrip("/")
        if candidate_origin == allowed_origin:
            return True

    if request.path == "/login" and "username" not in session and not saw_origin_signal and not fetch_site:
        return True

    return False


def _get_or_create_csrf_token():
    token = str(session.get(CSRF_SESSION_KEY) or "").strip()
    if token:
        return token

    token = secrets.token_urlsafe(32)
    session[CSRF_SESSION_KEY] = token
    return token


def _get_submitted_csrf_token():
    for header_name in CSRF_HEADER_NAMES:
        header_value = (request.headers.get(header_name) or "").strip()
        if header_value:
            return header_value

    form_value = (request.form.get(CSRF_FORM_FIELD) or "").strip()
    if form_value:
        return form_value

    if request.is_json:
        payload = request.get_json(silent=True)
        if isinstance(payload, dict):
            return str(payload.get(CSRF_FORM_FIELD) or "").strip()

    return ""


def _csrf_tokens_match(expected_token, submitted_token):
    expected_text = str(expected_token or "").strip()
    submitted_text = str(submitted_token or "").strip()
    if not expected_text or not submitted_text:
        return False
    return secrets.compare_digest(expected_text, submitted_text)


def _get_safe_referrer_path():
    raw_referrer = (request.referrer or "").strip()
    if not raw_referrer:
        return ""

    parsed_referrer = urlparse(raw_referrer)
    candidate_origin = f"{parsed_referrer.scheme}://{parsed_referrer.netloc}".rstrip("/")
    allowed_origin = request.host_url.rstrip("/")
    if candidate_origin != allowed_origin:
        return ""

    safe_path = parsed_referrer.path or ""
    if parsed_referrer.query:
        safe_path += f"?{parsed_referrer.query}"
    return safe_path


def _csrf_failure_response():
    if request.is_json:
        return jsonify({"success": False, "message": "Invalid or missing CSRF token"}), 400

    flash("Your session expired. Refresh the page and try again.", "danger")
    redirect_target = _get_safe_referrer_path() or ("/dashboard" if "username" in session else "/login")
    return redirect(redirect_target)


def _csrf_protection_applies(endpoint_name):
    if request.method in {"GET", "HEAD", "OPTIONS", "TRACE"}:
        return False

    if not endpoint_name:
        return False

    exempt_endpoints = set(app.config.get("CSRF_EXEMPT_ENDPOINTS") or ())
    if endpoint_name in exempt_endpoints:
        return False

    if app.config.get("CSRF_ENFORCE_ALL_POSTS", False):
        return True

    protected_endpoints = set(app.config.get("CSRF_PROTECTED_ENDPOINTS") or ())
    return endpoint_name in protected_endpoints


def _add_csrf_protected_endpoints(*endpoint_names):
    current_endpoints = set(app.config.get("CSRF_PROTECTED_ENDPOINTS") or ())
    for endpoint_name in endpoint_names:
        if endpoint_name:
            current_endpoints.add(endpoint_name)
    app.config["CSRF_PROTECTED_ENDPOINTS"] = tuple(sorted(current_endpoints))


_add_csrf_protected_endpoints(
    "login",
    "delete_job",
    "delete_job_attachment",
    "delete_quotation",
    "job_spare_billing",
    "close_case",
    "reopen_case",
    "add_staff_member",
    "edit_staff_member",
    "delete_staff_member",
    "delete_revenue_entry",
    "bulk_delete_revenue_entries",
)


def _build_global_nav_user():
    username = str(session.get("username") or "").strip()
    role = str(session.get("role") or "").strip()
    branch = str(session.get("branch") or "").strip()

    if not username:
        return {
            "is_authenticated": False,
            "username": "",
            "display_name": "",
            "initial": "U",
            "profile_picture_url": "",
            "role": role,
            "branch": branch,
            "show_settings": False,
        }

    nav_user = {
        "is_authenticated": True,
        "username": username,
        "display_name": username.upper(),
        "initial": username[:1].upper(),
        "profile_picture_url": "",
        "role": role,
        "branch": branch,
        "show_settings": role == "super_admin",
    }

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute(
            "SELECT username, profile_picture FROM users WHERE username=%s",
            (username,),
        )
        user_row = cursor.fetchone() or {}
        resolved_username = str(user_row.get("username") or username).strip() or username
        nav_user["username"] = resolved_username
        nav_user["display_name"] = resolved_username.upper()
        nav_user["initial"] = resolved_username[:1].upper()
        if user_row.get("profile_picture"):
            nav_user["profile_picture_url"] = f"/profile-picture/{resolved_username}"
    except Exception:
        pass
    finally:
        _safe_close(cursor, db)

    return nav_user


@app.context_processor
def _inject_global_nav_user():
    return {
        "global_nav_user": _build_global_nav_user(),
    }


@app.context_processor
def _inject_csrf_template_helpers():
    return {
        "csrf_token": _get_or_create_csrf_token,
        "csrf_field_name": CSRF_FORM_FIELD,
        "csrf_header_name": CSRF_HEADER_NAME,
        "job_attachment_retention_days": JOB_ATTACHMENT_RETENTION_DAYS,
        "format_datetime_display": format_datetime_display,
        "format_date_display": format_date_display,
    }


def _normalize_whatsapp_message(message):
    cleaned_lines = [re.sub(r"\s+", " ", line).strip() for line in str(message or "").splitlines()]
    normalized_message = "\n".join(line for line in cleaned_lines if line)
    if not normalized_message:
        normalized_message = ""
    return normalized_message


def _normalize_whatsapp_number(phone_number):
    digits = re.sub(r"\D", "", str(phone_number or ""))
    if not digits:
        return ""
    if digits.startswith("00") and len(digits) > 2:
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"91{digits}"
    if 10 <= len(digits) <= 15:
        return digits
    return ""


def _build_whatsapp_chat_url(phone_number, message):
    normalized_message = _normalize_whatsapp_message(message)
    normalized_number = _normalize_whatsapp_number(phone_number)
    if normalized_number:
        return f"https://wa.me/{normalized_number}?text={quote(normalized_message)}"
    return f"https://wa.me/?text={quote(normalized_message)}"


def _build_whatsapp_share_url(message):
    return _build_whatsapp_chat_url("", message)


@app.context_processor
def _inject_communication_template_helpers():
    return {
        "build_whatsapp_chat_url": _build_whatsapp_chat_url,
        "build_whatsapp_share_url": _build_whatsapp_share_url,
    }


@app.before_request
def _enforce_same_origin_protection():
    if request.method in {"GET", "HEAD", "OPTIONS"}:
        return None
    if _request_origin_is_allowed():
        return None
    app.logger.warning("Blocked unsafe cross-origin request to %s", request.path)
    return "Forbidden", 403


@app.before_request
def _enforce_csrf_protection():
    if not _csrf_protection_applies(request.endpoint):
        return None

    if _csrf_tokens_match(session.get(CSRF_SESSION_KEY), _get_submitted_csrf_token()):
        return None

    app.logger.warning("Blocked request with invalid CSRF token to %s", request.path)
    return _csrf_failure_response()


@app.before_request
def _cleanup_expired_job_attachments_before_request():
    if request.endpoint == "static":
        return None
    _purge_expired_job_attachments()
    return None


@app.after_request
def _add_security_headers(response):
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "same-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net"
    )
    if request.is_secure:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


@app.errorhandler(RequestEntityTooLarge)
def _handle_request_entity_too_large(error):
    max_size_mb = max(1, int((app.config.get("MAX_CONTENT_LENGTH") or 0) / (1024 * 1024)))
    if request.is_json:
        return jsonify({
            "success": False,
            "message": f"Upload too large. Total request size must be under {max_size_mb} MB.",
        }), 413

    flash(f"Upload too large. Total request size must be under {max_size_mb} MB.", "danger")
    return redirect(_get_safe_referrer_path() or request.url)


def _prune_login_attempts(client_key):
    now = time.time()
    attempts = [ts for ts in _login_attempts.get(client_key, []) if now - ts < LOGIN_ATTEMPT_WINDOW_SECONDS]
    _login_attempts[client_key] = attempts
    return attempts


def _record_failed_login_attempt(client_key):
    attempts = _prune_login_attempts(client_key)
    attempts.append(time.time())
    _login_attempts[client_key] = attempts


def _clear_login_attempts(client_key):
    _login_attempts.pop(client_key, None)


def _login_rate_limited(client_key):
    return len(_prune_login_attempts(client_key)) >= LOGIN_ATTEMPT_LIMIT


def _get_client_key():
    return (request.remote_addr or "unknown").strip() or "unknown"


def _is_password_hash(value):
    text = str(value or "")
    return text.startswith(PASSWORD_HASH_PREFIXES)


def _hash_password(password):
    return generate_password_hash(password)


def _has_legacy_password(stored_password):
    stored_password = str(stored_password or "").strip()
    return bool(stored_password) and not _is_password_hash(stored_password)


def _verify_password(stored_password, candidate_password):
    stored_password = str(stored_password or "")
    candidate_password = str(candidate_password or "")
    if not stored_password:
        return False, False
    if _has_legacy_password(stored_password):
        return False, True
    return check_password_hash(stored_password, candidate_password), False


def _flash_internal_error(user_message, exc=None):
    if exc is not None:
        app.logger.exception(user_message)
    flash(user_message, "danger")


def _json_internal_error(user_message, status_code=500, exc=None, payload=None):
    if exc is not None:
        app.logger.exception(user_message)
    body = payload or {}
    body.setdefault("success", False)
    body.setdefault("message", user_message)
    return jsonify(body), status_code


def _sniff_image_extension(file_storage):
    header = file_storage.stream.read(16)
    file_storage.stream.seek(0)

    if header.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if header.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if header.startswith(b"GIF87a") or header.startswith(b"GIF89a"):
        return ".gif"
    if header.startswith(b"RIFF") and header[8:12] == b"WEBP":
        return ".webp"
    return ""


def _save_image_upload(file_storage, destination_folder, prefix):
    original_name = secure_filename(file_storage.filename or "")
    provided_ext = os.path.splitext(original_name)[1].lower()
    sniffed_ext = _sniff_image_extension(file_storage)

    if not sniffed_ext:
        raise ValueError("Upload must be a valid image file")

    normalized_ext = ".jpg" if provided_ext == ".jpeg" else provided_ext
    if normalized_ext not in ALLOWED_IMAGE_EXTENSIONS or sniffed_ext != normalized_ext:
        raise ValueError("Only JPG, PNG, GIF and WEBP images are allowed")

    stored_name = f"{prefix}_{uuid4().hex}{sniffed_ext}"
    file_storage.save(os.path.join(destination_folder, stored_name))
    return stored_name


def _get_job_upload_files(files):
    uploads = [file for file in files.getlist("photos") if file and file.filename]

    legacy_photo = files.get("photo")
    if legacy_photo and legacy_photo.filename:
        uploads.append(legacy_photo)

    return uploads


def _remove_saved_images(filenames, destination_folder):
    for filename in filenames or []:
        if not filename:
            continue
        try:
            file_path = os.path.join(destination_folder, filename)
            if os.path.exists(file_path):
                os.remove(file_path)
        except OSError:
            app.logger.warning("Could not remove uploaded image %s", filename)


def _save_job_images(upload_files, job_key):
    saved_filenames = []
    for file_index, photo in enumerate(upload_files, start=1):
        saved_filenames.append(
            _save_image_upload(
                photo,
                app.config["JOB_PHOTO_FOLDER"],
                f"job_{job_key}_{file_index}",
            )
        )
    return saved_filenames


def _insert_job_attachments(cursor, job_id, filenames):
    if not filenames:
        return

    cursor.executemany(
        "INSERT INTO job_attachments (job_id, filename) VALUES (%s, %s)",
        [(job_id, filename) for filename in filenames],
    )


def _load_job_attachments(cursor, job_id, primary_photo=None):
    cursor.execute(
        "SELECT id, filename, uploaded_at FROM job_attachments WHERE job_id=%s ORDER BY id ASC",
        (job_id,),
    )
    attachment_rows = cursor.fetchall() or []

    attachments = []
    seen_filenames = set()
    for row in attachment_rows:
        filename = str(row.get("filename") or "").strip()
        if not filename or filename in seen_filenames:
            continue
        seen_filenames.add(filename)
        attachments.append(
            {
                "id": row.get("id"),
                "filename": filename,
                "uploaded_at": row.get("uploaded_at"),
                "legacy": False,
            }
        )

    primary_filename = str(primary_photo or "").strip()
    if primary_filename and primary_filename not in seen_filenames:
        attachments.insert(
            0,
            {
                "id": None,
                "filename": primary_filename,
                "uploaded_at": None,
                "legacy": True,
            },
        )

    return attachments


def _job_can_access_attachment(cursor, job, filename):
    normalized_filename = str(filename or "").strip()
    if not normalized_filename:
        return False

    if normalized_filename == str(job.get("photo") or "").strip():
        return True

    cursor.execute(
        "SELECT id FROM job_attachments WHERE job_id=%s AND filename=%s LIMIT 1",
        (job["id"], normalized_filename),
    )
    return cursor.fetchone() is not None


def _normalize_spares_billing_status(value, has_used_spares=False):
    lookup = {status.lower(): status for status in ALLOWED_SPARES_BILLING_STATUSES}
    normalized = lookup.get(str(value or "").strip().lower(), "")
    if normalized == "Not Required" and has_used_spares:
        return "Pending"
    if normalized:
        return normalized
    return "Pending" if has_used_spares else "Not Required"


def _job_used_spares_signature(used_spares):
    signature = []
    for item in used_spares or []:
        spare_name = str(item.get("spare_name") or "").strip()
        try:
            amount_value = float(item.get("amount") or 0)
        except (TypeError, ValueError):
            amount_value = 0.0
        signature.append((spare_name, f"{amount_value:.2f}"))
    return signature


def _saved_used_spares_preserved(existing_spares, updated_spares):
    remaining_signatures = list(_job_used_spares_signature(updated_spares))
    for signature in _job_used_spares_signature(existing_spares):
        if signature not in remaining_signatures:
            return False
        remaining_signatures.remove(signature)
    return True


def _set_job_spares_billing(cursor, job_id, status, invoice_no=None, invoice_date=None, billed_by=None, notes=None):
    normalized_status = _normalize_spares_billing_status(status, False)
    normalized_invoice_no = str(invoice_no or "").strip() or None
    normalized_notes = str(notes or "").strip() or None
    cursor.execute(
        """
        UPDATE jobs
        SET spares_billing_status=%s,
            spares_invoice_no=%s,
            spares_invoice_date=%s,
            spares_billed_by=%s,
            spares_billing_notes=%s
        WHERE id=%s
        """,
        (normalized_status, normalized_invoice_no, invoice_date, billed_by, normalized_notes, job_id),
    )


def _build_spares_billing_summary(job, used_spares):
    has_used_spares = bool(used_spares)
    invoice_no = str(job.get("spares_invoice_no") or "").strip()
    invoice_date = job.get("spares_invoice_date")
    status = _normalize_spares_billing_status(job.get("spares_billing_status"), has_used_spares)

    if has_used_spares and invoice_no:
        status = "Billed"
    elif has_used_spares:
        status = "Pending"
    else:
        status = "Not Required"

    invoice_date_value = ""
    if invoice_date:
        if hasattr(invoice_date, "strftime"):
            invoice_date_value = invoice_date.strftime("%Y-%m-%d")
        else:
            invoice_date_value = _normalize_date_input(str(invoice_date))

    return {
        "has_used_spares": has_used_spares,
        "used_spares_count": len(used_spares or []),
        "status": status,
        "invoice_no": invoice_no,
        "invoice_date": invoice_date,
        "invoice_date_value": invoice_date_value,
        "invoice_date_display": format_date_display(invoice_date),
        "billed_by": str(job.get("spares_billed_by") or "").strip(),
        "notes": str(job.get("spares_billing_notes") or "").strip(),
        "pending": has_used_spares and status != "Billed",
    }


def _normalize_attachment_filename(filename):
    return os.path.basename(str(filename or "").strip())


def _user_can_manage_job_attachments(role, job):
    if role not in ["super_admin", "admin", "coordinator"]:
        return False
    if job.get("closure_status") and role not in ["super_admin", "admin"]:
        return False
    return True


def _delete_job_attachment_record(cursor, job, filename):
    normalized_filename = _normalize_attachment_filename(filename)
    attachments = _load_job_attachments(cursor, job["id"], job.get("photo"))
    matching_attachment = next((item for item in attachments if item.get("filename") == normalized_filename), None)
    if not matching_attachment:
        return "", []

    remaining_filenames = [item["filename"] for item in attachments if item.get("filename") != normalized_filename]
    cursor.execute("DELETE FROM job_attachments WHERE job_id=%s AND filename=%s", (job["id"], normalized_filename))

    if normalized_filename == str(job.get("photo") or "").strip():
        next_primary = remaining_filenames[0] if remaining_filenames else None
        cursor.execute("UPDATE jobs SET photo=%s WHERE id=%s", (next_primary, job["id"]))

    return normalized_filename, remaining_filenames


def _job_all_attachment_filenames(cursor, job):
    filenames = []
    seen_filenames = set()
    for attachment in _load_job_attachments(cursor, job["id"], job.get("photo")):
        filename = str(attachment.get("filename") or "").strip()
        if not filename or filename in seen_filenames:
            continue
        seen_filenames.add(filename)
        filenames.append(filename)
    return filenames


def _delete_job_record(cursor, job_id):
    cursor.execute("DELETE FROM job_status_logs WHERE job_id=%s", (job_id,))
    cursor.execute("DELETE FROM job_attachments WHERE job_id=%s", (job_id,))
    cursor.execute("DELETE FROM used_spares WHERE job_id=%s", (job_id,))
    cursor.execute("DELETE FROM jobs WHERE id=%s", (job_id,))
    return cursor.rowcount


def _purge_expired_job_attachments(force=False):
    now_seconds = time.time()
    last_run = float(_job_attachment_cleanup_state.get("last_run") or 0.0)
    if not force and now_seconds - last_run < JOB_ATTACHMENT_CLEANUP_INTERVAL_SECONDS:
        return 0

    _job_attachment_cleanup_state["last_run"] = now_seconds
    cutoff = business_now_naive() - timedelta(days=JOB_ATTACHMENT_RETENTION_DAYS)
    db = None
    cursor = None
    filenames_to_remove = set()
    job_ids_to_clear = set()

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT id, photo
            FROM jobs
            WHERE closure_date IS NOT NULL
              AND closure_date <= %s
              AND photo IS NOT NULL
              AND TRIM(photo) <> ''
            """,
            (cutoff,),
        )
        for row in cursor.fetchall() or []:
            job_ids_to_clear.add(row["id"])
            normalized_filename = _normalize_attachment_filename(row.get("photo"))
            if normalized_filename:
                filenames_to_remove.add(normalized_filename)

        cursor.execute(
            """
            SELECT ja.job_id, ja.filename
            FROM job_attachments ja
            INNER JOIN jobs j ON j.id = ja.job_id
            WHERE j.closure_date IS NOT NULL
              AND j.closure_date <= %s
            """,
            (cutoff,),
        )
        for row in cursor.fetchall() or []:
            job_ids_to_clear.add(row["job_id"])
            normalized_filename = _normalize_attachment_filename(row.get("filename"))
            if normalized_filename:
                filenames_to_remove.add(normalized_filename)

        if not job_ids_to_clear and not filenames_to_remove:
            return 0

        cursor.execute(
            """
            DELETE ja
            FROM job_attachments ja
            INNER JOIN jobs j ON j.id = ja.job_id
            WHERE j.closure_date IS NOT NULL
              AND j.closure_date <= %s
            """,
            (cutoff,),
        )

        if job_ids_to_clear:
            placeholders = ", ".join(["%s"] * len(job_ids_to_clear))
            cursor.execute(
                f"UPDATE jobs SET photo=NULL WHERE id IN ({placeholders})",
                tuple(sorted(job_ids_to_clear)),
            )

        db.commit()
    except Exception:
        if db:
            db.rollback()
        app.logger.exception("Failed to purge expired job attachments")
        return 0
    finally:
        _safe_close(cursor, db)

    _remove_saved_images(sorted(filenames_to_remove), app.config["JOB_PHOTO_FOLDER"])
    return len(filenames_to_remove)

def get_age_group(days):
    if days <= 2:
        return "0-2"
    if days <= 5:
        return "3-5"
    if days <= 10:
        return "6-10"
    if days <= 15:
        return "11-15"
    if days <= 30:
        return "16-30"
    if days <= 90:
        return "31-90"
    if days <= 180:
        return "91-180"
    return "181+"


def format_datetime_display(dt_value):
    """Format datetime as dd/mm/yyyy hh:mm for UI display."""
    return format_datetime_display_value(dt_value)


def format_date_display(date_value):
    return format_date_display_value(date_value)


def _normalize_date_input(date_text):
    """Return YYYY-MM-DD if valid, else empty string."""
    if not date_text:
        return ""
    s = date_text.strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return ""


def _get_multi_values(args, key):
    """Read repeated query params and normalize as a unique list."""
    values = [v.strip() for v in args.getlist(key) if v and v.strip()]
    if not values:
        one_value = args.get(key, "").strip()
        if one_value:
            values = [one_value]
    seen = set()
    result = []
    for v in values:
        if v not in seen:
            seen.add(v)
            result.append(v)
    return result


def _normalize_option_list(values, keep_all_first=False):
    """Trim and dedupe dropdown option values; optionally keep single ALL at top."""
    cleaned = []
    seen = set()
    has_all = False

    for raw in values:
        if raw is None:
            continue
        v = str(raw).strip()
        if not v:
            continue

        if keep_all_first and v.upper() == "ALL":
            has_all = True
            continue

        key = v.upper()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(v)

    cleaned.sort(key=lambda x: x.upper())
    if keep_all_first and has_all:
        return ["ALL"] + cleaned
    return cleaned


def _load_engineer_usernames(cursor):
    engineer_values = []
    queries = [
        """
        SELECT username AS engineer_name
        FROM users
        WHERE LOWER(TRIM(COALESCE(role, '')))='engineer'
          AND username IS NOT NULL
          AND TRIM(username) <> ''
        """,
        """
        SELECT DISTINCT assigned_engineer AS engineer_name
        FROM jobs
        WHERE assigned_engineer IS NOT NULL
          AND TRIM(assigned_engineer) <> ''
        """,
        """
        SELECT DISTINCT staff_name AS engineer_name
        FROM staff_directory
        WHERE staff_name IS NOT NULL
          AND TRIM(staff_name) <> ''
          AND resigned_date IS NULL
        """,
    ]
    for query in queries:
        try:
            cursor.execute(query)
            engineer_values.extend(
                str(row.get("engineer_name") or "").strip()
                for row in cursor.fetchall()
                if str(row.get("engineer_name") or "").strip()
            )
        except Error:
            continue
    return _normalize_option_list(engineer_values)


def _validate_assigned_engineer_name(assigned_engineer, engineer_usernames):
    normalized = str(assigned_engineer or "").strip()
    allowed_lookup = {
        str(name or "").strip().casefold(): str(name or "").strip()
        for name in engineer_usernames
        if str(name or "").strip()
    }
    return allowed_lookup.get(normalized.casefold(), "")


def _resolve_upload_headers(headers, required_groups):
    normalized_map = {}
    for idx, header in enumerate(headers):
        key = re.sub(r'[^a-z0-9]+', '', str(header or '').strip().lower())
        if key and key not in normalized_map:
            normalized_map[key] = idx

    resolved = {}
    for field_name, aliases in required_groups.items():
        match_idx = None
        for alias in aliases:
            if alias in normalized_map:
                match_idx = normalized_map[alias]
                break
        if match_idx is None:
            raise ValueError(f"Missing required column: {field_name}")
        resolved[field_name] = match_idx

    return resolved


def _iter_user_branch_upload_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    required_headers = {
        "username": ["username", "user", "username"],
        "branch": ["branch", "branchname"],
    }

    if ext == ".csv":
        raw_bytes = uploaded_file.stream.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("utf-8", errors="ignore")

        reader = csv.reader(io.StringIO(text))
        headers = next(reader, None)
        if not headers:
            raise ValueError("The file is empty")

        header_map = _resolve_upload_headers(headers, required_headers)
        for row_number, row in enumerate(reader, start=2):
            username = row[header_map["username"]] if len(row) > header_map["username"] else ""
            branch = row[header_map["branch"]] if len(row) > header_map["branch"] else ""
            yield row_number, str(username or "").strip(), str(branch or "").strip()
        return

    if ext == ".xlsx":
        if load_workbook is None:
            raise ValueError("Excel upload needs openpyxl installed. For now, use CSV or install openpyxl.")

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                raise ValueError("The file is empty")

            header_map = _resolve_upload_headers(headers, required_headers)
            for row_number, row in enumerate(rows, start=2):
                row = row or ()
                username = row[header_map["username"]] if len(row) > header_map["username"] else ""
                branch = row[header_map["branch"]] if len(row) > header_map["branch"] else ""
                yield row_number, str(username or "").strip(), str(branch or "").strip()
        finally:
            workbook.close()
        return

    raise ValueError("Please upload a .xlsx or .csv file")


def _iter_users_upload_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    required_headers = {
        "username": ["username", "user"],
        "password": ["password", "pass"],
        "role": ["role", "rolename"],
    }
    if ext == ".csv":
        raw_bytes = uploaded_file.stream.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader, None)
        if not headers:
            raise ValueError("The file is empty")
        header_map = _resolve_upload_headers(headers, required_headers)
        for row_number, row in enumerate(reader, start=2):
            username = row[header_map["username"]] if len(row) > header_map["username"] else ""
            password = row[header_map["password"]] if len(row) > header_map["password"] else ""
            role = row[header_map["role"]] if len(row) > header_map["role"] else ""
            yield row_number, str(username or "").strip(), str(password or "").strip(), str(role or "").strip()
        return
    if ext == ".xlsx":
        if load_workbook is None:
            raise ValueError("Excel needs openpyxl. Use CSV or install openpyxl.")
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                raise ValueError("The file is empty")
            header_map = _resolve_upload_headers(headers, required_headers)
            for row_number, row in enumerate(rows, start=2):
                row = row or ()
                username = row[header_map["username"]] if len(row) > header_map["username"] else ""
                password = row[header_map["password"]] if len(row) > header_map["password"] else ""
                role = row[header_map["role"]] if len(row) > header_map["role"] else ""
                yield row_number, str(username or "").strip(), str(password or "").strip(), str(role or "").strip()
        finally:
            workbook.close()
        return
    raise ValueError("Please upload .xlsx or .csv")


def _iter_revenue_targets_upload_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    required_headers = {"branch": ["branch", "branchname"], "target": ["target", "totaltarget", "total_target"]}
    if ext == ".csv":
        raw_bytes = uploaded_file.stream.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader, None)
        if not headers:
            raise ValueError("The file is empty")
        header_map = _resolve_upload_headers(headers, required_headers)
        for row_number, row in enumerate(reader, start=2):
            branch = row[header_map["branch"]] if len(row) > header_map["branch"] else ""
            target = row[header_map["target"]] if len(row) > header_map["target"] else ""
            yield row_number, str(branch or "").strip(), str(target or "").strip()
        return
    if ext == ".xlsx":
        if load_workbook is None:
            raise ValueError("Excel needs openpyxl. Use CSV or install openpyxl.")
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                raise ValueError("The file is empty")
            header_map = _resolve_upload_headers(headers, required_headers)
            for row_number, row in enumerate(rows, start=2):
                row = row or ()
                branch = row[header_map["branch"]] if len(row) > header_map["branch"] else ""
                target = row[header_map["target"]] if len(row) > header_map["target"] else ""
                yield row_number, str(branch or "").strip(), str(target or "").strip()
        finally:
            workbook.close()
        return
    raise ValueError("Please upload .xlsx or .csv")


def _iter_print_profiles_upload_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    ext = os.path.splitext(filename)[1].lower()

    def _normalize_header_key(value):
        return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())

    def _build_header_index(headers):
        index = {}
        for idx, header in enumerate(headers or []):
            key = _normalize_header_key(header)
            if key and key not in index:
                index[key] = idx
        return index

    header_aliases = {
        "branch_name": ["branch", "branchname"],
        "company_name": ["company", "companyname"],
        "address_line1": ["addressline1", "address1"],
        "address_line2": ["addressline2", "address2"],
        "gst_no": ["gst", "gstno", "gstin"],
        "mobile1": ["mobile1", "phone1"],
        "mobile2": ["mobile2", "phone2"],
        "mobile3": ["mobile3", "phone3"],
        "terms_text": ["terms", "termstext", "termsconditions", "termsandconditions"],
        "quotation_terms": [
            "quotationterms",
            "quotationtermstext",
            "quotationtermsconditions",
            "quotationtermsandconditions",
            "quotationtc",
        ],
    }

    def _read_field(row, header_index, field_name):
        for alias in header_aliases[field_name]:
            idx = header_index.get(alias)
            if idx is not None and len(row) > idx:
                return str(row[idx] or "").strip()
        return ""

    def _iter_from_rows(rows):
        headers = next(rows, None)
        if not headers:
            raise ValueError("The file is empty")

        header_index = _build_header_index(headers)
        if not any(alias in header_index for alias in header_aliases["branch_name"]):
            raise ValueError("Missing required column: branch")

        for row_number, row in enumerate(rows, start=2):
            row = row or ()
            branch_name = _read_field(row, header_index, "branch_name")
            company_name = _read_field(row, header_index, "company_name")
            address_line1 = _read_field(row, header_index, "address_line1")
            address_line2 = _read_field(row, header_index, "address_line2")
            gst_no = _read_field(row, header_index, "gst_no")
            mobile1 = _read_field(row, header_index, "mobile1")
            mobile2 = _read_field(row, header_index, "mobile2")
            mobile3 = _read_field(row, header_index, "mobile3")
            terms_text = _read_field(row, header_index, "terms_text")
            quotation_terms = _read_field(row, header_index, "quotation_terms")

            yield (
                row_number,
                branch_name,
                company_name,
                address_line1,
                address_line2,
                gst_no,
                mobile1,
                mobile2,
                mobile3,
                terms_text,
                quotation_terms,
            )

    if ext == ".csv":
        raw_bytes = uploaded_file.stream.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        yield from _iter_from_rows(reader)
        return

    if ext == ".xlsx":
        if load_workbook is None:
            raise ValueError("Excel needs openpyxl. Use CSV or install openpyxl.")

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            yield from _iter_from_rows(rows)
        finally:
            workbook.close()
        return

    raise ValueError("Please upload .xlsx or .csv")


def build_dashboard_filters(args, role, session_branch):
    """
    Parse request args and build a WHERE clause that covers:
    branch scope, date range, job card status, closure status, engineer.
    Returns a dict with where_sql, params tuple, and all filter values.
    """
    from_date = _normalize_date_input(args.get("from_date", ""))
    to_date = _normalize_date_input(args.get("to_date", ""))

    date_field_values = _get_multi_values(args, "date_field")
    has_date_all = any(v.lower() == "all" for v in date_field_values)
    date_fields = [v for v in date_field_values if v in ["created", "closed"]]
    if has_date_all:
        date_fields = ["created", "closed"]
    if not date_fields:
        date_fields = ["created"]

    branch_values = _get_multi_values(args, "filter_branch")
    has_branch_all = any(v.upper() == "ALL" for v in branch_values)
    filter_branches = [v for v in branch_values if v.upper() != "ALL"]

    status_values = _get_multi_values(args, "filter_status")
    has_status_all = any(v.upper() == "ALL" for v in status_values)
    filter_statuses = [v for v in status_values if v.upper() != "ALL"]

    closure_values = _get_multi_values(args, "filter_closure")
    has_closure_all = any(v.upper() == "ALL" for v in closure_values)
    filter_closures = [v for v in closure_values if v.upper() != "ALL"]

    engineer_values = _get_multi_values(args, "filter_engineer")
    has_engineer_all = any(v.upper() == "ALL" for v in engineer_values)
    filter_engineers = [v for v in engineer_values if v.upper() != "ALL"]

    conditions = []
    params = []

    # Branch scope
    if role in ["super_admin", "admin"] and session_branch == "ALL":
        if filter_branches and not has_branch_all:
            branch_parts = []
            for filter_branch in filter_branches:
                branch_sql, branch_params = _build_job_transfer_branch_scope_clause(filter_branch)
                branch_parts.append(branch_sql)
                params.extend(branch_params)
            conditions.append("(" + " OR ".join(branch_parts) + ")")
        # no restriction when filter_branch is blank
    else:
        branch_sql, branch_params = _build_job_transfer_branch_scope_clause(session_branch)
        conditions.append(branch_sql)
        params.extend(branch_params)

    # Date range
    if from_date:
        from_ts = from_date + " 00:00:00"
        if len(date_fields) == 1:
            date_col = "closure_date" if date_fields[0] == "closed" else "created_at"
            conditions.append(f"{date_col} >= %s")
            params.append(from_ts)
        else:
            conditions.append("((created_at >= %s) OR (closure_date >= %s))")
            params.extend([from_ts, from_ts])

    if to_date:
        to_ts = to_date + " 23:59:59"
        if len(date_fields) == 1:
            date_col = "closure_date" if date_fields[0] == "closed" else "created_at"
            conditions.append(f"{date_col} <= %s")
            params.append(to_ts)
        else:
            conditions.append("((created_at <= %s) OR (closure_date <= %s))")
            params.extend([to_ts, to_ts])

    # Job card status filter
    if filter_statuses and not has_status_all:
        closed_selected = "Closed" in filter_statuses
        normal_statuses = [s for s in filter_statuses if s != "Closed"]
        status_parts = []

        if normal_statuses:
            placeholders = ", ".join(["%s"] * len(normal_statuses))
            status_parts.append(f"status IN ({placeholders})")
            params.extend(normal_statuses)

        if closed_selected:
            status_parts.append("((closure_status IS NOT NULL AND closure_status <> '') OR status='Closed')")

        if status_parts:
            conditions.append("(" + " OR ".join(status_parts) + ")")

    # Closure status filter
    if filter_closures and not has_closure_all:
        placeholders = ", ".join(["%s"] * len(filter_closures))
        conditions.append(f"closure_status IN ({placeholders})")
        params.extend(filter_closures)

    # Engineer filter
    if filter_engineers and not has_engineer_all:
        engineer_sql, engineer_params = _build_job_transfer_engineer_filter_clause(filter_engineers)
        conditions.append(engineer_sql)
        params.extend(engineer_params)

    where_sql = " AND ".join(conditions) if conditions else "1=1"

    return {
        "where_sql": where_sql,
        "params": tuple(params),
        "from_date": from_date,
        "to_date": to_date,
        "date_field_values": date_field_values,
        "date_fields": date_fields,
        "filter_branch_values": branch_values,
        "filter_branches": filter_branches,
        "filter_status_values": status_values,
        "filter_statuses": filter_statuses,
        "filter_closure_values": closure_values,
        "filter_closures": filter_closures,
        "filter_engineer_values": engineer_values,
        "filter_engineers": filter_engineers,
        # Backward-compatible single-value aliases
        "date_field": date_fields[0] if date_fields else "created",
        "filter_branch": filter_branches[0] if filter_branches else "",
        "filter_status": filter_statuses[0] if filter_statuses else "",
        "filter_closure": filter_closures[0] if filter_closures else "",
        "filter_engineer": filter_engineers[0] if filter_engineers else "",
    }


# Branch list fallback (matches the login branch datalist)
DEFAULT_BRANCHES = [
    "ALL",
    "ALAPPUZHA",
    "CALICUT",
    "CHENGANNUR",
    "DREAMS MALL KOTTIYAM",
    "FORUM MALL COCHIN",
    "KADAVANTHRA",
    "KANHANGAD",
    "KANNUR",
    "KOLLAM",
    "KOTTAKKAL",
    "KOTTAYAM",
    "LULU COCHIN",
    "LULU TVM",
    "MANJERI",
    "MOT",
    "NILAMBUR",
    "OBERON",
    "PALA",
    "PATHANAMTHITTA",
    "PERINTHALMANNA",
    "PERUMBAVOOR",
    "RAVIPURAM",
    "SOBHA CITY",
    "THALASSERY",
    "THODUPUZHA",
    "THRISSUR",
    "TRIPUNITHURA",
    "Y MALL",
    "CHENNAI",
    "COIMBATORE",
]


class _PooledConnectionHandle:
    def __init__(self, connection, release_callback):
        self._connection = connection
        self._release_callback = release_callback
        self._closed = False

    def close(self):
        if self._closed:
            return
        try:
            self._connection.close()
        finally:
            self._closed = True
            self._release_callback()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()
        return False

    def __getattr__(self, name):
        return getattr(self._connection, name)


def _ensure_db_pool():
    global _db_pool, _db_pool_semaphore
    if _db_pool is not None and _db_pool_semaphore is not None:
        return

    with _db_pool_lock:
        if _db_pool is None:
            _db_pool = pooling.MySQLConnectionPool(
                pool_name=DB_POOL_NAME,
                pool_size=DB_POOL_SIZE,
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASSWORD,
                database=DB_NAME,
                autocommit=False,
                raise_on_warnings=True,
            )
        if _db_pool_semaphore is None:
            _db_pool_semaphore = BoundedSemaphore(DB_POOL_SIZE)


def _configure_db_connection_session(connection):
    cursor = None
    try:
        cursor = connection.cursor()
        cursor.execute("SET time_zone = %s", (mysql_session_timezone_value(),))
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass


def get_db():
    _ensure_db_pool()

    acquired = _db_pool_semaphore.acquire(timeout=DB_POOL_ACQUIRE_TIMEOUT_SECONDS)
    if not acquired:
        raise PoolError(
            f"Timed out waiting for DB connection after {DB_POOL_ACQUIRE_TIMEOUT_SECONDS:.1f} seconds"
        )

    try:
        connection = _db_pool.get_connection()
    except Exception:
        _db_pool_semaphore.release()
        raise

    try:
        _configure_db_connection_session(connection)
    except Exception:
        try:
            connection.close()
        finally:
            _db_pool_semaphore.release()
        raise

    return _PooledConnectionHandle(connection, _db_pool_semaphore.release)


def _safe_close(*resources):
    for resource in resources:
        if not resource:
            continue
        try:
            resource.close()
        except Exception:
            pass


def ensure_performance_indexes():
    """Create essential indexes used by dashboard and report filters."""
    db = get_db()
    cursor = db.cursor()

    try:
        index_plan = [
            ("jobs", "idx_jobs_branch_name", "CREATE INDEX idx_jobs_branch_name ON jobs (branch_name)"),
            ("jobs", "idx_jobs_status", "CREATE INDEX idx_jobs_status ON jobs (status)"),
            ("jobs", "idx_jobs_closure_status", "CREATE INDEX idx_jobs_closure_status ON jobs (closure_status)"),
            ("jobs", "idx_jobs_assigned_engineer", "CREATE INDEX idx_jobs_assigned_engineer ON jobs (assigned_engineer)"),
            ("jobs", "idx_jobs_created_at", "CREATE INDEX idx_jobs_created_at ON jobs (created_at)"),
            ("jobs", "idx_jobs_closure_date", "CREATE INDEX idx_jobs_closure_date ON jobs (closure_date)"),
            ("jobs", "idx_jobs_branch_created", "CREATE INDEX idx_jobs_branch_created ON jobs (branch_name, created_at)"),
            ("jobs", "idx_jobs_branch_closure", "CREATE INDEX idx_jobs_branch_closure ON jobs (branch_name, closure_date)"),
            ("jobs", "idx_jobs_scope_active", "CREATE INDEX idx_jobs_scope_active ON jobs (branch_name, closure_status, status)"),
            (
                "branch_revenue_entries",
                "idx_revenue_entry_branch_date",
                "CREATE INDEX idx_revenue_entry_branch_date ON branch_revenue_entries (branch_name, entry_date)",
            ),
            (
                "syscare_memberships",
                "idx_syscare_branch_record_date",
                "CREATE INDEX idx_syscare_branch_record_date ON syscare_memberships (branch_name, record_date)",
            ),
        ]

        for table_name, index_name, create_sql in index_plan:
            cursor.execute(
                """
                SELECT COUNT(*)
                FROM information_schema.statistics
                WHERE table_schema=%s AND table_name=%s AND index_name=%s
                """,
                (DB_NAME, table_name, index_name),
            )
            exists = cursor.fetchone()[0] > 0
            if not exists:
                try:
                    cursor.execute(create_sql)
                except Error:
                    # Skip if table/column is not present in an older DB.
                    pass

        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_created_at_column():
    """Ensure jobs.created_at exists and is populated."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            "SELECT COUNT(*) FROM information_schema.columns WHERE table_schema=%s AND table_name='jobs' AND column_name='created_at'",
            (DB_NAME,)
        )
        has_column = cursor.fetchone()[0] > 0

        if not has_column:
            cursor.execute(
                "ALTER TABLE jobs ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP"
            )
            cursor.execute(
                "UPDATE jobs SET created_at = NOW() WHERE created_at IS NULL"
            )
            db.commit()

    except Error:
        pass

    finally:
        _safe_close(cursor, db)


def ensure_closure_columns():
    """Ensure closure fields exist in jobs table."""
    db = get_db()
    cursor = db.cursor()

    try:
        columns = {
            "closure_status": "VARCHAR(255) NULL",
            "closure_service_type": "VARCHAR(32) NULL",
            "closure_notes": "TEXT NULL",
            "closure_date": "DATETIME NULL",
            "closed_by": "VARCHAR(255) NULL",
        }

        for col_name, col_def in columns.items():
            cursor.execute(
                "SELECT COUNT(*) FROM information_schema.columns "
                "WHERE table_schema=%s AND table_name='jobs' AND column_name=%s",
                (DB_NAME, col_name),
            )
            has_column = cursor.fetchone()[0] > 0

            if not has_column:
                cursor.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {col_def}")

        db.commit()

    except Error:
        pass

    finally:
        _safe_close(cursor, db)


def ensure_job_billing_columns():
    """Ensure billing and payment columns exist in jobs table."""
    db = get_db()
    cursor = db.cursor()

    try:
        columns = {
            "service_charges": "DECIMAL(12,2) NOT NULL DEFAULT 0",
            "payment_cash": "DECIMAL(12,2) NOT NULL DEFAULT 0",
            "payment_upi": "DECIMAL(12,2) NOT NULL DEFAULT 0",
            "payment_card": "DECIMAL(12,2) NOT NULL DEFAULT 0",
            "spares_billing_status": "VARCHAR(32) NOT NULL DEFAULT 'Not Required'",
            "spares_invoice_no": "VARCHAR(100) NULL",
            "spares_invoice_date": "DATE NULL",
            "spares_billed_by": "VARCHAR(255) NULL",
            "spares_billing_notes": "TEXT NULL",
        }

        for col_name, col_def in columns.items():
            cursor.execute(
                "SELECT COUNT(*) FROM information_schema.columns "
                "WHERE table_schema=%s AND table_name='jobs' AND column_name=%s",
                (DB_NAME, col_name),
            )
            has_column = cursor.fetchone()[0] > 0
            if not has_column:
                cursor.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {col_def}")

        cursor.execute(
            """
            UPDATE jobs j
            LEFT JOIN (
                SELECT job_id, COUNT(*) AS spare_count
                FROM used_spares
                GROUP BY job_id
            ) us ON us.job_id = j.id
            SET j.spares_billing_status = CASE
                WHEN COALESCE(us.spare_count, 0) = 0 THEN 'Not Required'
                WHEN j.spares_invoice_no IS NOT NULL AND TRIM(j.spares_invoice_no) <> '' THEN 'Billed'
                ELSE 'Pending'
            END
            WHERE j.spares_billing_status IS NULL
               OR TRIM(j.spares_billing_status) = ''
               OR TRIM(j.spares_billing_status) NOT IN ('Pending', 'Billed', 'Not Required')
            """
        )

        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_profile_picture_column():
    """Ensure profile_picture column exists in users table."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            "SELECT COUNT(*) FROM information_schema.columns "
            "WHERE table_schema=%s AND table_name='users' AND column_name='profile_picture'",
            (DB_NAME,)
        )
        has_column = cursor.fetchone()[0] > 0

        if not has_column:
            cursor.execute(
                "ALTER TABLE users ADD COLUMN profile_picture VARCHAR(500) NULL"
            )
            db.commit()

    except Error:
        pass

    finally:
        _safe_close(cursor, db)


def ensure_password_hash_column_capacity():
    """Ensure users.password can store modern password hashes."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            SELECT data_type, character_maximum_length, is_nullable
            FROM information_schema.columns
            WHERE table_schema=%s AND table_name='users' AND column_name='password'
            """,
            (DB_NAME,)
        )
        column_info = cursor.fetchone() or ()
        data_type = str(column_info[0] if len(column_info) > 0 else "").lower()
        max_length = int(column_info[1] if len(column_info) > 1 and column_info[1] is not None else 0)
        is_nullable = str(column_info[2] if len(column_info) > 2 else "YES").upper() == "YES"

        if data_type in {"varchar", "char"} and max_length < 255:
            nullable_sql = "NULL" if is_nullable else "NOT NULL"
            cursor.execute(f"ALTER TABLE users MODIFY COLUMN password VARCHAR(255) {nullable_sql}")
            db.commit()

    except Error:
        pass

    finally:
        _safe_close(cursor, db)


def ensure_status_update_columns():
    """Ensure status update audit fields exist in jobs table."""
    db = get_db()
    cursor = db.cursor()

    try:
        columns = {
            "status_update_notes": "TEXT NULL",
            "status_updated_by": "VARCHAR(255) NULL",
            "status_updated_at": "DATETIME NULL",
        }

        for col_name, col_def in columns.items():
            cursor.execute(
                "SELECT COUNT(*) FROM information_schema.columns "
                "WHERE table_schema=%s AND table_name='jobs' AND column_name=%s",
                (DB_NAME, col_name),
            )
            has_column = cursor.fetchone()[0] > 0

            if not has_column:
                cursor.execute(f"ALTER TABLE jobs ADD COLUMN {col_name} {col_def}")

        db.commit()

    except Error:
        pass

    finally:
        _safe_close(cursor, db)


def ensure_used_spares_table():
    """Ensure used_spares table exists for multiple spare rows per job."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS used_spares (
                id INT AUTO_INCREMENT PRIMARY KEY,
                job_id INT NOT NULL,
                spare_name VARCHAR(255) NOT NULL,
                amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_used_spares_job (job_id),
                CONSTRAINT fk_used_spares_job FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_job_attachments_table():
    """Ensure job_attachments table exists for storing multiple images per job."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS job_attachments (
                id INT AUTO_INCREMENT PRIMARY KEY,
                job_id INT NOT NULL,
                filename VARCHAR(255) NOT NULL,
                uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_job_attachments_job_id (job_id),
                CONSTRAINT fk_job_attachments_job FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_branch_print_profiles_table():
    """Ensure table for branch-wise customer print header settings exists."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS branch_print_profiles (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch_name VARCHAR(255) NOT NULL UNIQUE,
                company_name VARCHAR(255) NOT NULL DEFAULT 'SYSMANTECH',
                address_line1 VARCHAR(255) NULL,
                address_line2 VARCHAR(255) NULL,
                gst_no VARCHAR(100) NULL,
                mobile1 VARCHAR(50) NULL,
                mobile2 VARCHAR(50) NULL,
                mobile3 VARCHAR(50) NULL,
                terms_text TEXT NULL,
                quotation_terms TEXT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )

        cursor.execute(
            "SELECT COUNT(*) FROM information_schema.columns "
            "WHERE table_schema=%s AND table_name='branch_print_profiles' AND column_name='quotation_terms'",
            (DB_NAME,),
        )
        has_quotation_terms = cursor.fetchone()[0] > 0
        if not has_quotation_terms:
            cursor.execute("ALTER TABLE branch_print_profiles ADD COLUMN quotation_terms TEXT NULL")

        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_branch_revenue_targets_table():
    """Ensure table for branch-wise combined revenue target exists."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS branch_revenue_targets (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch_name VARCHAR(255) NOT NULL UNIQUE,
                total_target DECIMAL(12,2) NOT NULL DEFAULT 0,
                sales_target DECIMAL(12,2) NOT NULL DEFAULT 0,
                service_target DECIMAL(12,2) NOT NULL DEFAULT 0,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        db.commit()
    except Error:
        pass

    # Always try to add total_target column in case table was created without it
    try:
        cursor.execute(
            "ALTER TABLE branch_revenue_targets ADD COLUMN total_target DECIMAL(12,2) NOT NULL DEFAULT 0"
        )
        cursor.execute(
            "UPDATE branch_revenue_targets SET total_target = COALESCE(sales_target, 0) + COALESCE(service_target, 0) "
            "WHERE total_target IS NULL OR total_target = 0"
        )
        db.commit()
    except Error:
        pass  # Column already exists — that's fine

    cursor.close()
    db.close()


def ensure_branch_revenue_entries_table():
    """Ensure date-wise manual revenue entries table exists."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS branch_revenue_entries (
                id INT AUTO_INCREMENT PRIMARY KEY,
                entry_date DATE NOT NULL,
                branch_name VARCHAR(255) NOT NULL,
                sales_profit DECIMAL(12,2) NOT NULL DEFAULT 0,
                service_charges DECIMAL(12,2) NOT NULL DEFAULT 0,
                total_profit DECIMAL(12,2) NOT NULL DEFAULT 0,
                zone VARCHAR(100) NULL,
                created_by VARCHAR(255) NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_revenue_entry_date_branch (entry_date, branch_name),
                INDEX idx_revenue_entry_date (entry_date),
                INDEX idx_revenue_branch (branch_name)
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_revenue_entry_period_locks_table():
    """Ensure table for locking revenue entry periods (month-wise)."""
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS revenue_entry_period_locks (
                id INT AUTO_INCREMENT PRIMARY KEY,
                from_date DATE NOT NULL,
                to_date DATE NOT NULL,
                locked_by VARCHAR(255) NULL,
                locked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_rev_entry_lock (from_date, to_date)
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_branch_cashflow_entries_table():
    """Ensure daily branch cashflow entry table exists."""
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS branch_cashflow_entries (
                id INT AUTO_INCREMENT PRIMARY KEY,
                entry_date DATE NOT NULL,
                branch_name VARCHAR(255) NOT NULL,
                cash_amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                card_amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                upi_amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                total_amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                remarks VARCHAR(255) NULL,
                created_by VARCHAR(255) NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_cashflow_entry_date_branch (entry_date, branch_name),
                INDEX idx_cashflow_entry_date (entry_date),
                INDEX idx_cashflow_entry_branch (branch_name)
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_branch_cash_transfer_requests_table():
    """Ensure branch cash transfer request table exists."""
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS branch_cash_transfer_requests (
                id INT AUTO_INCREMENT PRIMARY KEY,
                branch_name VARCHAR(255) NOT NULL,
                request_date DATE NOT NULL,
                amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                transfer_to VARCHAR(255) NULL,
                requested_notes TEXT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'Pending',
                requested_by VARCHAR(255) NULL,
                requested_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                reviewed_by VARCHAR(255) NULL,
                reviewed_at DATETIME NULL,
                review_notes TEXT NULL,
                transfer_reference VARCHAR(120) NULL,
                INDEX idx_cash_transfer_branch_status (branch_name, status),
                INDEX idx_cash_transfer_request_date (request_date)
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_job_service_transfers_table():
    """Ensure specialist branch transfer table exists for main service jobs."""
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS job_service_transfers (
                id INT AUTO_INCREMENT PRIMARY KEY,
                job_id INT NOT NULL,
                from_branch_name VARCHAR(255) NOT NULL,
                to_branch_name VARCHAR(255) NOT NULL,
                specialist_engineer VARCHAR(255) NULL,
                service_type VARCHAR(255) NULL,
                request_notes TEXT NULL,
                status_notes TEXT NULL,
                internal_service_charge DECIMAL(12,2) NOT NULL DEFAULT 0,
                status VARCHAR(50) NOT NULL DEFAULT 'Sent',
                sent_by VARCHAR(255) NULL,
                updated_by VARCHAR(255) NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                accepted_at DATETIME NULL,
                completed_at DATETIME NULL,
                returned_at DATETIME NULL,
                INDEX idx_job_service_transfers_job (job_id),
                INDEX idx_job_service_transfers_target_status (to_branch_name, status),
                INDEX idx_job_service_transfers_engineer_status (specialist_engineer, status),
                CONSTRAINT fk_job_service_transfers_job FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_quotations_tables():
    """Ensure standalone quotations master and items tables exist."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS quotations (
                id INT AUTO_INCREMENT PRIMARY KEY,
                quote_number VARCHAR(100) NOT NULL UNIQUE,
                quote_date DATE NOT NULL,
                branch_name VARCHAR(255) NOT NULL,
                customer_name VARCHAR(255) NOT NULL,
                customer_mobile VARCHAR(50) NULL,
                customer_address TEXT NULL,
                customer_gst_no VARCHAR(100) NULL,
                engineer_name VARCHAR(255) NULL,
                engineer_mobile VARCHAR(50) NULL,
                terms_text TEXT NULL,
                grand_total DECIMAL(12,2) NOT NULL DEFAULT 0,
                created_by VARCHAR(255) NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_quote_date (quote_date),
                INDEX idx_quote_branch (branch_name)
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS quotation_items (
                id INT AUTO_INCREMENT PRIMARY KEY,
                quotation_id INT NOT NULL,
                line_no INT NOT NULL DEFAULT 1,
                item_name VARCHAR(255) NOT NULL,
                narration TEXT NULL,
                qty DECIMAL(12,2) NOT NULL DEFAULT 0,
                amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                final_amount DECIMAL(12,2) NOT NULL DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_quotation_items_quote (quotation_id),
                CONSTRAINT fk_quotation_items_quote FOREIGN KEY (quotation_id) REFERENCES quotations(id) ON DELETE CASCADE
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_sequence_counters_table():
    """Ensure atomic counters table exists for concurrent document number allocation."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS sequence_counters (
                sequence_key VARCHAR(100) PRIMARY KEY,
                `last_value` BIGINT NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def ensure_staff_directory_table():
    """Ensure staff management table exists for super admin payroll tracking."""
    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS staff_directory (
                id INT AUTO_INCREMENT PRIMARY KEY,
                staff_name VARCHAR(255) NOT NULL,
                contact_number VARCHAR(50) NOT NULL,
                branch_name VARCHAR(255) NOT NULL,
                salary DECIMAL(12,2) NOT NULL DEFAULT 0,
                esi DECIMAL(12,2) NOT NULL DEFAULT 0,
                pf DECIMAL(12,2) NOT NULL DEFAULT 0,
                room DECIMAL(12,2) NOT NULL DEFAULT 0,
                rent DECIMAL(12,2) NOT NULL DEFAULT 0,
                joined_date DATE NOT NULL,
                resigned_date DATE NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_staff_directory_branch (branch_name),
                INDEX idx_staff_directory_joined_date (joined_date),
                INDEX idx_staff_directory_name (staff_name)
            )
            """
        )

        staff_columns = {
            "staff_name": "ADD COLUMN staff_name VARCHAR(255) NOT NULL DEFAULT '' AFTER id",
            "contact_number": "ADD COLUMN contact_number VARCHAR(50) NOT NULL DEFAULT '' AFTER staff_name",
            "branch_name": "ADD COLUMN branch_name VARCHAR(255) NOT NULL DEFAULT '' AFTER contact_number",
            "salary": "ADD COLUMN salary DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER branch_name",
            "esi": "ADD COLUMN esi DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER salary",
            "pf": "ADD COLUMN pf DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER esi",
            "room": "ADD COLUMN room DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER pf",
            "rent": "ADD COLUMN rent DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER room",
            "joined_date": "ADD COLUMN joined_date DATE NULL AFTER rent",
            "resigned_date": "ADD COLUMN resigned_date DATE NULL AFTER joined_date",
            "created_at": "ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP AFTER resigned_date",
            "updated_at": "ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP AFTER created_at",
        }
        for column_name, alter_clause in staff_columns.items():
            cursor.execute("SHOW COLUMNS FROM staff_directory LIKE %s", (column_name,))
            if not cursor.fetchone():
                cursor.execute(f"ALTER TABLE staff_directory {alter_clause}")

        staff_column_modifications = {
            "staff_name": "MODIFY COLUMN staff_name VARCHAR(255) NOT NULL DEFAULT ''",
            "contact_number": "MODIFY COLUMN contact_number VARCHAR(50) NOT NULL DEFAULT ''",
            "branch_name": "MODIFY COLUMN branch_name VARCHAR(255) NOT NULL DEFAULT ''",
            "salary": "MODIFY COLUMN salary DECIMAL(12,2) NOT NULL DEFAULT 0",
            "esi": "MODIFY COLUMN esi DECIMAL(12,2) NOT NULL DEFAULT 0",
            "pf": "MODIFY COLUMN pf DECIMAL(12,2) NOT NULL DEFAULT 0",
            "room": "MODIFY COLUMN room DECIMAL(12,2) NOT NULL DEFAULT 0",
            "rent": "MODIFY COLUMN rent DECIMAL(12,2) NOT NULL DEFAULT 0",
            "joined_date": "MODIFY COLUMN joined_date DATE NULL",
            "resigned_date": "MODIFY COLUMN resigned_date DATE NULL",
        }
        for alter_clause in staff_column_modifications.values():
            try:
                cursor.execute(f"ALTER TABLE staff_directory {alter_clause}")
            except Error:
                continue

        db.commit()
    except Error:
        pass
    finally:
        _safe_close(cursor, db)


def get_branch_print_profile(cursor, branch_name):
    """Get branch print profile; fallback to ALL profile; fallback to defaults."""
    default_terms = (
        "1. Goods once sold/service delivered are not returnable.\n"
        "2. Service warranty as per invoice/estimate approval terms.\n"
        "3. Please keep this job copy for future reference."
    )
    default_quotation_terms = (
        "1. Quotation is valid for 7 days from the quote date.\n"
        "2. Delivery timeline starts after order confirmation and advance payment (if applicable).\n"
        "3. Taxes and transport charges are extra unless explicitly stated."
    )

    profile = None
    if branch_name:
        cursor.execute("SELECT * FROM branch_print_profiles WHERE branch_name=%s", (branch_name,))
        profile = cursor.fetchone()

    if not profile:
        cursor.execute("SELECT * FROM branch_print_profiles WHERE branch_name=%s", ("ALL",))
        profile = cursor.fetchone()

    if not profile:
        profile = {
            "branch_name": branch_name or "ALL",
            "company_name": "SYSMANTECH",
            "address_line1": "",
            "address_line2": "",
            "gst_no": "",
            "mobile1": "",
            "mobile2": "",
            "mobile3": "",
            "terms_text": default_terms,
            "quotation_terms": default_quotation_terms,
        }
    elif not profile.get("terms_text"):
        profile["terms_text"] = default_terms

    if not profile.get("quotation_terms"):
        profile["quotation_terms"] = default_quotation_terms

    return profile


def _extract_branch_print_profile_payload(form):
    branch_name = re.sub(r"\s+", " ", str(form.get("branch_name") or "").strip()).upper()
    company_name = re.sub(r"\s+", " ", str(form.get("company_name") or "").strip()) or "SYSMANTECH"
    return {
        "branch_name": branch_name,
        "company_name": company_name,
        "address_line1": str(form.get("address_line1") or "").strip(),
        "address_line2": str(form.get("address_line2") or "").strip(),
        "gst_no": str(form.get("gst_no") or "").strip(),
        "mobile1": str(form.get("mobile1") or "").strip(),
        "mobile2": str(form.get("mobile2") or "").strip(),
        "mobile3": str(form.get("mobile3") or "").strip(),
        "terms_text": str(form.get("terms_text") or "").strip(),
        "quotation_terms": str(form.get("quotation_terms") or "").strip(),
    }


def get_branch_revenue_target(cursor, branch_name):
    """Get branch revenue target; fallback to ALL target; fallback to 0 values."""
    target = None

    if branch_name:
        cursor.execute("SELECT * FROM branch_revenue_targets WHERE branch_name=%s", (branch_name,))
        target = cursor.fetchone()

    if not target:
        cursor.execute("SELECT * FROM branch_revenue_targets WHERE branch_name=%s", ("ALL",))
        target = cursor.fetchone()

    if not target:
        target = {
            "branch_name": branch_name or "ALL",
            "total_target": 0,
        }

    if "total_target" not in target or target.get("total_target") is None:
        target["total_target"] = float(target.get("sales_target") or 0) + float(target.get("service_target") or 0)

    return target


def _get_effective_revenue_target_amount(target_row):
    if not target_row:
        return 0.0

    total_target = target_row.get("total_target")
    if total_target is None:
        total_target = float(target_row.get("sales_target") or 0) + float(target_row.get("service_target") or 0)

    return round(float(total_target or 0), 2)


def _get_revenue_target_for_scope(cursor, filter_branch=""):
    scope_branch = str(filter_branch or "").strip()

    if scope_branch and scope_branch.upper() != "ALL":
        target_row = get_branch_revenue_target(cursor, scope_branch)
        target_value = _get_effective_revenue_target_amount(target_row)
        target_label = str(target_row.get("branch_name") or scope_branch or "Target").strip() or "Target"
        return target_value, target_label

    all_target_row = get_branch_revenue_target(cursor, "ALL")
    all_target_value = _get_effective_revenue_target_amount(all_target_row)
    if all_target_value > 0:
        return all_target_value, "All Branches"

    cursor.execute("SELECT * FROM branch_revenue_targets WHERE UPPER(branch_name) <> %s", ("ALL",))
    target_rows = cursor.fetchall()
    total_target_value = round(sum(_get_effective_revenue_target_amount(row) for row in target_rows), 2)
    return total_target_value, "All Branches"


def _parse_money(value):
    text = str(value or "").strip()
    if not text:
        return 0.0
    text = text.replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in [".", "-", "-."]:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_cash_transfer_role(role):
    return (role or "").strip() in {"super_admin", "admin", "cash_admin", "coordinator", "user"}


def _can_request_cash_transfer(role):
    return _is_cash_transfer_role(role)


def _can_approve_cash_transfer(role):
    return _is_cash_transfer_role(role)


def _can_cancel_cash_transfer(role):
    return (role or "").strip() in {"super_admin", "admin"}


def _merge_cashflow_amount_maps(*maps):
    merged = {}
    for amount_map in maps:
        for branch_name, amount in (amount_map or {}).items():
            normalized_branch = str(branch_name or "").strip()
            if not normalized_branch:
                continue
            merged[normalized_branch] = round(float(merged.get(normalized_branch) or 0) + float(amount or 0), 2)
    return merged


def _merge_cashflow_summary_rows(*row_groups):
    merged = {}
    for row_group in row_groups:
        for row in (row_group or []):
            branch_name = str(row.get("branch_name") or "").strip()
            if not branch_name:
                continue
            bucket = merged.setdefault(
                branch_name,
                {
                    "branch_name": branch_name,
                    "cash_total": 0.0,
                    "card_total": 0.0,
                    "upi_total": 0.0,
                    "total_collected": 0.0,
                },
            )
            bucket["cash_total"] = round(float(bucket.get("cash_total") or 0) + float(row.get("cash_total") or 0), 2)
            bucket["card_total"] = round(float(bucket.get("card_total") or 0) + float(row.get("card_total") or 0), 2)
            bucket["upi_total"] = round(float(bucket.get("upi_total") or 0) + float(row.get("upi_total") or 0), 2)
            bucket["total_collected"] = round(float(bucket.get("total_collected") or 0) + float(row.get("total_collected") or 0), 2)
    return [merged[branch_name] for branch_name in sorted(merged)]


def _fetch_closed_job_cashflow_summary_rows(cursor, filter_branch="", from_date="", to_date=""):
    where = [
        "status=%s",
        "closure_date IS NOT NULL",
        "(COALESCE(payment_cash, 0) <> 0 OR COALESCE(payment_card, 0) <> 0 OR COALESCE(payment_upi, 0) <> 0)",
    ]
    params = ["Closed"]
    normalized_branch = (filter_branch or "").strip()

    if from_date:
        where.append("closure_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("closure_date < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(to_date)
    if normalized_branch and normalized_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(normalized_branch)

    cursor.execute(
        f"""
        SELECT
            branch_name,
            COALESCE(SUM(payment_cash), 0) AS cash_total,
            COALESCE(SUM(payment_card), 0) AS card_total,
            COALESCE(SUM(payment_upi), 0) AS upi_total,
            COALESCE(SUM(COALESCE(payment_cash, 0) + COALESCE(payment_card, 0) + COALESCE(payment_upi, 0)), 0) AS total_collected
        FROM jobs
        WHERE {' AND '.join(where)}
        GROUP BY branch_name
        ORDER BY branch_name ASC
        """,
        tuple(params),
    )
    return cursor.fetchall()


def _fetch_closed_job_cash_totals_by_branch(cursor, filter_branch="", to_date=""):
    return {
        str(row.get("branch_name") or "").strip(): float(row.get("cash_total") or 0)
        for row in _fetch_closed_job_cashflow_summary_rows(cursor, filter_branch=filter_branch, to_date=to_date)
        if str(row.get("branch_name") or "").strip()
    }


def _fetch_closed_job_cashflow_rows(cursor, filter_branch="", from_date="", to_date=""):
    where = [
        "status=%s",
        "closure_date IS NOT NULL",
        "(COALESCE(payment_cash, 0) <> 0 OR COALESCE(payment_card, 0) <> 0 OR COALESCE(payment_upi, 0) <> 0)",
    ]
    params = ["Closed"]
    normalized_branch = (filter_branch or "").strip()

    if from_date:
        where.append("closure_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("closure_date < DATE_ADD(%s, INTERVAL 1 DAY)")
        params.append(to_date)
    if normalized_branch and normalized_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(normalized_branch)

    cursor.execute(
        f"""
        SELECT
            id,
            DATE(closure_date) AS entry_date,
            branch_name,
            COALESCE(payment_cash, 0) AS cash_amount,
            COALESCE(payment_card, 0) AS card_amount,
            COALESCE(payment_upi, 0) AS upi_amount,
            COALESCE(payment_cash, 0) + COALESCE(payment_card, 0) + COALESCE(payment_upi, 0) AS total_amount,
            CONCAT('Auto from closed job #', COALESCE(CAST(job_number AS CHAR), CAST(id AS CHAR))) AS remarks,
            closed_by AS created_by,
            closure_date AS updated_at
        FROM jobs
        WHERE {' AND '.join(where)}
        ORDER BY closure_date DESC, branch_name ASC, id DESC
        """,
        tuple(params),
    )
    return cursor.fetchall()


def _fetch_cashflow_cash_totals_by_branch(cursor, filter_branch="", to_date=""):
    return _fetch_closed_job_cash_totals_by_branch(cursor, filter_branch=filter_branch, to_date=to_date)


def _fetch_cashflow_transfer_totals_by_branch(cursor, filter_branch="", from_date="", to_date="", statuses=None):
    where = ["1=1"]
    params = []
    normalized_branch = (filter_branch or "").strip()
    normalized_statuses = [str(status).strip() for status in (statuses or []) if str(status).strip()]

    if from_date:
        where.append("request_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("request_date <= %s")
        params.append(to_date)
    if normalized_branch and normalized_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(normalized_branch)
    if normalized_statuses:
        where.append("status IN (" + ", ".join(["%s"] * len(normalized_statuses)) + ")")
        params.extend(normalized_statuses)

    cursor.execute(
        f"""
        SELECT branch_name, COALESCE(SUM(amount), 0) AS amount
        FROM branch_cash_transfer_requests
        WHERE {' AND '.join(where)}
        GROUP BY branch_name
        """,
        tuple(params),
    )
    return {
        str(row.get("branch_name") or "").strip(): float(row.get("amount") or 0)
        for row in cursor.fetchall()
        if str(row.get("branch_name") or "").strip()
    }


def _fetch_cashflow_transfer_request_count(cursor, filter_branch="", statuses=None):
    where = ["1=1"]
    params = []
    normalized_branch = (filter_branch or "").strip()
    normalized_statuses = [str(status).strip() for status in (statuses or []) if str(status).strip()]

    if normalized_branch and normalized_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(normalized_branch)
    if normalized_statuses:
        where.append("status IN (" + ", ".join(["%s"] * len(normalized_statuses)) + ")")
        params.extend(normalized_statuses)

    cursor.execute(
        f"""
        SELECT COUNT(*) AS request_count
        FROM branch_cash_transfer_requests
        WHERE {' AND '.join(where)}
        """,
        tuple(params),
    )
    row = cursor.fetchone() or {}
    return int(row.get("request_count") or 0)


def _get_branch_available_cash(cursor, branch_name, to_date=""):
    normalized_branch = (branch_name or "").strip()
    if not normalized_branch:
        return 0.0

    cash_totals = _fetch_cashflow_cash_totals_by_branch(cursor, normalized_branch, to_date)
    approved_transfers = _fetch_cashflow_transfer_totals_by_branch(
        cursor,
        normalized_branch,
        to_date=to_date,
        statuses=["Approved"],
    )
    return round(
        float(cash_totals.get(normalized_branch) or 0) - float(approved_transfers.get(normalized_branch) or 0),
        2,
    )


def _get_cashflow_dashboard_snapshot(cursor, filter_branch=""):
    today_str = business_now_naive().strftime("%Y-%m-%d")
    normalized_branch = (filter_branch or "").strip()
    branch_rows = _fetch_closed_job_cashflow_summary_rows(
        cursor,
        filter_branch=normalized_branch,
        from_date=today_str,
        to_date=today_str,
    )
    branch_rows.sort(key=lambda row: str(row.get("branch_name") or "").upper())
    branch_rows.sort(key=lambda row: float(row.get("total_collected") or 0), reverse=True)

    totals_row = {
        "cash_total": round(sum(float(row.get("cash_total") or 0) for row in branch_rows), 2),
        "card_total": round(sum(float(row.get("card_total") or 0) for row in branch_rows), 2),
        "upi_total": round(sum(float(row.get("upi_total") or 0) for row in branch_rows), 2),
        "total_collected": round(sum(float(row.get("total_collected") or 0) for row in branch_rows), 2),
    }

    cash_totals_map = _fetch_cashflow_cash_totals_by_branch(cursor, normalized_branch, today_str)
    approved_totals_map = _fetch_cashflow_transfer_totals_by_branch(
        cursor,
        normalized_branch,
        to_date=today_str,
        statuses=["Approved"],
    )
    pending_totals_map = _fetch_cashflow_transfer_totals_by_branch(
        cursor,
        normalized_branch,
        statuses=["Pending"],
    )
    for row in branch_rows:
        branch_name = str(row.get("branch_name") or "").strip()
        available_cash = float(cash_totals_map.get(branch_name) or 0) - float(approved_totals_map.get(branch_name) or 0)
        row["available_cash"] = round(available_cash, 2)
        row["pending_transfer"] = round(float(pending_totals_map.get(branch_name) or 0), 2)

    branch_rows = branch_rows[:6]

    available_cash_total = round(
        sum(float(cash_totals_map.get(branch_name) or 0) - float(approved_totals_map.get(branch_name) or 0) for branch_name in cash_totals_map),
        2,
    )
    pending_transfer_total = round(sum(float(amount or 0) for amount in pending_totals_map.values()), 2)
    pending_transfer_count = _fetch_cashflow_transfer_request_count(cursor, normalized_branch, statuses=["Pending"])

    return {
        "today_date": today_str,
        "cash_total": float(totals_row.get("cash_total") or 0),
        "card_total": float(totals_row.get("card_total") or 0),
        "upi_total": float(totals_row.get("upi_total") or 0),
        "total_collected": float(totals_row.get("total_collected") or 0),
        "available_cash": available_cash_total,
        "pending_transfers": pending_transfer_total,
        "pending_transfer_count": pending_transfer_count,
        "branch_rows": branch_rows,
        "has_data": bool(branch_rows) or float(totals_row.get("total_collected") or 0) > 0,
    }


def _parse_pasted_revenue_rows(raw_text):
    """Parse pasted table rows (tab/comma separated) to normalized revenue rows."""
    rows = []
    for line in (raw_text or "").splitlines():
        s = (line or "").strip()
        if not s:
            continue

        parts = [p.strip() for p in (s.split("\t") if "\t" in s else s.split(","))]
        parts = [p for p in parts if p != ""]
        if len(parts) < 3:
            continue

        first = parts[0].lower().replace(" ", "")
        if first in ["branch", "branchname"]:
            continue

        branch_name = parts[0]
        sales = _parse_money(parts[1])
        service = _parse_money(parts[2])

        total = sales + service
        zone = ""

        if len(parts) >= 4:
            fourth = parts[3]
            fourth_money = _parse_money(fourth)
            if re.search(r"[a-zA-Z]", fourth):
                zone = fourth
            else:
                total = fourth_money

        if len(parts) >= 5:
            zone = parts[4]

        if total <= 0:
            total = sales + service

        rows.append(
            {
                "branch_name": branch_name,
                "sales_profit": max(sales, 0),
                "service_charges": max(service, 0),
                "total_profit": max(total, 0),
                "zone": zone,
            }
        )

    return rows


def _iter_revenue_entry_excel_rows(uploaded_file):
    """Parse .xlsx/.csv for revenue entry bulk upload.
    Columns: Branch, Sales Profit, Service Charges, Total Profit (optional), Zone (optional)
    """
    filename = secure_filename(uploaded_file.filename or "")
    ext = os.path.splitext(filename)[1].lower()

    branch_aliases = ["branch", "branchname"]
    sales_aliases = ["salesprofit", "sales_profit", "sales", "salesprofitamount"]
    service_aliases = ["servicecharges", "service_charges", "service", "servicecharge"]
    total_aliases = ["totalprofit", "total_profit", "total", "totalamount"]
    zone_aliases = ["zone", "region", "area"]

    def _hkey(s):
        return re.sub(r'[^a-z0-9]+', '', str(s or '').strip().lower())

    def _build_map(headers):
        m = {}
        for i, h in enumerate(headers or []):
            k = _hkey(h)
            if k and k not in m:
                m[k] = i
        return m

    def _pick(m, aliases):
        for a in aliases:
            if a in m:
                return m[a]
        return None

    def _cell(row, idx):
        if idx is None or idx >= len(row):
            return ""
        return row[idx] if row[idx] is not None else ""

    def _process_rows(rows_iter, header_map):
        idx_branch = _pick(header_map, branch_aliases)
        idx_sales = _pick(header_map, sales_aliases)
        idx_service = _pick(header_map, service_aliases)
        idx_total = _pick(header_map, total_aliases)
        idx_zone = _pick(header_map, zone_aliases)

        if idx_branch is None:
            raise ValueError("Missing required column: Branch")
        if idx_sales is None and idx_service is None and idx_total is None:
            raise ValueError("Missing amount columns. Need: Sales Profit, Service Charges, or Total Profit")

        for row_number, row in rows_iter:
            row = list(row) if not isinstance(row, list) else row
            branch = str(_cell(row, idx_branch) or "").strip()
            if not branch:
                continue
            first = branch.lower().replace(" ", "")
            if first in ["branch", "branchname"]:
                continue  # skip header row if included in data
            sales = _parse_money(_cell(row, idx_sales)) if idx_sales is not None else 0.0
            service = _parse_money(_cell(row, idx_service)) if idx_service is not None else 0.0
            total = _parse_money(_cell(row, idx_total)) if idx_total is not None else 0.0
            if total <= 0:
                total = sales + service
            zone = str(_cell(row, idx_zone) or "").strip() if idx_zone is not None else ""
            yield row_number, branch, max(sales, 0), max(service, 0), max(total, 0), zone

    if ext == ".csv":
        raw_bytes = uploaded_file.stream.read()
        try:
            text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader, None)
        if not headers:
            raise ValueError("The file is empty")
        header_map = _build_map(headers)
        yield from _process_rows(enumerate(reader, start=2), header_map)
        return

    if ext == ".xlsx":
        if load_workbook is None:
            raise ValueError("openpyxl not installed. Use CSV or install openpyxl.")
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                raise ValueError("The file is empty")
            header_map = _build_map(headers)
            yield from _process_rows(enumerate(rows_iter, start=2), header_map)
        finally:
            workbook.close()
        return

    raise ValueError("Please upload .xlsx or .csv")


def _extract_used_spares_from_form(form):
    """Parse repeated spare rows and return clean [{'spare_name','amount'}]."""
    spare_names = form.getlist("used_spare_name[]")
    spare_amounts = form.getlist("used_spare_amount[]")

    items = []
    max_len = max(len(spare_names), len(spare_amounts))

    for i in range(max_len):
        name = (spare_names[i] if i < len(spare_names) else "").strip()
        raw_amount = (spare_amounts[i] if i < len(spare_amounts) else "").strip()

        if not name and not raw_amount:
            continue
        if not name:
            continue

        try:
            amount = float(raw_amount) if raw_amount else 0.0
        except (ValueError, TypeError):
            amount = 0.0

        if amount < 0:
            amount = 0.0

        items.append({"spare_name": name, "amount": amount})

    return items


def _user_has_all_branch_scope(role, session_branch):
    return role in ["super_admin", "admin"] and session_branch == "ALL"


def _branch_in_scope(role, session_branch, target_branch):
    if _user_has_all_branch_scope(role, session_branch):
        return True
    return (target_branch or "").strip() == (session_branch or "").strip()


def _get_branch_scope(role, session_branch):
    normalized_branch = (session_branch or "").strip()
    if _user_has_all_branch_scope(role, normalized_branch):
        return None
    if not normalized_branch or normalized_branch.upper() == "ALL":
        raise PermissionError("Invalid branch scope")
    return normalized_branch


def _resolve_branch_input(role, session_branch, requested_branch):
    normalized_requested = (requested_branch or "").strip()
    branch_scope = _get_branch_scope(role, session_branch)
    if branch_scope is None:
        return normalized_requested
    if normalized_requested and normalized_requested != branch_scope:
        raise PermissionError("Access denied")
    return branch_scope


def _load_known_branches(cursor):
    branch_values = [branch for branch in DEFAULT_BRANCHES if branch and branch != "ALL"]
    queries = [
        "SELECT DISTINCT branch_name FROM jobs WHERE branch_name IS NOT NULL AND branch_name <> ''",
        "SELECT DISTINCT branch_name FROM user_branches WHERE branch_name IS NOT NULL AND branch_name <> ''",
        "SELECT DISTINCT branch_name FROM staff_directory WHERE branch_name IS NOT NULL AND branch_name <> ''",
        "SELECT DISTINCT branch_name FROM branch_print_profiles WHERE branch_name IS NOT NULL AND branch_name <> '' AND branch_name <> 'ALL'",
        "SELECT DISTINCT branch_name FROM branch_revenue_targets WHERE branch_name IS NOT NULL AND branch_name <> '' AND branch_name <> 'ALL'",
        "SELECT value AS branch_name FROM dropdown_options WHERE type='branch'",
    ]

    for query in queries:
        try:
            cursor.execute(query)
            for row in cursor.fetchall():
                if isinstance(row, dict):
                    branch_values.append(str(row.get("branch_name") or "").strip())
                else:
                    branch_values.append(str((row[0] if row else "") or "").strip())
        except Error:
            continue

    deduped = []
    seen = set()
    for branch_name in branch_values:
        key = branch_name.upper()
        if not branch_name or key in seen:
            continue
        seen.add(key)
        deduped.append(branch_name)
    return sorted(deduped, key=lambda value: value.upper())


def _flash_staff_database_error(action_label, exc):
    app.logger.exception("%s failed", action_label)
    error_text = str(exc).strip() or exc.__class__.__name__
    flash(f"{action_label}: {error_text}", "danger")


def _resolve_known_branch(cursor, branch_name):
    normalized_branch = (branch_name or "").strip()
    if not normalized_branch:
        return None
    if normalized_branch.upper() == "ALL":
        return "ALL"

    for known_branch in _load_known_branches(cursor):
        if known_branch.upper() == normalized_branch.upper():
            return known_branch

    return None


def _is_known_branch(cursor, branch_name):
    return _resolve_known_branch(cursor, branch_name) is not None


def _build_report_datetime_bounds(from_date, to_date):
    return f"{from_date} 00:00:00", f"{to_date} 23:59:59"


def _get_syscare_membership_columns(cursor):
    try:
        cursor.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema=%s AND table_name='syscare_memberships'
            """,
            (DB_NAME,),
        )
        return {str(row.get("column_name") or "").strip().lower() for row in cursor.fetchall() if row.get("column_name")}
    except Error:
        return set()


def _load_admin_report_engineer_options(cursor, syscare_membership_columns):
    option_values = []
    queries = [
        "SELECT DISTINCT assigned_engineer AS engineer_name FROM jobs WHERE assigned_engineer IS NOT NULL AND assigned_engineer <> ''",
        "SELECT DISTINCT username AS engineer_name FROM users WHERE role='engineer' AND username IS NOT NULL AND username <> ''",
    ]

    if "assigned_engineer" in syscare_membership_columns:
        queries.append(
            "SELECT DISTINCT assigned_engineer AS engineer_name FROM syscare_memberships WHERE assigned_engineer IS NOT NULL AND assigned_engineer <> ''"
        )
    if "incharge" in syscare_membership_columns:
        queries.append(
            "SELECT DISTINCT incharge AS engineer_name FROM syscare_memberships WHERE incharge IS NOT NULL AND incharge <> ''"
        )

    for query in queries:
        try:
            cursor.execute(query)
            option_values.extend([str(row.get("engineer_name") or "").strip() for row in cursor.fetchall() if row.get("engineer_name")])
        except Error:
            continue

    deduped = []
    seen = set()
    for value in option_values:
        key = value.upper()
        if not value or key in seen:
            continue
        seen.add(key)
        deduped.append(value)
    return sorted(deduped, key=lambda value: value.upper())


def _build_admin_report_scope(role, session_branch, requested_branch, requested_engineer, syscare_membership_columns):
    branch_scope = _get_branch_scope(role, session_branch)
    normalized_branch = (requested_branch or "").strip()
    normalized_engineer = (requested_engineer or "").strip()

    if branch_scope is None:
        effective_branch = normalized_branch
    else:
        effective_branch = branch_scope

    job_where = ["1=1"]
    job_params = []
    syscare_where = ["1=1"]
    syscare_params = []

    if effective_branch:
        job_where.append("branch_name=%s")
        job_params.append(effective_branch)
        if syscare_membership_columns:
            syscare_where.append("branch_name=%s")
            syscare_params.append(effective_branch)

    if normalized_engineer:
        job_where.append("assigned_engineer=%s")
        job_params.append(normalized_engineer)
        if syscare_membership_columns:
            engineer_clauses = []
            engineer_params = []
            if "assigned_engineer" in syscare_membership_columns:
                engineer_clauses.append("assigned_engineer=%s")
                engineer_params.append(normalized_engineer)
            if "incharge" in syscare_membership_columns:
                engineer_clauses.append("incharge=%s")
                engineer_params.append(normalized_engineer)
            if engineer_clauses:
                syscare_where.append("(" + " OR ".join(engineer_clauses) + ")")
                syscare_params.extend(engineer_params)
            else:
                syscare_where.append("1=0")

    return {
        "branch_scope": branch_scope,
        "effective_branch": effective_branch,
        "effective_engineer": normalized_engineer,
        "job_where_sql": " AND ".join(job_where),
        "job_params": tuple(job_params),
        "syscare_where_sql": " AND ".join(syscare_where),
        "syscare_params": tuple(syscare_params),
    }


def _fetch_job_report_snapshot(cursor, job_where_sql, job_params, from_date, to_date):
    start_dt, end_dt = _build_report_datetime_bounds(from_date, to_date)
    closed_sql = "((closure_status IS NOT NULL AND closure_status <> '') OR status='Closed')"
    cursor.execute(
        f"""
        SELECT
            COALESCE(SUM(CASE WHEN created_at >= %s AND created_at <= %s THEN 1 ELSE 0 END), 0) AS new_calls,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s THEN 1 ELSE 0 END), 0) AS closed_calls,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s AND created_at IS NOT NULL AND TIMESTAMPDIFF(DAY, created_at, closure_date) > 7 THEN 1 ELSE 0 END), 0) AS closed_after_7_days,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s AND COALESCE(service_charges, 0) > 0 THEN 1 ELSE 0 END), 0) AS service_charge_count,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s THEN COALESCE(service_charges, 0) ELSE 0 END), 0) AS service_charge_total
        FROM jobs
        WHERE {job_where_sql}
        """,
        (
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            *job_params,
        ),
    )
    row = cursor.fetchone() or {}
    return {
        "new_calls": int(row.get("new_calls") or 0),
        "closed_calls": int(row.get("closed_calls") or 0),
        "closed_after_7_days": int(row.get("closed_after_7_days") or 0),
        "service_charge_count": int(row.get("service_charge_count") or 0),
        "service_charge_total": round(float(row.get("service_charge_total") or 0), 2),
    }


def _fetch_syscare_report_snapshot(cursor, syscare_where_sql, syscare_params, from_date, to_date, syscare_membership_columns):
    if not syscare_membership_columns:
        return {"syscare_count": 0, "syscare_amount": 0.0}

    cursor.execute(
        f"""
        SELECT COUNT(*) AS syscare_count, COALESCE(SUM(COALESCE(amount, 0)), 0) AS syscare_amount
        FROM syscare_memberships
        WHERE {syscare_where_sql}
          AND record_date >= %s
          AND record_date <= %s
        """,
        (*syscare_params, from_date, to_date),
    )
    row = cursor.fetchone() or {}
    return {
        "syscare_count": int(row.get("syscare_count") or 0),
        "syscare_amount": round(float(row.get("syscare_amount") or 0), 2),
    }


def _fetch_job_report_summary_rows(cursor, job_where_sql, job_params, group_expr, group_alias, from_date, to_date):
    start_dt, end_dt = _build_report_datetime_bounds(from_date, to_date)
    closed_sql = "((closure_status IS NOT NULL AND closure_status <> '') OR status='Closed')"
    cursor.execute(
        f"""
        SELECT
            {group_expr} AS {group_alias},
            COALESCE(SUM(CASE WHEN created_at >= %s AND created_at <= %s THEN 1 ELSE 0 END), 0) AS new_calls,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s THEN 1 ELSE 0 END), 0) AS closed_calls,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s AND created_at IS NOT NULL AND TIMESTAMPDIFF(DAY, created_at, closure_date) > 7 THEN 1 ELSE 0 END), 0) AS closed_after_7_days,
            COALESCE(SUM(CASE WHEN {closed_sql} AND closure_date >= %s AND closure_date <= %s THEN COALESCE(service_charges, 0) ELSE 0 END), 0) AS service_charge_total
        FROM jobs
        WHERE {job_where_sql}
        GROUP BY {group_expr}
        ORDER BY {group_alias}
        """,
        (
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            start_dt,
            end_dt,
            *job_params,
        ),
    )

    rows = []
    for row in cursor.fetchall():
        rows.append(
            {
                group_alias: row.get(group_alias),
                "new_calls": int(row.get("new_calls") or 0),
                "closed_calls": int(row.get("closed_calls") or 0),
                "closed_after_7_days": int(row.get("closed_after_7_days") or 0),
                "service_charge_total": round(float(row.get("service_charge_total") or 0), 2),
            }
        )
    return rows


def _fetch_syscare_report_summary_rows(cursor, syscare_where_sql, syscare_params, group_expr, group_alias, from_date, to_date, syscare_membership_columns):
    if not syscare_membership_columns:
        return []

    cursor.execute(
        f"""
        SELECT
            {group_expr} AS {group_alias},
            COUNT(*) AS syscare_count,
            COALESCE(SUM(COALESCE(amount, 0)), 0) AS syscare_amount
        FROM syscare_memberships
        WHERE {syscare_where_sql}
          AND record_date >= %s
          AND record_date <= %s
        GROUP BY {group_expr}
        ORDER BY {group_alias}
        """,
        (*syscare_params, from_date, to_date),
    )

    rows = []
    for row in cursor.fetchall():
        rows.append(
            {
                group_alias: row.get(group_alias),
                "syscare_count": int(row.get("syscare_count") or 0),
                "syscare_amount": round(float(row.get("syscare_amount") or 0), 2),
            }
        )
    return rows


def _merge_report_summary_rows(job_rows, syscare_rows, key_name):
    default_label = "Unknown" if key_name == "branch_name" else "Unassigned"
    merged = {}

    for row in job_rows or []:
        key = str(row.get(key_name) or "").strip() or default_label
        merged[key] = {
            key_name: key,
            "new_calls": int(row.get("new_calls") or 0),
            "closed_calls": int(row.get("closed_calls") or 0),
            "closed_after_7_days": int(row.get("closed_after_7_days") or 0),
            "service_charge_total": round(float(row.get("service_charge_total") or 0), 2),
            "syscare_count": 0,
            "syscare_amount": 0.0,
        }

    for row in syscare_rows or []:
        key = str(row.get(key_name) or "").strip() or default_label
        entry = merged.setdefault(
            key,
            {
                key_name: key,
                "new_calls": 0,
                "closed_calls": 0,
                "closed_after_7_days": 0,
                "service_charge_total": 0.0,
                "syscare_count": 0,
                "syscare_amount": 0.0,
            },
        )
        entry["syscare_count"] = int(row.get("syscare_count") or 0)
        entry["syscare_amount"] = round(float(row.get("syscare_amount") or 0), 2)

    return sorted(merged.values(), key=lambda item: str(item.get(key_name) or "").upper())


def ensure_engineer_revenue_entries_table():
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS engineer_revenue_entries (
                id INT AUTO_INCREMENT PRIMARY KEY,
                entry_date DATE NOT NULL,
                engineer_name VARCHAR(255) NOT NULL,
                employee_code VARCHAR(255) DEFAULT '',
                branch_name VARCHAR(255) NOT NULL,
                sales_revenue DECIMAL(12,2) NOT NULL DEFAULT 0,
                service_charges DECIMAL(12,2) NOT NULL DEFAULT 0,
                total_revenue DECIMAL(12,2) NOT NULL DEFAULT 0,
                source_type VARCHAR(64) DEFAULT 'manual',
                created_by VARCHAR(255) DEFAULT '',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY idx_engineer_revenue_unique (entry_date, engineer_name, branch_name, employee_code),
                INDEX idx_engineer_revenue_date (entry_date),
                INDEX idx_engineer_revenue_engineer (engineer_name),
                INDEX idx_engineer_revenue_branch (branch_name)
            ) ENGINE=InnoDB
            """
        )
        db.commit()
    except Exception:
        pass
    finally:
        _safe_close(cursor, db)


def _fetch_staff_ranking_rows(cursor, job_where_sql, job_params, from_date, to_date):
    start_dt, end_dt = _build_report_datetime_bounds(from_date, to_date)
    cursor.execute(
        f"""
        SELECT
            COALESCE(TRIM(assigned_engineer), 'Unassigned') AS engineer_name,
            COALESCE(TRIM(branch_name), 'Unknown') AS branch_name,
            COUNT(*) AS total_calls,
            SUM(CASE WHEN (COALESCE(TRIM(status), '') <> 'Closed' AND (closure_status IS NULL OR TRIM(closure_status) = '')) THEN 1 ELSE 0 END) AS open_calls,
            SUM(CASE WHEN (LOWER(COALESCE(closure_status, '')) LIKE '%success%' OR (LOWER(COALESCE(status, '')) = 'closed' AND LOWER(COALESCE(closure_status, '')) NOT LIKE '%failed%')) THEN 1 ELSE 0 END) AS closed_success,
            SUM(CASE WHEN LOWER(COALESCE(closure_status, '')) LIKE '%failed%' THEN 1 ELSE 0 END) AS closed_failed
        FROM jobs
        WHERE {job_where_sql}
          AND created_at >= %s
          AND created_at <= %s
        GROUP BY COALESCE(TRIM(assigned_engineer), 'Unassigned'), COALESCE(TRIM(branch_name), 'Unknown')
        ORDER BY closed_success DESC, total_calls DESC, engineer_name ASC, branch_name ASC
        """,
        (*job_params, start_dt, end_dt),
    )
    rows = []
    for row in cursor.fetchall():
        rows.append(
            {
                "engineer_name": str(row.get("engineer_name") or "Unassigned").strip() or "Unassigned",
                "employee_code": "",
                "branch_name": str(row.get("branch_name") or "Unknown").strip() or "Unknown",
                "total_calls": int(row.get("total_calls") or 0),
                "open_calls": int(row.get("open_calls") or 0),
                "closed_success": int(row.get("closed_success") or 0),
                "closed_failed": int(row.get("closed_failed") or 0),
                "sales_revenue": 0.0,
                "service_charges": 0.0,
                "total_revenue": 0.0,
            }
        )
    return rows


def _fetch_staff_revenue_rows(cursor, filter_branch, filter_engineer, from_date, to_date):
    where_clauses = ["1=1"]
    params = []

    if filter_branch:
        where_clauses.append("branch_name=%s")
        params.append(filter_branch)
    if filter_engineer:
        where_clauses.append("engineer_name=%s")
        params.append(filter_engineer)

    where_clauses.append("entry_date >= %s")
    params.append(from_date)
    where_clauses.append("entry_date <= %s")
    params.append(to_date)

    cursor.execute(
        f"""
        SELECT
            COALESCE(TRIM(engineer_name), 'Unassigned') AS engineer_name,
            COALESCE(TRIM(branch_name), 'Unknown') AS branch_name,
            COALESCE(SUM(COALESCE(sales_revenue, 0)), 0) AS sales_revenue,
            COALESCE(SUM(COALESCE(service_charges, 0)), 0) AS service_charges,
            COALESCE(SUM(COALESCE(total_revenue, 0)), 0) AS total_revenue
        FROM engineer_revenue_entries
        WHERE {' AND '.join(where_clauses)}
        GROUP BY COALESCE(TRIM(engineer_name), 'Unassigned'), COALESCE(TRIM(branch_name), 'Unknown')
        ORDER BY total_revenue DESC, engineer_name ASC, branch_name ASC
        """,
        tuple(params),
    )
    rows = []
    for row in cursor.fetchall():
        rows.append(
            {
                "engineer_name": str(row.get("engineer_name") or "Unassigned").strip() or "Unassigned",
                "employee_code": "",
                "branch_name": str(row.get("branch_name") or "Unknown").strip() or "Unknown",
                "sales_revenue": round(float(row.get("sales_revenue") or 0), 2),
                "service_charges": round(float(row.get("service_charges") or 0), 2),
                "total_revenue": round(float(row.get("total_revenue") or 0), 2),
            }
        )
    return rows


def _merge_staff_rank_rows(call_rows, revenue_rows):
    revenue_map = {}
    for row in revenue_rows or []:
        key = (str(row.get("engineer_name") or "Unassigned").strip() or "Unassigned", str(row.get("branch_name") or "Unknown").strip() or "Unknown")
        revenue_map[key] = row

    merged = []
    for row in call_rows or []:
        key = (row.get("engineer_name"), row.get("branch_name"))
        revenue_row = revenue_map.get(key, {})
        merged_row = dict(row)
        merged_row["sales_revenue"] = round(float(revenue_row.get("sales_revenue") or 0), 2)
        merged_row["service_charges"] = round(float(revenue_row.get("service_charges") or 0), 2)
        merged_row["total_revenue"] = round(float(revenue_row.get("total_revenue") or 0), 2)
        merged.append(merged_row)

    for row in revenue_rows or []:
        key = (str(row.get("engineer_name") or "Unassigned").strip() or "Unassigned", str(row.get("branch_name") or "Unknown").strip() or "Unknown")
        if any(item.get("engineer_name") == key[0] and item.get("branch_name") == key[1] for item in merged):
            continue
        merged.append(
            {
                "engineer_name": key[0],
                "employee_code": "",
                "branch_name": key[1],
                "total_calls": 0,
                "open_calls": 0,
                "closed_success": 0,
                "closed_failed": 0,
                "sales_revenue": round(float(row.get("sales_revenue") or 0), 2),
                "service_charges": round(float(row.get("service_charges") or 0), 2),
                "total_revenue": round(float(row.get("total_revenue") or 0), 2),
            }
        )

    return merged


def _get_admin_reports_context(args, role, session_branch):
    db = None
    cursor = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        today = business_now_naive().date()
        month_start = today.replace(day=1)
        today_str = today.strftime("%Y-%m-%d")
        month_start_str = month_start.strftime("%Y-%m-%d")

        from_date = _normalize_date_input(args.get("from_date", "")) or month_start_str
        to_date = _normalize_date_input(args.get("to_date", "")) or today_str
        if from_date > to_date:
            from_date, to_date = to_date, from_date

        syscare_membership_columns = _get_syscare_membership_columns(cursor)
        branch_scope = _get_branch_scope(role, session_branch)
        branch_options = _load_known_branches(cursor) if branch_scope is None else [branch_scope]

        requested_branch = (args.get("filter_branch") or "").strip()
        filter_branch = _resolve_known_branch(cursor, requested_branch) if requested_branch else ""
        if branch_scope is not None:
            filter_branch = branch_scope

        engineer_options = _load_admin_report_engineer_options(cursor, syscare_membership_columns)
        filter_engineer = (args.get("filter_engineer") or "").strip()
        if filter_engineer and filter_engineer not in engineer_options:
            engineer_options = sorted(engineer_options + [filter_engineer], key=lambda value: value.upper())

        scope = _build_admin_report_scope(role, session_branch, filter_branch, filter_engineer, syscare_membership_columns)

        today_jobs = _fetch_job_report_snapshot(cursor, scope["job_where_sql"], scope["job_params"], today_str, today_str)
        today_syscare = _fetch_syscare_report_snapshot(cursor, scope["syscare_where_sql"], scope["syscare_params"], today_str, today_str, syscare_membership_columns)
        month_jobs = _fetch_job_report_snapshot(cursor, scope["job_where_sql"], scope["job_params"], month_start_str, today_str)
        month_syscare = _fetch_syscare_report_snapshot(cursor, scope["syscare_where_sql"], scope["syscare_params"], month_start_str, today_str, syscare_membership_columns)
        selected_jobs = _fetch_job_report_snapshot(cursor, scope["job_where_sql"], scope["job_params"], from_date, to_date)
        selected_syscare = _fetch_syscare_report_snapshot(cursor, scope["syscare_where_sql"], scope["syscare_params"], from_date, to_date, syscare_membership_columns)

        branch_group_expr = "COALESCE(NULLIF(TRIM(branch_name), ''), 'Unknown')"
        engineer_group_expr = "COALESCE(NULLIF(TRIM(assigned_engineer), ''), 'Unassigned')"
        syscare_engineer_expr = (
            "COALESCE(NULLIF(TRIM(assigned_engineer), ''), NULLIF(TRIM(incharge), ''), 'Unassigned')"
            if "assigned_engineer" in syscare_membership_columns
            else "COALESCE(NULLIF(TRIM(incharge), ''), 'Unassigned')"
        )

        branch_job_rows = _fetch_job_report_summary_rows(
            cursor,
            scope["job_where_sql"],
            scope["job_params"],
            branch_group_expr,
            "branch_name",
            from_date,
            to_date,
        )
        engineer_job_rows = _fetch_job_report_summary_rows(
            cursor,
            scope["job_where_sql"],
            scope["job_params"],
            engineer_group_expr,
            "engineer_name",
            from_date,
            to_date,
        )
        branch_syscare_rows = _fetch_syscare_report_summary_rows(
            cursor,
            scope["syscare_where_sql"],
            scope["syscare_params"],
            branch_group_expr,
            "branch_name",
            from_date,
            to_date,
            syscare_membership_columns,
        )
        engineer_syscare_rows = _fetch_syscare_report_summary_rows(
            cursor,
            scope["syscare_where_sql"],
            scope["syscare_params"],
            syscare_engineer_expr,
            "engineer_name",
            from_date,
            to_date,
            syscare_membership_columns,
        )

        staff_call_rows = _fetch_staff_ranking_rows(
            cursor,
            scope["job_where_sql"],
            scope["job_params"],
            from_date,
            to_date,
        )
        staff_revenue_rows = _fetch_staff_revenue_rows(
            cursor,
            filter_branch,
            filter_engineer,
            from_date,
            to_date,
        )
        staff_rank_rows = _merge_staff_rank_rows(staff_call_rows, staff_revenue_rows)
        staff_closed_rank_rows = sorted(
            staff_rank_rows,
            key=lambda row: (
                -(int(row.get("closed_success") or 0)),
                -(int(row.get("total_calls") or 0)),
                str(row.get("engineer_name") or "").upper(),
                str(row.get("branch_name") or "").upper(),
            ),
        )
        staff_revenue_rank_rows = sorted(
            staff_rank_rows,
            key=lambda row: (
                -(float(row.get("total_revenue") or 0)),
                -(float(row.get("sales_revenue") or 0)),
                str(row.get("engineer_name") or "").upper(),
                str(row.get("branch_name") or "").upper(),
            ),
        )

        return {
            "from_date": from_date,
            "to_date": to_date,
            "filter_branch": filter_branch,
            "filter_engineer": filter_engineer,
            "branch_options": branch_options,
            "engineer_options": engineer_options,
            "has_global_scope": branch_scope is None,
            "today_date": today_str,
            "month_start": month_start_str,
            "today_jobs": today_jobs,
            "today_syscare": today_syscare,
            "month_jobs": month_jobs,
            "month_syscare": month_syscare,
            "selected_jobs": selected_jobs,
            "selected_syscare": selected_syscare,
            "branch_summary_rows": _merge_report_summary_rows(branch_job_rows, branch_syscare_rows, "branch_name"),
            "engineer_summary_rows": _merge_report_summary_rows(engineer_job_rows, engineer_syscare_rows, "engineer_name"),
            "staff_closed_rank_rows": staff_closed_rank_rows,
            "staff_revenue_rank_rows": staff_revenue_rank_rows,
        }
    finally:
        _safe_close(cursor, db)


def _get_staff_management_context(args):
    db = None
    cursor = None

    try:
        ensure_staff_directory_table()
        db = get_db()
        cursor = db.cursor(dictionary=True)

        branch_options = _load_known_branches(cursor)
        requested_branch = (args.get("filter_branch") or "").strip()
        filter_branch = _resolve_known_branch(cursor, requested_branch) or requested_branch
        search_q = (args.get("q") or "").strip()

        where_clauses = ["1=1"]
        params = []

        if filter_branch:
            where_clauses.append("branch_name=%s")
            params.append(filter_branch)
        if search_q:
            like_value = f"%{search_q}%"
            where_clauses.append("(staff_name LIKE %s OR contact_number LIKE %s OR branch_name LIKE %s)")
            params.extend([like_value, like_value, like_value])

        where_sql = " AND ".join(where_clauses)

        cursor.execute(
            f"SELECT * FROM staff_directory WHERE {where_sql} AND resigned_date IS NULL ORDER BY staff_name ASC, id ASC",
            tuple(params),
        )
        staff_rows = cursor.fetchall()

        cursor.execute(
            f"SELECT * FROM staff_directory WHERE {where_sql} AND resigned_date IS NOT NULL ORDER BY resigned_date DESC, staff_name ASC, id ASC",
            tuple(params),
        )
        resigned_staff_rows = cursor.fetchall()

        staff_count = 0
        resigned_count = len(resigned_staff_rows)
        payable_total = 0.0
        salary_total = 0.0
        deduction_total = 0.0

        for row in staff_rows:
            row["room_rent"] = _compute_staff_room_rent(row)
            row["total_payable"] = _compute_staff_total_payable(row)
            staff_count += 1
            payable_total += row["total_payable"]
            salary_total += float(row.get("salary") or 0)
            deduction_total += (
                float(row.get("esi") or 0)
                + float(row.get("pf") or 0)
                + row["room_rent"]
            )

        for row in resigned_staff_rows:
            row["room_rent"] = _compute_staff_room_rent(row)
            row["total_payable"] = _compute_staff_total_payable(row)

        return {
            "staff_rows": staff_rows,
            "resigned_staff_rows": resigned_staff_rows,
            "branch_options": branch_options,
            "filter_branch": filter_branch,
            "search_q": search_q,
            "staff_count": staff_count,
            "resigned_count": resigned_count,
            "payable_total": round(payable_total, 2),
            "salary_total": round(salary_total, 2),
            "deduction_total": round(deduction_total, 2),
        }
    finally:
        _safe_close(cursor, db)


def _compute_staff_total_payable(staff_row):
    return round(
        float(staff_row.get("salary") or 0)
        + float(staff_row.get("esi") or 0)
        + float(staff_row.get("pf") or 0)
        + _compute_staff_room_rent(staff_row),
        2,
    )


def _compute_staff_room_rent(staff_row):
    return round(float(staff_row.get("room") or 0) + float(staff_row.get("rent") or 0), 2)


def _extract_staff_member_payload(cursor, form):
    staff_name = (form.get("staff_name") or "").strip()
    contact_number = (form.get("contact_number") or "").strip()
    branch_name = (form.get("branch_name") or "").strip()
    joined_date = _normalize_date_input(form.get("joined_date", ""))
    resigned_date = _normalize_date_input(form.get("resigned_date", "")) or None
    if "room_rent" in form:
        room_rent = max(_parse_money(form.get("room_rent")), 0)
    else:
        room_rent = max(_parse_money(form.get("room")), 0) + max(_parse_money(form.get("rent")), 0)

    payload = {
        "staff_name": staff_name,
        "contact_number": contact_number,
        "branch_name": _resolve_known_branch(cursor, branch_name) or branch_name,
        "salary": max(_parse_money(form.get("salary")), 0),
        "esi": max(_parse_money(form.get("esi")), 0),
        "pf": max(_parse_money(form.get("pf")), 0),
        "room": room_rent,
        "rent": 0,
        "joined_date": joined_date,
        "resigned_date": resigned_date,
    }

    errors = []
    if not payload["staff_name"]:
        errors.append("Staff name is required")
    if not payload["contact_number"]:
        errors.append("Contact number is required")
    if not payload["branch_name"]:
        errors.append("Branch is required")
    if not payload["joined_date"]:
        errors.append("Joined date is required")
    if payload["joined_date"] and payload["resigned_date"] and payload["resigned_date"] < payload["joined_date"]:
        errors.append("Resigned date cannot be before joined date")

    return payload, errors


def _build_job_transfer_branch_scope_clause(branch_name, table_alias=""):
    normalized_branch = (branch_name or "").strip()
    if not normalized_branch or normalized_branch.upper() == "ALL":
        return "1=1", ()

    prefix = f"{table_alias}." if table_alias else ""
    branch_expr = f"{prefix}branch_name"
    job_id_expr = f"{prefix}id"
    status_expr = f"{prefix}status"
    active_placeholders = ", ".join(["%s"] * len(ACTIVE_TRANSFER_STATUSES))
    completed_placeholders = ", ".join(["%s"] * len(COMPLETED_TRANSFER_STATUSES))
    clause = (
        f"({branch_expr}=%s OR "
        f"{job_id_expr} IN ("
        "SELECT transfer.job_id FROM job_service_transfers transfer "
        f"WHERE transfer.to_branch_name=%s AND transfer.status IN ({active_placeholders})"
        ") OR "
        f"({status_expr}='Closed' AND {job_id_expr} IN ("
        "SELECT transfer.job_id FROM job_service_transfers transfer "
        f"WHERE transfer.to_branch_name=%s AND transfer.status IN ({completed_placeholders})"
        ")))"
    )
    return (
        clause,
        (
            normalized_branch,
            normalized_branch,
            *ACTIVE_TRANSFER_STATUSES,
            normalized_branch,
            *COMPLETED_TRANSFER_STATUSES,
        ),
    )


def _build_job_transfer_engineer_filter_clause(engineer_names, table_alias=""):
    normalized_names = [str(name or "").strip() for name in (engineer_names or []) if str(name or "").strip()]
    if not normalized_names:
        return "", ()

    prefix = f"{table_alias}." if table_alias else ""
    assigned_expr = f"{prefix}assigned_engineer"
    job_id_expr = f"{prefix}id"
    status_expr = f"{prefix}status"
    engineer_placeholders = ", ".join(["%s"] * len(normalized_names))
    active_placeholders = ", ".join(["%s"] * len(ACTIVE_TRANSFER_STATUSES))
    completed_placeholders = ", ".join(["%s"] * len(COMPLETED_TRANSFER_STATUSES))
    clause = (
        f"({assigned_expr} IN ({engineer_placeholders}) OR EXISTS ("
        "SELECT 1 FROM job_service_transfers transfer "
        f"WHERE transfer.job_id={job_id_expr} "
        f"AND transfer.specialist_engineer IN ({engineer_placeholders}) "
        f"AND (transfer.status IN ({active_placeholders}) OR ({status_expr}='Closed' AND transfer.status IN ({completed_placeholders})))"
        "))"
    )
    return (
        clause,
        (
            *normalized_names,
            *normalized_names,
            *ACTIVE_TRANSFER_STATUSES,
            *COMPLETED_TRANSFER_STATUSES,
        ),
    )


def _build_job_active_transfer_exists_clause(table_alias="jobs"):
    job_ref = f"{table_alias}.id" if table_alias else "id"
    active_placeholders = ", ".join(["%s"] * len(ACTIVE_TRANSFER_STATUSES))
    clause = (
        "EXISTS ("
        "SELECT 1 FROM job_service_transfers transfer "
        f"WHERE transfer.job_id={job_ref} AND transfer.status IN ({active_placeholders})"
        ")"
    )
    return clause, tuple(ACTIVE_TRANSFER_STATUSES)


def _prepare_job_service_transfer_row(row):
    prepared = dict(row or {})
    prepared["status"] = normalize_transfer_status(prepared.get("status"), default="Sent")
    prepared["specialist_engineer"] = str(prepared.get("specialist_engineer") or "").strip()
    prepared["service_type"] = str(prepared.get("service_type") or "").strip()
    prepared["request_notes"] = str(prepared.get("request_notes") or "").strip()
    prepared["status_notes"] = str(prepared.get("status_notes") or "").strip()
    prepared["sent_by"] = str(prepared.get("sent_by") or "").strip()
    prepared["updated_by"] = str(prepared.get("updated_by") or "").strip()
    prepared["internal_service_charge"] = round(float(prepared.get("internal_service_charge") or 0), 2)
    prepared["created_on"] = format_datetime_display(prepared.get("created_at"))
    prepared["updated_on"] = format_datetime_display(prepared.get("updated_at"))
    prepared["accepted_on"] = format_datetime_display(prepared.get("accepted_at"))
    prepared["completed_on"] = format_datetime_display(prepared.get("completed_at"))
    prepared["returned_on"] = format_datetime_display(prepared.get("returned_at"))
    return prepared


def _fetch_job_service_transfer_rows(cursor, job_id):
    cursor.execute("SELECT * FROM job_service_transfers WHERE job_id=%s ORDER BY id DESC", (job_id,))
    return [_prepare_job_service_transfer_row(row) for row in cursor.fetchall()]


def _fetch_job_service_transfer_by_id(cursor, transfer_id):
    cursor.execute("SELECT * FROM job_service_transfers WHERE id=%s", (transfer_id,))
    row = cursor.fetchone()
    if not row:
        return None
    return _prepare_job_service_transfer_row(row)


def _fetch_job_service_transfer_summary_map(cursor, job_ids):
    normalized_job_ids = sorted({int(job_id) for job_id in (job_ids or []) if str(job_id or "").strip()})
    if not normalized_job_ids:
        return {}

    job_placeholders = ", ".join(["%s"] * len(normalized_job_ids))
    completed_placeholders = ", ".join(["%s"] * len(COMPLETED_TRANSFER_STATUSES))
    active_placeholders = ", ".join(["%s"] * len(ACTIVE_TRANSFER_STATUSES))
    summary_map = {job_id: {"specialist_service_total": 0.0, "active_transfer": None, "latest_completed_transfer": None} for job_id in normalized_job_ids}

    cursor.execute(
        f"""
        SELECT job_id, COALESCE(SUM(internal_service_charge), 0) AS specialist_service_total
        FROM job_service_transfers
        WHERE job_id IN ({job_placeholders})
          AND status IN ({completed_placeholders})
        GROUP BY job_id
        """,
        tuple(normalized_job_ids + list(COMPLETED_TRANSFER_STATUSES)),
    )
    for row in cursor.fetchall():
        job_id = int(row.get("job_id") or 0)
        if job_id in summary_map:
            summary_map[job_id]["specialist_service_total"] = round(float(row.get("specialist_service_total") or 0), 2)

    cursor.execute(
        f"""
        SELECT *
        FROM job_service_transfers
        WHERE job_id IN ({job_placeholders})
          AND status IN ({active_placeholders})
        ORDER BY job_id ASC, id DESC
        """,
        tuple(normalized_job_ids + list(ACTIVE_TRANSFER_STATUSES)),
    )
    for row in cursor.fetchall():
        job_id = int(row.get("job_id") or 0)
        if job_id in summary_map and summary_map[job_id]["active_transfer"] is None:
            summary_map[job_id]["active_transfer"] = _prepare_job_service_transfer_row(row)

    cursor.execute(
        f"""
        SELECT *
        FROM job_service_transfers
        WHERE job_id IN ({job_placeholders})
          AND status IN ({completed_placeholders})
        ORDER BY job_id ASC, id DESC
        """,
        tuple(normalized_job_ids + list(COMPLETED_TRANSFER_STATUSES)),
    )
    for row in cursor.fetchall():
        job_id = int(row.get("job_id") or 0)
        if job_id in summary_map and summary_map[job_id]["latest_completed_transfer"] is None:
            summary_map[job_id]["latest_completed_transfer"] = _prepare_job_service_transfer_row(row)

    return summary_map


def _decorate_job_rows_with_transfer_summary(cursor, rows):
    row_list = rows or []
    summary_map = _fetch_job_service_transfer_summary_map(cursor, [row.get("id") for row in row_list])

    for row in row_list:
        job_id = int(row.get("id") or 0)
        service_total = float(row.get("service_charges") or 0)
        summary = summary_map.get(job_id) or {}
        split = compute_job_transfer_split(service_total, summary.get("specialist_service_total") or 0)
        highlighted_transfer = summary.get("active_transfer") or summary.get("latest_completed_transfer")
        row["specialist_service_total"] = split["specialist_service_total"]
        row["closing_branch_service_margin"] = split["closing_branch_service_margin"]
        row["active_specialist_transfer"] = summary.get("active_transfer")
        row["latest_completed_transfer"] = summary.get("latest_completed_transfer")
        row["specialist_transfer_is_active"] = bool(summary.get("active_transfer"))
        row["specialist_transfer_branch_name"] = highlighted_transfer.get("to_branch_name") if highlighted_transfer else ""
        row["specialist_transfer_engineer"] = highlighted_transfer.get("specialist_engineer") if highlighted_transfer else ""
        row["specialist_transfer_status"] = highlighted_transfer.get("status") if highlighted_transfer else ""
        row["has_specialist_transfer_history"] = bool(highlighted_transfer) or split["specialist_service_total"] > 0

    return row_list


def _user_can_manage_job_core(role, session_branch, job):
    if role not in ["super_admin", "admin", "coordinator"]:
        return False
    return _branch_in_scope(role, session_branch, job.get("branch_name"))


def _user_can_send_job_transfer(role, session_branch, job):
    if role not in ["super_admin", "admin", "coordinator"]:
        return False
    return _branch_in_scope(role, session_branch, job.get("branch_name"))


def _user_can_update_job_transfer(role, session_branch, job, transfer_row):
    if role not in ["super_admin", "admin", "coordinator", "engineer"]:
        return False
    if _user_has_all_branch_scope(role, session_branch):
        return True

    normalized_branch = (session_branch or "").strip()
    if not normalized_branch or normalized_branch.upper() == "ALL":
        return False

    if role in ["super_admin", "admin", "coordinator"] and normalized_branch == str(job.get("branch_name") or "").strip():
        return True
    return normalized_branch == str(transfer_row.get("to_branch_name") or "").strip()


def _fetch_scoped_job(cursor, job_id, role, session_branch):
    branch_scope = _get_branch_scope(role, session_branch)
    if branch_scope is None:
        cursor.execute("SELECT * FROM jobs WHERE id=%s", (job_id,))
    else:
        scope_sql, scope_params = _build_job_transfer_branch_scope_clause(branch_scope)
        cursor.execute(f"SELECT * FROM jobs WHERE id=%s AND {scope_sql}", (job_id, *scope_params))
    return cursor.fetchone()


def _quotation_branch_options(cursor, role, session_branch):
    if _user_has_all_branch_scope(role, session_branch):
        return _load_known_branches(cursor)

    return [(session_branch or "").strip()]


def _extract_quotation_items_from_form(form):
    item_names = form.getlist("item_name[]")
    narrations = form.getlist("narration[]")
    qty_list = form.getlist("qty[]")
    amount_list = form.getlist("amount[]")

    items = []
    max_len = max(len(item_names), len(narrations), len(qty_list), len(amount_list))

    for i in range(max_len):
        item_name = (item_names[i] if i < len(item_names) else "").strip()
        narration = (narrations[i] if i < len(narrations) else "").strip()
        qty = max(_parse_money(qty_list[i] if i < len(qty_list) else 0), 0)
        amount = max(_parse_money(amount_list[i] if i < len(amount_list) else 0), 0)

        if not item_name and not narration and qty <= 0 and amount <= 0:
            continue
        if not item_name:
            continue

        final_amount = round(qty * amount, 2)

        items.append(
            {
                "item_name": item_name,
                "narration": narration,
                "qty": qty,
                "amount": amount,
                "final_amount": final_amount,
            }
        )

    return items


def _next_quote_number(cursor):
    prefix = business_now_naive().strftime("QTN-%Y%m")
    next_no = _next_sequence_value(
        cursor,
        f"quotations:{prefix}",
        """
        SELECT COALESCE(MAX(CAST(SUBSTRING_INDEX(quote_number, '-', -1) AS UNSIGNED)), 0) AS seed_value
        FROM quotations
        WHERE quote_number LIKE %s
        """,
        (prefix + "-%",),
    )
    return f"{prefix}-{next_no:04d}"


def _next_sequence_value(cursor, sequence_key, seed_sql, seed_params=()):
    cursor.execute(
        "SELECT `last_value` FROM sequence_counters WHERE sequence_key = %s FOR UPDATE",
        (sequence_key,),
    )
    row = cursor.fetchone()
    cursor.execute(seed_sql, seed_params)
    seed_row = cursor.fetchone() or {}
    seed_value = seed_row.get("seed_value") if isinstance(seed_row, dict) else (seed_row[0] if seed_row else 0)
    seed_value = int(seed_value or 0)
    if not row:
        cursor.execute(
            """
            INSERT INTO sequence_counters (sequence_key, `last_value`)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE sequence_key = sequence_key
            """,
            (sequence_key, seed_value),
        )
        cursor.execute(
            "SELECT `last_value` FROM sequence_counters WHERE sequence_key = %s FOR UPDATE",
            (sequence_key,),
        )
        row = cursor.fetchone()

    last_value = row.get("last_value") if isinstance(row, dict) else (row[0] if row else 0)
    last_value = int(last_value or 0)
    if seed_value > last_value:
        last_value = seed_value
        cursor.execute(
            "UPDATE sequence_counters SET `last_value` = %s WHERE sequence_key = %s",
            (last_value, sequence_key),
        )

    next_value = last_value + 1
    cursor.execute(
        "UPDATE sequence_counters SET `last_value` = %s WHERE sequence_key = %s",
        (next_value, sequence_key),
    )
    return next_value


def _next_job_number(cursor):
    return _next_sequence_value(
        cursor,
        "jobs",
        "SELECT COALESCE(MAX(job_number), 79999) AS seed_value FROM jobs",
    )


# Ensure schema bootstrap does not crash app startup on hosting.
def bootstrap_schema_safely():
    try:
        ensure_created_at_column()
        ensure_closure_columns()
        ensure_job_billing_columns()
        ensure_profile_picture_column()
        ensure_password_hash_column_capacity()
        ensure_status_update_columns()
        ensure_used_spares_table()
        ensure_job_attachments_table()
        ensure_branch_print_profiles_table()
        ensure_branch_revenue_targets_table()
        ensure_branch_revenue_entries_table()
        ensure_revenue_entry_period_locks_table()
        ensure_branch_cashflow_entries_table()
        ensure_branch_cash_transfer_requests_table()
        ensure_job_service_transfers_table()
        ensure_quotations_tables()
        ensure_sequence_counters_table()
        ensure_staff_directory_table()
        ensure_engineer_revenue_entries_table()
        ensure_performance_indexes()
        ensure_job_status_logs_table()
    except Exception as e:
        print(f"Schema bootstrap skipped: {e}")
# Create job_status_logs table for full job status history
def ensure_job_status_logs_table():
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            '''
            CREATE TABLE IF NOT EXISTS job_status_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                job_id INT NOT NULL,
                status VARCHAR(255) NOT NULL,
                notes TEXT,
                updated_by VARCHAR(255) NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
                INDEX idx_job_status_logs_job (job_id)
            )
            '''
        )
        db.commit()
    except Exception:
        pass
    finally:
        _safe_close(cursor, db)


bootstrap_schema_safely()
_purge_expired_job_attachments(force=True)


# ---------------- LOGIN ---------------- #

@app.route("/")
def home():
    if "username" in session:
        return redirect("/dashboard")
    return redirect("/login")

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        db = None
        cursor = None
        client_key = _get_client_key()

        if _login_rate_limited(client_key):
            flash("Too many failed login attempts. Try again later.", "danger")
            return redirect("/login")

        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        branch = (request.form.get("branch") or "").strip()

        if not all([username, password, branch]):
            flash("All fields are required", "danger")
            return redirect("/login")

        try:

            db = get_db()
            cursor = db.cursor(dictionary=True)

            cursor.execute(
                "SELECT id, username, password, role FROM users WHERE LOWER(username)=LOWER(%s)",
                (username,)
            )

            user = cursor.fetchone()
            password_ok = False
            password_reset_required = False

            if user:
                password_ok, password_reset_required = _verify_password(user.get("password"), password)

            if user and password_reset_required:
                _record_failed_login_attempt(client_key)
                app.logger.warning(
                    "Blocked login for user %s because the stored password is not hashed",
                    user.get("username"),
                )
                flash("Invalid username or password", "danger")
                return redirect("/login")

            if not user or not password_ok:
                _record_failed_login_attempt(client_key)
                flash("Invalid username or password", "danger")
                return redirect("/login")

            role = user["role"]
            username = user["username"]
            canonical_branch = _resolve_known_branch(cursor, branch)

            if not canonical_branch:
                flash("Select a valid branch", "danger")
                return redirect("/login")

            if role in ["super_admin", "admin"]:
                session.clear()
                session.permanent = True
                session["username"] = username
                session["role"] = role
                session["branch"] = canonical_branch
                session["dashboard_branch_scope"] = "ALL"
                _clear_login_attempts(client_key)
                return redirect("/dashboard")

            # Normal user → check branch
            cursor.execute(
                "SELECT 1 FROM user_branches WHERE LOWER(username)=LOWER(%s) AND UPPER(branch_name)=UPPER(%s)",
                (username, canonical_branch)
            )

            if not cursor.fetchone():
                _record_failed_login_attempt(client_key)
                flash("No access to selected branch", "danger")
                return redirect("/login")

            session.clear()
            session.permanent = True
            session["username"] = username
            session["role"] = role
            session["branch"] = canonical_branch
            _clear_login_attempts(client_key)

            return redirect("/dashboard")

        except Error as e:
            _flash_internal_error("Login failed. Try again later.", e)
            return redirect("/login")

        finally:
            _safe_close(cursor, db)

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        branches = ["ALL"] + _load_known_branches(cursor)
    except Error:
        branches = DEFAULT_BRANCHES
    finally:
        _safe_close(cursor, db)

    return render_template("login.html", branches=branches)


# ---------------- LOGOUT ---------------- #

@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully", "info")
    return redirect("/login")


def _get_jobs_view_definition(raw_view):
    view_map = {
        "all": ("1=1", "All Cases"),
        "active": ("(closure_status IS NULL OR closure_status='') AND status <> 'Closed'", "Active Cases"),
        "open": ("status='Open' AND (closure_status IS NULL OR closure_status='')", "Open Cases"),
        "closed": ("((closure_status IS NOT NULL AND closure_status <> '') OR status='Closed')", "Closed Cases"),
        "ongoing": ("(status IS NULL OR status <> 'Closed')", "Total Job Cards"),
        "pending": (
            "(status IS NULL OR status <> 'Closed') AND COALESCE(TRIM(status), '') NOT IN ('Ready-Waiting for Customer', 'Non-Repairable Waiting for Customer')",
            "Pending Jobs",
        ),
        "completed": (
            "(status IS NULL OR status <> 'Closed') AND COALESCE(TRIM(status), '') IN ('Ready-Waiting for Customer', 'Non-Repairable Waiting for Customer')",
            "Completed Jobs",
        ),
    }
    current_view = (raw_view or "active").strip().lower()
    if current_view not in view_map:
        current_view = "active"
    filter_sql, view_title = view_map[current_view]
    return current_view, filter_sql, view_title


def _build_jobs_listing_context(args, role, branch):
    current_view, filter_sql, view_title = _get_jobs_view_definition(args.get("view", "active"))
    search_query = (args.get("search") or "").strip()
    status_filters = _get_multi_values(args, "status")
    branch_filters = _get_multi_values(args, "branch_name")
    transfer_filters = [value.lower() for value in _get_multi_values(args, "transfer_filter")]
    age_bucket_filters = _get_multi_values(args, "age_bucket")
    call_type_filters = _get_multi_values(args, "call_type_filter")
    device_filters = _get_multi_values(args, "device_filter")
    priority_filters = _get_multi_values(args, "priority_filter")
    complaint_type_filters = _get_multi_values(args, "complaint_type_filter")
    engineer_filters = _get_multi_values(args, "engineer_filter")
    closure_results = [value.lower() for value in _get_multi_values(args, "closure_result")]
    chart_from_date = _normalize_date_input(args.get("chart_from_date", ""))
    chart_to_date = _normalize_date_input(args.get("chart_to_date", ""))
    default_chart_mode = "closed" if current_view == "closed" else "created"
    chart_mode = (args.get("chart_mode", default_chart_mode) or default_chart_mode).strip().lower()

    valid_buckets = ["0-2", "3-5", "6-10", "11-15", "16-30", "31-90", "91-180", "181+"]
    age_bucket_filters = [value for value in age_bucket_filters if value in valid_buckets]
    if chart_mode not in ["created", "closed"]:
        chart_mode = default_chart_mode
    transfer_filters = [value for value in transfer_filters if value == "active"]
    closure_results = [value for value in closure_results if value in ["success", "failed"]]

    where_clauses = [filter_sql]
    params = []

    if role in ["super_admin", "admin"] and branch == "ALL":
        branch_scope_clauses = []
        for branch_filter in branch_filters:
            branch_sql, branch_params = _build_job_transfer_branch_scope_clause(branch_filter)
            if branch_sql and branch_sql != "1=1":
                branch_scope_clauses.append(f"({branch_sql})")
                params.extend(branch_params)
        if branch_scope_clauses:
            where_clauses.append("(" + " OR ".join(branch_scope_clauses) + ")")
    else:
        branch_sql, branch_params = _build_job_transfer_branch_scope_clause(branch)
        where_clauses.append(branch_sql)
        params.extend(branch_params)

    if status_filters:
        where_clauses.append("status IN (" + ", ".join(["%s"] * len(status_filters)) + ")")
        params.extend(status_filters)

    if search_query:
        where_clauses.append(
            "(CAST(job_number AS CHAR) LIKE %s OR serial_number LIKE %s OR customer_name LIKE %s OR mobile LIKE %s OR alt_no LIKE %s)"
        )
        like_term = f"%{search_query}%"
        params.extend([like_term, like_term, like_term, like_term, like_term])

    exact_filters = [
        ("TRIM(call_type)", call_type_filters),
        ("device", device_filters),
        ("priority", priority_filters),
        ("complaint_type", complaint_type_filters),
    ]
    for column, values in exact_filters:
        if values:
            where_clauses.append(f"{column} IN (" + ", ".join(["%s"] * len(values)) + ")")
            params.extend(values)

    if engineer_filters:
        engineer_sql, engineer_params = _build_job_transfer_engineer_filter_clause(engineer_filters, "jobs")
        if engineer_sql:
            where_clauses.append(engineer_sql)
            params.extend(engineer_params)

    if "active" in transfer_filters:
        transfer_sql, transfer_params = _build_job_active_transfer_exists_clause("jobs")
        where_clauses.append(transfer_sql)
        params.extend(transfer_params)

    closure_clauses = []
    if "success" in closure_results:
        closure_clauses.append("LOWER(COALESCE(closure_status,'')) LIKE 'closed success%'")
    if "failed" in closure_results:
        closure_clauses.append("LOWER(COALESCE(closure_status,'')) LIKE 'closed failed%'")
    if closure_clauses:
        where_clauses.append("(" + " OR ".join(closure_clauses) + ")")

    if chart_from_date:
        date_col = "closure_date" if chart_mode == "closed" else "created_at"
        where_clauses.append(f"{date_col} >= %s")
        params.append(chart_from_date + " 00:00:00")

    if chart_to_date:
        date_col = "closure_date" if chart_mode == "closed" else "created_at"
        where_clauses.append(f"{date_col} <= %s")
        params.append(chart_to_date + " 23:59:59")

    return {
        "current_view": current_view,
        "view_title": view_title,
        "where_sql": " AND ".join(where_clauses),
        "params": tuple(params),
        "search_query": search_query,
        "status_filters": status_filters,
        "branch_filters": branch_filters,
        "transfer_filters": transfer_filters,
        "age_bucket_filters": age_bucket_filters,
        "call_type_filters": call_type_filters,
        "device_filters": device_filters,
        "priority_filters": priority_filters,
        "complaint_type_filters": complaint_type_filters,
        "engineer_filters": engineer_filters,
        "closure_results": closure_results,
        "chart_from_date": chart_from_date,
        "chart_to_date": chart_to_date,
        "chart_mode": chart_mode,
    }


def _get_jobs_filter_options(cursor, role, branch):
    scope_sql, scope_params = ("1=1", ())
    if not (role in ["super_admin", "admin"] and branch == "ALL"):
        scope_sql, scope_params = _build_job_transfer_branch_scope_clause(branch, "jobs")

    option_columns = {
        "status": "status",
        "branch": "branch_name",
        "device": "device",
        "priority": "priority",
        "call_type": "call_type",
        "complaint_type": "complaint_type",
        "engineer": "assigned_engineer",
    }
    options = {}
    for key, column in option_columns.items():
        cursor.execute(
            f"""
            SELECT DISTINCT TRIM(jobs.{column}) AS value
            FROM jobs
            WHERE {scope_sql}
              AND jobs.{column} IS NOT NULL
              AND TRIM(jobs.{column}) <> ''
            ORDER BY value
            """,
            scope_params,
        )
        options[key] = [row["value"] for row in cursor.fetchall() if row.get("value")]

    try:
        cursor.execute(
            f"""
            SELECT DISTINCT TRIM(transfer.specialist_engineer) AS value
            FROM job_service_transfers transfer
            JOIN jobs ON jobs.id=transfer.job_id
            WHERE {scope_sql}
              AND transfer.specialist_engineer IS NOT NULL
              AND TRIM(transfer.specialist_engineer) <> ''
            ORDER BY value
            """,
            scope_params,
        )
        specialist_engineers = [row["value"] for row in cursor.fetchall() if row.get("value")]
        options["engineer"] = sorted(set(options.get("engineer", []) + specialist_engineers), key=str.lower)
    except Exception:
        pass

    return options


def _annotate_job_rows(rows):
    now = business_now_naive()
    for job in rows:
        created_at = normalize_display_datetime(job.get("created_at"))
        closure_date = job.get("closure_date")
        age_days = None
        age_group = ""
        created_on = ""

        if created_at:
            age_days = (now.date() - created_at.date()).days
            if age_days < 0:
                age_days = 0
            age_group = get_age_group(age_days)
            created_on = format_datetime_display(created_at)

        job["age_days"] = age_days
        job["age_group"] = age_group
        job["created_on"] = created_on
        job["closed_on"] = format_datetime_display(closure_date)

    return rows


def _build_jobs_export_response(cursor, rows, filename_prefix):
    _decorate_job_rows_with_transfer_summary(cursor, rows)
    cursor.execute("SHOW COLUMNS FROM jobs")
    column_rows = cursor.fetchall()
    base_columns = [c.get("Field") for c in column_rows if c.get("Field")]

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        base_columns
        + [
            "created_on",
            "closed_on",
            "age",
            "age_group",
            "specialist_transfer_branch_name",
            "specialist_transfer_engineer",
            "specialist_transfer_status",
            "specialist_service_total",
            "closing_branch_service_margin",
        ]
    )

    for row in rows:
        output_row = [row.get(col, "") for col in base_columns]
        output_row.extend(
            [
                row.get("created_on", ""),
                row.get("closed_on", ""),
                row.get("age_days", ""),
                row.get("age_group", ""),
                row.get("specialist_transfer_branch_name", ""),
                row.get("specialist_transfer_engineer", ""),
                row.get("specialist_transfer_status", ""),
                row.get("specialist_service_total", ""),
                row.get("closing_branch_service_margin", ""),
            ]
        )
        writer.writerow(output_row)

    filename = f"{filename_prefix}_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------- JOB LIST ---------------- #

@app.route("/jobs")
def jobs():

    if "username" not in session:
        return redirect("/login")

    try:

        db = get_db()
        cursor = db.cursor(dictionary=True)

        branch = session.get("branch")
        role = session.get("role")
        jobs_context = _build_jobs_listing_context(request.args, role, branch)
        query = f"SELECT * FROM jobs WHERE {jobs_context['where_sql']} ORDER BY id DESC"
        cursor.execute(query, jobs_context["params"])

        jobs_list = cursor.fetchall()
        _annotate_job_rows(jobs_list)

        if jobs_context["age_bucket_filters"]:
            jobs_list = [job for job in jobs_list if job.get("age_group") in jobs_context["age_bucket_filters"]]

        # Pagination (20 rows per page)
        try:
            page = int(request.args.get("page", 1))
        except (ValueError, TypeError):
            page = 1
        per_page = 20
        total_count = len(jobs_list)
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))
        jobs_list = jobs_list[(page - 1) * per_page : page * per_page]
        _decorate_job_rows_with_transfer_summary(cursor, jobs_list)

        applied_filters = []
        if jobs_context["branch_filters"] and role in ["super_admin", "admin"] and branch == "ALL":
            applied_filters.append("Branch: " + ", ".join(jobs_context["branch_filters"]))
        if jobs_context["status_filters"]:
            applied_filters.append("Status: " + ", ".join(jobs_context["status_filters"]))
        if "active" in jobs_context["transfer_filters"]:
            applied_filters.append("Transfer: Active")
        if jobs_context["age_bucket_filters"]:
            applied_filters.append("Age: " + ", ".join(f"{value} days" for value in jobs_context["age_bucket_filters"]))
        if jobs_context["call_type_filters"]:
            applied_filters.append("Call Type: " + ", ".join(jobs_context["call_type_filters"]))
        if jobs_context["device_filters"]:
            applied_filters.append("Device: " + ", ".join(jobs_context["device_filters"]))
        if jobs_context["priority_filters"]:
            applied_filters.append("Priority: " + ", ".join(jobs_context["priority_filters"]))
        if jobs_context["complaint_type_filters"]:
            applied_filters.append("Complaint Type: " + ", ".join(jobs_context["complaint_type_filters"]))
        if jobs_context["engineer_filters"]:
            applied_filters.append("Engineer: " + ", ".join(jobs_context["engineer_filters"]))
        if "success" in jobs_context["closure_results"]:
            applied_filters.append("Closure: Closed Success")
        if "failed" in jobs_context["closure_results"]:
            applied_filters.append("Closure: Closed Failed")
        if jobs_context["chart_from_date"] or jobs_context["chart_to_date"]:
            applied_filters.append(f"Date: {jobs_context['chart_from_date'] or '-'} to {jobs_context['chart_to_date'] or '-'}")
        if jobs_context["search_query"]:
            applied_filters.append(f"Search: {jobs_context['search_query']}")

        if applied_filters:
            jobs_context["view_title"] = f"{jobs_context['view_title']} | " + " | ".join(applied_filters)

        jobs_filter_options = _get_jobs_filter_options(cursor, role, branch)
        def build_jobs_url(endpoint="jobs", **updates):
            query_args = request.args.to_dict(flat=False)
            query_args.pop("page", None)
            for key, value in updates.items():
                if value is None:
                    query_args.pop(key, None)
                elif isinstance(value, (list, tuple)):
                    query_args[key] = [str(item) for item in value if str(item)]
                else:
                    query_args[key] = [str(value)]
            query_string = urlencode(query_args, doseq=True)
            return ("/export-jobs" if endpoint == "export_jobs" else "/jobs") + (f"?{query_string}" if query_string else "")

        return render_template(
            "jobs.html",
            jobs=jobs_list,
            branch=branch,
            jobs_filter_options=jobs_filter_options,
            build_jobs_url=build_jobs_url,
            current_view=jobs_context["current_view"],
            view_title=jobs_context["view_title"],
            status_filters=jobs_context["status_filters"],
            branch_filters=jobs_context["branch_filters"],
            transfer_filters=jobs_context["transfer_filters"],
            age_bucket_filters=jobs_context["age_bucket_filters"],
            call_type_filters=jobs_context["call_type_filters"],
            device_filters=jobs_context["device_filters"],
            priority_filters=jobs_context["priority_filters"],
            complaint_type_filters=jobs_context["complaint_type_filters"],
            engineer_filters=jobs_context["engineer_filters"],
            closure_results=jobs_context["closure_results"],
            chart_from_date=jobs_context["chart_from_date"],
            chart_to_date=jobs_context["chart_to_date"],
            chart_mode=jobs_context["chart_mode"],
            search_query=jobs_context["search_query"],
            page=page,
            total_pages=total_pages,
            total_count=total_count,
        )

    except Error as e:
        _flash_internal_error("Error loading jobs", e)
        return redirect("/login")

    finally:
        _safe_close(cursor, db)


@app.route("/delete-job/<int:job_id>", methods=["POST"])
def delete_job(job_id):

    if "username" not in session:
        return redirect("/login")

    role = str(session.get("role") or "").strip().lower()
    if role != "super_admin":
        flash("Only super admin can permanently delete cases", "danger")
        return redirect(_get_safe_referrer_path() or "/jobs")

    return_to = (request.form.get("return_to") or "").strip()
    if not return_to.startswith("/jobs"):
        return_to = "/jobs?view=all"

    db = None
    cursor = None
    filenames_to_remove = []

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, role, session.get("branch"))
        if not job:
            flash("Case not found", "danger")
            return redirect(return_to)
        if not _branch_in_scope(role, session.get("branch"), job.get("branch_name")):
            flash("Only the owning branch can delete this case", "danger")
            return redirect(return_to)

        filenames_to_remove = _job_all_attachment_filenames(cursor, job)
        deleted_rows = _delete_job_record(cursor, job_id)
        if not deleted_rows:
            db.rollback()
            flash("Case not found", "danger")
            return redirect(return_to)

        db.commit()
        _remove_saved_images(filenames_to_remove, app.config["JOB_PHOTO_FOLDER"])
        flash(f"Case {job.get('job_number') or job_id} deleted permanently", "success")
    except Error as e:
        if db is not None:
            db.rollback()
        _flash_internal_error("Failed to delete case", e)
    finally:
        _safe_close(cursor, db)

    return redirect(return_to)


@app.route("/export-jobs")
def export_jobs():

    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        branch = session.get("branch")
        role = session.get("role")
        jobs_context = _build_jobs_listing_context(request.args, role, branch)

        cursor.execute(
            f"SELECT * FROM jobs WHERE {jobs_context['where_sql']} ORDER BY id DESC",
            jobs_context["params"],
        )
        rows = cursor.fetchall()
        _annotate_job_rows(rows)

        if jobs_context["age_bucket_filters"]:
            rows = [row for row in rows if row.get("age_group") in jobs_context["age_bucket_filters"]]

        return _build_jobs_export_response(cursor, rows, f"{jobs_context['current_view']}_jobs")

    except Error as e:
        _flash_internal_error("Job export failed", e)
        return redirect("/jobs")

    finally:
        _safe_close(cursor, db)


@app.route("/export-cases/<view_name>")
def export_cases(view_name):

    if "username" not in session:
        return redirect("/login")

    allowed_views = ["total", "active", "open", "closed"]
    if view_name not in allowed_views:
        flash("Invalid export type", "danger")
        return redirect("/dashboard")

    role = session.get("role")
    branch = session.get("branch")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        # Build filter-aware WHERE clause from query params
        dashboard_filters = build_dashboard_filters(request.args, role, branch)
        scope_sql = dashboard_filters["where_sql"]
        params = list(dashboard_filters["params"])

        if view_name == "total":
            view_sql = "1=1"
        elif view_name == "active":
            view_sql = "(closure_status IS NULL OR closure_status='') AND status <> 'Closed'"
        elif view_name == "open":
            view_sql = "status='Open' AND (closure_status IS NULL OR closure_status='')"
        else:  # closed
            view_sql = "((closure_status IS NOT NULL AND closure_status <> '') OR status='Closed')"

        cursor.execute(
            f"SELECT * FROM jobs WHERE {scope_sql} AND {view_sql} ORDER BY id DESC",
            tuple(params),
        )
        rows = cursor.fetchall()
        _annotate_job_rows(rows)
        return _build_jobs_export_response(cursor, rows, f"{view_name}_cases")

    except Error as e:
        _flash_internal_error("Export failed", e)
        return redirect("/dashboard")

    finally:
        _safe_close(cursor, db)


def _normalize_used_spares_view(value):
    normalized = str(value or "pending").strip().lower()
    if normalized in {"pending", "billed", "closed"}:
        return normalized
    return "pending"


@app.route("/used-spares")
def used_spares_page():

    if "username" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        role = session.get("role")
        branch = session.get("branch")
        filter_branch = str(request.args.get("filter_branch", "") or "").strip()

        from_date = _normalize_date_input(request.args.get("from_date", ""))
        to_date = _normalize_date_input(request.args.get("to_date", ""))
        current_view = _normalize_used_spares_view(request.args.get("view", "pending"))

        where = ["1=1"]
        params = []
        invoice_present_sql = "(j.spares_invoice_no IS NOT NULL AND TRIM(j.spares_invoice_no) <> '')"
        closed_sql = "((j.closure_status IS NOT NULL AND j.closure_status <> '') OR j.status='Closed')"

        if role in ["super_admin", "admin"] and branch == "ALL" and filter_branch:
            where.append("j.branch_name=%s")
            params.append(filter_branch)
        elif not (role in ["super_admin", "admin"] and branch == "ALL"):
            where.append("j.branch_name=%s")
            params.append(branch)

        if from_date:
            where.append("j.created_at >= %s")
            params.append(from_date + " 00:00:00")
        if to_date:
            where.append("j.created_at <= %s")
            params.append(to_date + " 23:59:59")

        if current_view == "pending":
            where.append(f"NOT {invoice_present_sql}")
            where.append(f"NOT {closed_sql}")
        elif current_view == "billed":
            where.append(invoice_present_sql)
            where.append(f"NOT {closed_sql}")
        else:
            where.append(invoice_present_sql)
            where.append(closed_sql)

        cursor.execute(
            f"""
            SELECT
                j.id AS job_id,
                j.job_number,
                j.customer_name,
                j.branch_name,
                j.spares_billing_status,
                j.spares_invoice_no,
                j.spares_invoice_date,
                j.spares_billed_by,
                j.closure_date,
                us.spare_name,
                us.amount,
                us.created_at
            FROM used_spares us
            JOIN jobs j ON j.id = us.job_id
            WHERE {' AND '.join(where)}
            ORDER BY j.job_number DESC, us.id ASC
            """,
            tuple(params),
        )
        rows = cursor.fetchall()

        for row in rows:
            row["spares_billing_status_display"] = _normalize_spares_billing_status(row.get("spares_billing_status"), True)
            row["spares_invoice_date_display"] = format_date_display(row.get("spares_invoice_date"))
            row["closed_on"] = format_datetime_display(row.get("closure_date"))

        total_amount = sum(float(r.get("amount") or 0) for r in rows)

        return render_template(
            "used_spares.html",
            rows=rows,
            total_amount=total_amount,
            branch=branch,
            filter_branch=filter_branch,
            from_date=from_date,
            to_date=to_date,
            current_view=current_view,
        )

    except Error as e:
        _flash_internal_error("Error loading used spare report", e)
        return redirect("/dashboard")

    finally:
        _safe_close(cursor, db)


@app.route("/used-spares/export")
def used_spares_export_page():

    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        role = session.get("role")
        branch = session.get("branch")
        filter_branch = str(request.args.get("filter_branch", "") or "").strip()

        from_date = _normalize_date_input(request.args.get("from_date", ""))
        to_date = _normalize_date_input(request.args.get("to_date", ""))
        current_view = _normalize_used_spares_view(request.args.get("view", "pending"))

        where = ["1=1"]
        params = []
        invoice_present_sql = "(j.spares_invoice_no IS NOT NULL AND TRIM(j.spares_invoice_no) <> '')"
        closed_sql = "((j.closure_status IS NOT NULL AND j.closure_status <> '') OR j.status='Closed')"

        if role in ["super_admin", "admin"] and branch == "ALL" and filter_branch:
            where.append("j.branch_name=%s")
            params.append(filter_branch)
        elif not (role in ["super_admin", "admin"] and branch == "ALL"):
            where.append("j.branch_name=%s")
            params.append(branch)

        if from_date:
            where.append("j.created_at >= %s")
            params.append(from_date + " 00:00:00")
        if to_date:
            where.append("j.created_at <= %s")
            params.append(to_date + " 23:59:59")

        if current_view == "pending":
            where.append(f"NOT {invoice_present_sql}")
            where.append(f"NOT {closed_sql}")
        elif current_view == "billed":
            where.append(invoice_present_sql)
            where.append(f"NOT {closed_sql}")
        else:
            where.append(invoice_present_sql)
            where.append(closed_sql)

        cursor.execute(
            f"""
            SELECT
                j.job_number,
                j.customer_name,
                j.branch_name,
                j.spares_billing_status,
                j.spares_invoice_no,
                j.spares_invoice_date,
                j.spares_billed_by,
                j.closure_date,
                us.spare_name,
                us.amount,
                us.created_at
            FROM used_spares us
            JOIN jobs j ON j.id = us.job_id
            WHERE {' AND '.join(where)}
            ORDER BY j.job_number DESC, us.id ASC
            """,
            tuple(params),
        )
        rows = cursor.fetchall()

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Job No", "Customer", "Branch", "Used Spare", "Amount", "Billing Status", "Invoice No", "Invoice Date", "Billed By", "Closed On", "Created On"])

        for row in rows:
            writer.writerow(
                [
                    row.get("job_number", ""),
                    row.get("customer_name", ""),
                    row.get("branch_name", ""),
                    row.get("spare_name", ""),
                    row.get("amount", ""),
                    _normalize_spares_billing_status(row.get("spares_billing_status"), True),
                    row.get("spares_invoice_no", ""),
                    format_date_display(row.get("spares_invoice_date")),
                    row.get("spares_billed_by", ""),
                    format_datetime_display(row.get("closure_date")),
                    format_datetime_display(row.get("created_at")),
                ]
            )

        filename = f"used_spares_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            buffer.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Error as e:
        _flash_internal_error("Used spare export failed", e)
        return redirect("/used-spares")

    finally:
        _safe_close(cursor, db)


def _get_locked_monthly_revenue_summary(cursor, from_date, to_date, filter_branch=""):
    where = [
        "entry_date >= %s",
        "entry_date <= %s",
        """
        EXISTS (
            SELECT 1
            FROM revenue_entry_period_locks AS lock_period
            WHERE branch_revenue_entries.entry_date >= lock_period.from_date
              AND branch_revenue_entries.entry_date <= lock_period.to_date
        )
        """,
    ]
    params = [from_date, to_date]

    if filter_branch and filter_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(filter_branch)

    cursor.execute(
        f"""
            SELECT
                YEAR(entry_date) AS year_num,
                MONTH(entry_date) AS month_num,
                SUM(COALESCE(sales_profit, 0)) AS total_sales,
                SUM(COALESCE(service_charges, 0)) AS total_service,
                SUM(COALESCE(total_profit, 0)) AS total_profit,
                COUNT(*) AS entry_count
            FROM branch_revenue_entries
            WHERE {' AND '.join(where)}
            GROUP BY YEAR(entry_date), MONTH(entry_date)
            ORDER BY YEAR(entry_date), MONTH(entry_date)
        """,
        tuple(params),
    )

    monthly_rows = cursor.fetchall()
    year_count = len({int(row.get("year_num") or 0) for row in monthly_rows if row.get("year_num") is not None})
    target_value, target_scope_label = _get_revenue_target_for_scope(cursor, filter_branch)

    labels = []
    full_labels = []
    sales_series = []
    service_series = []
    total_series = []
    target_series = []
    normalized_rows = []

    for row in monthly_rows:
        year_num = int(row.get("year_num") or 0)
        month_num = int(row.get("month_num") or 0)
        month_key = f"{year_num:04d}-{month_num:02d}-01" if year_num and month_num else ""
        try:
            month_value = datetime.strptime(month_key, "%Y-%m-%d")
        except (TypeError, ValueError):
            month_value = None

        if month_value:
            label = month_value.strftime("%b %Y") if year_count > 1 else month_value.strftime("%b")
            full_label = month_value.strftime("%B %Y")
        else:
            label = month_key or "Month"
            full_label = label

        sales_value = round(float(row.get("total_sales") or 0), 2)
        service_value = round(float(row.get("total_service") or 0), 2)
        total_value = round(float(row.get("total_profit") or 0), 2)
        entry_count = int(row.get("entry_count") or 0)

        labels.append(label)
        full_labels.append(full_label)
        sales_series.append(sales_value)
        service_series.append(service_value)
        total_series.append(total_value)
        target_series.append(target_value)
        normalized_rows.append(
            {
                "label": label,
                "full_label": full_label,
                "sales_profit": sales_value,
                "service_charges": service_value,
                "total_profit": total_value,
                "target_value": target_value,
                "entry_count": entry_count,
            }
        )

    return {
        "locked_month_labels": labels,
        "locked_month_full_labels": full_labels,
        "locked_sales_series": sales_series,
        "locked_service_series": service_series,
        "locked_total_series": total_series,
        "locked_target_series": target_series,
        "locked_month_rows": normalized_rows,
        "locked_month_count": len(normalized_rows),
        "locked_total_sales": round(sum(sales_series), 2),
        "locked_total_service": round(sum(service_series), 2),
        "locked_total_profit": round(sum(total_series), 2),
        "locked_target_value": target_value,
        "locked_target_scope_label": target_scope_label,
        "locked_chart_has_data": len(normalized_rows) > 0,
    }


def _get_default_locked_revenue_range(cursor, filter_branch=""):
    where = [
        """
        EXISTS (
            SELECT 1
            FROM revenue_entry_period_locks AS lock_period
            WHERE branch_revenue_entries.entry_date >= lock_period.from_date
              AND branch_revenue_entries.entry_date <= lock_period.to_date
        )
        """,
    ]
    params = []

    if filter_branch and filter_branch.upper() != "ALL":
        where.append("branch_name=%s")
        params.append(filter_branch)

    cursor.execute(
        f"""
            SELECT
                YEAR(entry_date) AS year_num,
                MONTH(entry_date) AS month_num,
                MIN(entry_date) AS min_entry_date,
                MAX(entry_date) AS max_entry_date
            FROM branch_revenue_entries
            WHERE {' AND '.join(where)}
            GROUP BY YEAR(entry_date), MONTH(entry_date)
            ORDER BY YEAR(entry_date) DESC, MONTH(entry_date) DESC
            LIMIT 12
        """,
        tuple(params),
    )

    month_rows = cursor.fetchall()
    if not month_rows:
        return "", ""

    sorted_rows = sorted(
        month_rows,
        key=lambda row: (int(row.get("year_num") or 0), int(row.get("month_num") or 0)),
    )
    start_date = sorted_rows[0].get("min_entry_date")
    end_date = sorted_rows[-1].get("max_entry_date")

    start_text = start_date.strftime("%Y-%m-%d") if hasattr(start_date, "strftime") else str(start_date or "").strip()
    end_text = end_date.strftime("%Y-%m-%d") if hasattr(end_date, "strftime") else str(end_date or "").strip()
    return start_text, end_text


def _get_revenue_view_context(args, can_manage_revenue, session_branch, include_edit_row=False, include_locked_periods=False, recent_limit=None, force_global_scope=False, allow_default_dates=True):
    today = business_now_naive().date()
    first_day = today.replace(day=1)
    requested_from_date = _normalize_date_input(args.get("from_date", ""))
    requested_to_date = _normalize_date_input(args.get("to_date", ""))
    filter_branch = (args.get("filter_branch") or "").strip()
    edit_id = (args.get("edit_id") or "").strip() if include_edit_row else ""

    has_global_scope = can_manage_revenue and (force_global_scope or (session_branch or "").strip().upper() == "ALL")
    if not has_global_scope:
        if not session_branch or session_branch.upper() == "ALL":
            raise PermissionError("Access Denied")
        filter_branch = session_branch
        edit_id = ""

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        if allow_default_dates and not requested_from_date and not requested_to_date:
            default_from_date, default_to_date = _get_default_locked_revenue_range(cursor, filter_branch)
        else:
            default_from_date, default_to_date = "", ""

        from_date = requested_from_date or default_from_date or (first_day.strftime("%Y-%m-%d") if allow_default_dates else "")
        to_date = requested_to_date or default_to_date or (today.strftime("%Y-%m-%d") if allow_default_dates else "")

        edit_row = {
            "entry_date": today.strftime("%Y-%m-%d"),
            "branch_name": "",
            "sales_profit": 0,
            "service_charges": 0,
            "total_profit": 0,
            "zone": "",
        }
        if include_edit_row and edit_id.isdigit():
            edit_sql = """
                SELECT id, entry_date, branch_name, sales_profit, service_charges, total_profit, zone
                FROM branch_revenue_entries
                WHERE id=%s
            """
            edit_params = [int(edit_id)]
            if not has_global_scope:
                edit_sql += " AND branch_name=%s"
                edit_params.append(session_branch)
            cursor.execute(edit_sql, tuple(edit_params))
            found_edit_row = cursor.fetchone()
            if found_edit_row:
                edit_row = found_edit_row
                if edit_row.get("entry_date") and hasattr(edit_row.get("entry_date"), "strftime"):
                    edit_row["entry_date"] = edit_row["entry_date"].strftime("%Y-%m-%d")

        where = ["1=1"]
        params = []
        if from_date:
            where.append("entry_date >= %s")
            params.append(from_date)
        if to_date:
            where.append("entry_date <= %s")
            params.append(to_date)
        if filter_branch and filter_branch.upper() != "ALL":
            where.append("branch_name=%s")
            params.append(filter_branch)

        select_sql = f"""
            SELECT id, entry_date, branch_name, sales_profit, service_charges, total_profit, zone, created_by, updated_at
            FROM branch_revenue_entries
            WHERE {' AND '.join(where)}
            ORDER BY entry_date DESC, branch_name ASC
        """
        select_params = list(params)
        if recent_limit and int(recent_limit) > 0:
            select_sql += " LIMIT %s"
            select_params.append(int(recent_limit))

        cursor.execute(select_sql, tuple(select_params))
        rows = cursor.fetchall()
        for row in rows:
            row["updated_on"] = format_datetime_display(row.get("updated_at"))

        total_sales = sum(float(r.get("sales_profit") or 0) for r in rows)
        total_service = sum(float(r.get("service_charges") or 0) for r in rows)
        total_profit = sum(float(r.get("total_profit") or 0) for r in rows)
        locked_chart_context = {
            "locked_month_labels": [],
            "locked_month_full_labels": [],
            "locked_sales_series": [],
            "locked_service_series": [],
            "locked_total_series": [],
            "locked_target_series": [],
            "locked_month_rows": [],
            "locked_month_count": 0,
            "locked_total_sales": 0,
            "locked_total_service": 0,
            "locked_total_profit": 0,
            "locked_target_value": 0,
            "locked_target_scope_label": filter_branch or session_branch or "All Branches",
            "locked_chart_has_data": False,
        }
        if from_date and to_date:
            locked_chart_context = _get_locked_monthly_revenue_summary(cursor, from_date, to_date, filter_branch)

        if has_global_scope:
            cursor.execute(
                "SELECT DISTINCT branch_name FROM branch_revenue_entries WHERE branch_name IS NOT NULL ORDER BY branch_name"
            )
            branch_options = [r.get("branch_name") for r in cursor.fetchall() if r.get("branch_name")]
            for b in _load_known_branches(cursor):
                if b and b != "ALL" and b not in branch_options:
                    branch_options.append(b)
            branch_options = sorted(set(branch_options), key=lambda x: x.upper())
        else:
            branch_options = [session_branch]

        locked_periods = []
        if can_manage_revenue and include_locked_periods:
            cursor.execute(
                "SELECT from_date, to_date, locked_by, locked_at FROM revenue_entry_period_locks ORDER BY from_date DESC"
            )
            locked_periods = cursor.fetchall()
            for lp in locked_periods:
                if lp.get("from_date") and hasattr(lp["from_date"], "strftime"):
                    lp["from_date"] = lp["from_date"].strftime("%Y-%m-%d")
                if lp.get("to_date") and hasattr(lp["to_date"], "strftime"):
                    lp["to_date"] = lp["to_date"].strftime("%Y-%m-%d")
                if lp.get("locked_at") and hasattr(lp["locked_at"], "strftime"):
                    lp["locked_at"] = lp["locked_at"].strftime("%d-%m-%Y %H:%M")

        return {
            "rows": rows,
            "from_date": from_date,
            "to_date": to_date,
            "filter_branch": filter_branch,
            "branch_options": branch_options,
            "total_sales": total_sales,
            "total_service": total_service,
            "total_profit": total_profit,
            "default_entry_date": today.strftime("%Y-%m-%d"),
            "edit_row": edit_row,
            "locked_periods": locked_periods,
            **locked_chart_context,
        }
    finally:
        _safe_close(cursor, db)


def _get_revenue_reports_redirect_target():
    return_to = (request.form.get("return_to") or "").strip()
    if return_to.startswith("/revenue-reports"):
        return return_to
    return _get_safe_referrer_path() or "/revenue-reports"


def _delete_revenue_entries(cursor, entry_ids, session_branch, has_global_scope):
    normalized_ids = []
    seen_ids = set()
    for raw_entry_id in entry_ids or ():
        entry_text = str(raw_entry_id or "").strip()
        if not entry_text.isdigit():
            continue
        entry_id = int(entry_text)
        if entry_id <= 0 or entry_id in seen_ids:
            continue
        normalized_ids.append(entry_id)
        seen_ids.add(entry_id)

    if not normalized_ids:
        return {"requested": 0, "found": 0, "deleted": 0, "locked": 0, "missing": 0}

    placeholders = ", ".join(["%s"] * len(normalized_ids))
    select_sql = f"""
        SELECT bre.id,
               bre.entry_date,
               EXISTS(
                   SELECT 1
                   FROM revenue_entry_period_locks AS lock_period
                   WHERE lock_period.from_date <= bre.entry_date
                     AND lock_period.to_date >= bre.entry_date
               ) AS is_locked
        FROM branch_revenue_entries AS bre
        WHERE bre.id IN ({placeholders})
    """
    select_params = list(normalized_ids)
    if not has_global_scope:
        select_sql += " AND bre.branch_name=%s"
        select_params.append(session_branch)

    cursor.execute(select_sql, tuple(select_params))
    matching_rows = cursor.fetchall() or []

    found_ids = {int(row.get("id") or 0) for row in matching_rows if int(row.get("id") or 0) > 0}
    deletable_ids = []
    locked_count = 0
    for row in matching_rows:
        if int(row.get("is_locked") or 0) > 0:
            locked_count += 1
            continue
        entry_id = int(row.get("id") or 0)
        if entry_id > 0:
            deletable_ids.append(entry_id)

    deleted_count = 0
    if deletable_ids:
        delete_placeholders = ", ".join(["%s"] * len(deletable_ids))
        delete_sql = f"DELETE FROM branch_revenue_entries WHERE id IN ({delete_placeholders})"
        delete_params = list(deletable_ids)
        if not has_global_scope:
            delete_sql += " AND branch_name=%s"
            delete_params.append(session_branch)
        cursor.execute(delete_sql, tuple(delete_params))
        deleted_count = max(int(cursor.rowcount or 0), 0)

    return {
        "requested": len(normalized_ids),
        "found": len(found_ids),
        "deleted": deleted_count,
        "locked": locked_count,
        "missing": max(len(normalized_ids) - len(found_ids), 0),
    }


def _get_cashflow_view_context(args, role, session_branch):
    today = business_now_naive().date()
    first_day = today.replace(day=1)
    requested_from_date = _normalize_date_input(args.get("from_date", ""))
    requested_to_date = _normalize_date_input(args.get("to_date", ""))
    filter_branch = (args.get("filter_branch") or "").strip()
    has_global_scope = _user_has_all_branch_scope(role, session_branch)

    if has_global_scope and filter_branch.upper() == "ALL":
        filter_branch = ""
    if not has_global_scope:
        if not session_branch or session_branch.upper() == "ALL":
            raise PermissionError("Access Denied")
        filter_branch = session_branch

    from_date = requested_from_date or first_day.strftime("%Y-%m-%d")
    to_date = requested_to_date or today.strftime("%Y-%m-%d")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        rows = _fetch_closed_job_cashflow_rows(
            cursor,
            filter_branch=filter_branch,
            from_date=from_date,
            to_date=to_date,
        )
        for row in rows:
            row["updated_on"] = format_datetime_display(row.get("updated_at"))
        rows.sort(key=lambda row: str(row.get("branch_name") or "").upper())
        rows.sort(key=lambda row: str(row.get("updated_at") or ""), reverse=True)
        rows.sort(key=lambda row: str(row.get("entry_date") or ""), reverse=True)

        branch_summary_rows = _fetch_closed_job_cashflow_summary_rows(
            cursor,
            filter_branch=filter_branch,
            from_date=from_date,
            to_date=to_date,
        )

        approved_transfers_by_branch = _fetch_cashflow_transfer_totals_by_branch(
            cursor,
            filter_branch,
            from_date=from_date,
            to_date=to_date,
            statuses=["Approved"],
        )
        pending_transfers_by_branch = _fetch_cashflow_transfer_totals_by_branch(
            cursor,
            filter_branch,
            from_date=from_date,
            to_date=to_date,
            statuses=["Pending"],
        )
        balance_cash_map = _fetch_cashflow_cash_totals_by_branch(cursor, filter_branch, to_date)
        balance_approved_map = _fetch_cashflow_transfer_totals_by_branch(
            cursor,
            filter_branch,
            to_date=to_date,
            statuses=["Approved"],
        )

        for row in branch_summary_rows:
            branch_name = str(row.get("branch_name") or "").strip()
            gross_cash = float(row.get("cash_total") or 0)
            approved_amount = float(approved_transfers_by_branch.get(branch_name) or 0)
            pending_amount = float(pending_transfers_by_branch.get(branch_name) or 0)
            cash_balance = float(balance_cash_map.get(branch_name) or 0) - float(balance_approved_map.get(branch_name) or 0)
            available_cash = max(cash_balance, 0.0)
            row["gross_cash_total"] = round(gross_cash, 2)
            row["cash_balance"] = round(available_cash, 2)
            row["approved_transfer_total"] = round(approved_amount, 2)
            row["pending_transfer_total"] = round(pending_amount, 2)
            row["available_cash"] = round(available_cash, 2)

        transfer_where = ["1=1"]
        transfer_params = []
        if from_date:
            transfer_where.append("request_date >= %s")
            transfer_params.append(from_date)
        if to_date:
            transfer_where.append("request_date <= %s")
            transfer_params.append(to_date)
        if filter_branch:
            transfer_where.append("branch_name=%s")
            transfer_params.append(filter_branch)

        cursor.execute(
            f"""
            SELECT
                id,
                branch_name,
                request_date,
                amount,
                transfer_to,
                requested_notes,
                status,
                requested_by,
                requested_at,
                reviewed_by,
                reviewed_at,
                review_notes,
                transfer_reference
            FROM branch_cash_transfer_requests
            WHERE {' AND '.join(transfer_where)}
            ORDER BY CASE status WHEN 'Pending' THEN 0 WHEN 'Approved' THEN 1 ELSE 2 END, request_date DESC, id DESC
            LIMIT 200
            """,
            tuple(transfer_params),
        )
        transfer_rows = cursor.fetchall()
        for row in transfer_rows:
            row["requested_on"] = format_datetime_display(row.get("requested_at"))
            row["reviewed_on"] = format_datetime_display(row.get("reviewed_at"))

        if has_global_scope:
            branch_options = _load_known_branches(cursor)
        else:
            branch_options = [session_branch]

        total_cash = round(sum(float(row.get("cash_balance") or 0) for row in branch_summary_rows), 2)
        total_card = round(sum(float(row.get("card_total") or 0) for row in branch_summary_rows), 2)
        total_upi = round(sum(float(row.get("upi_total") or 0) for row in branch_summary_rows), 2)
        total_collected = round(sum(float(row.get("total_collected") or 0) for row in branch_summary_rows), 2)
        total_gross_cash = round(sum(float(row.get("gross_cash_total") or 0) for row in branch_summary_rows), 2)
        total_available_cash = round(sum(float(row.get("available_cash") or 0) for row in branch_summary_rows), 2)

        overall_pending_map = _fetch_cashflow_transfer_totals_by_branch(cursor, filter_branch, statuses=["Pending"])
        pending_transfer_total = round(sum(float(amount or 0) for amount in overall_pending_map.values()), 2)

        transfer_request_branch_rows = []
        if has_global_scope and _can_request_cash_transfer(role):
            scoped_branch_names = _load_known_branches(cursor)
            if filter_branch:
                scoped_branch_names = [filter_branch]

            for branch_name in scoped_branch_names:
                normalized_branch = str(branch_name or "").strip()
                if not normalized_branch:
                    continue

                available_cash = round(
                max(
                    float(balance_cash_map.get(normalized_branch) or 0)
                    - float(balance_approved_map.get(normalized_branch) or 0),
                    0.0,
                ),
                )
                pending_amount = round(float(overall_pending_map.get(normalized_branch) or 0), 2)
                requestable_cash = round(max(available_cash - pending_amount, 0), 2)

                if available_cash <= 0 and pending_amount <= 0:
                    continue

                transfer_request_branch_rows.append(
                    {
                        "branch_name": normalized_branch,
                        "available_cash": available_cash,
                        "pending_transfer_total": pending_amount,
                        "requestable_cash": requestable_cash,
                    }
                )

            transfer_request_branch_rows.sort(key=lambda row: str(row.get("branch_name") or "").upper())

        return {
            "rows": rows,
            "branch_summary_rows": branch_summary_rows,
            "transfer_rows": transfer_rows,
            "transfer_request_branch_rows": transfer_request_branch_rows,
            "from_date": from_date,
            "to_date": to_date,
            "filter_branch": filter_branch,
            "branch_options": branch_options,
            "today_date": today.strftime("%Y-%m-%d"),
            "total_cash": total_cash,
            "total_card": total_card,
            "total_upi": total_upi,
            "total_collected": total_collected,
            "total_gross_cash": total_gross_cash,
            "total_available_cash": total_available_cash,
            "pending_transfer_total": pending_transfer_total,
            "has_global_scope": has_global_scope,
            "can_request_transfer": _can_request_cash_transfer(role),
            "can_approve_transfer": _can_approve_cash_transfer(role),
            "can_cancel_transfer": _can_cancel_cash_transfer(role),
        }
    finally:
        _safe_close(cursor, db)


@app.route("/revenue-dashboard")
def revenue_dashboard_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=False,
            recent_limit=None,
            force_global_scope=can_manage_revenue,
        )
        return render_template("revenue_dashboard.html", can_manage_revenue=can_manage_revenue, **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue dashboard error", e)
        return redirect("/dashboard")


@app.route("/revenue-entry", methods=["GET", "POST"])
def revenue_entry_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    has_global_scope = can_manage_revenue and session_branch.upper() == "ALL"

    if not can_manage_revenue:
        flash("Access limited: Revenue Dashboard only", "warning")
        return redirect("/revenue-dashboard")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        if request.method == "POST":
            if not can_manage_revenue:
                return "Access Denied"
            mode = (request.form.get("mode") or "bulk").strip().lower()
            entry_date = _normalize_date_input(request.form.get("entry_date", ""))
            if not entry_date:
                flash("Valid entry date is required", "danger")
                return redirect("/revenue-entry")

            rows_to_save = []
            if mode == "single":
                branch_name = (request.form.get("branch_name") or "").strip()
                if not has_global_scope:
                    branch_name = session_branch
                if not branch_name:
                    flash("Branch is required", "danger")
                    return redirect("/revenue-entry")

                sales = max(_parse_money(request.form.get("sales_profit")), 0)
                service = max(_parse_money(request.form.get("service_charges")), 0)
                total = _parse_money(request.form.get("total_profit"))
                if total <= 0:
                    total = sales + service
                zone = (request.form.get("zone") or "").strip()

                rows_to_save.append(
                    {
                        "branch_name": branch_name,
                        "sales_profit": sales,
                        "service_charges": service,
                        "total_profit": max(total, 0),
                        "zone": zone,
                    }
                )
            else:
                pasted_text = request.form.get("pasted_rows", "")
                rows_to_save = _parse_pasted_revenue_rows(pasted_text)
                if not rows_to_save:
                    flash("No valid rows found in pasted data", "warning")
                    return redirect("/revenue-entry")
                if not has_global_scope:
                    scoped_rows = []
                    for row in rows_to_save:
                        if (row.get("branch_name") or "").strip() != session_branch:
                            continue
                        scoped_rows.append(row)
                    rows_to_save = scoped_rows
                    if not rows_to_save:
                        flash("Pasted rows do not match your branch scope", "danger")
                        return redirect("/revenue-entry")

            cursor.execute(
                "SELECT COUNT(*) as cnt FROM revenue_entry_period_locks WHERE from_date <= %s AND to_date >= %s",
                (entry_date, entry_date),
            )
            lock_check = cursor.fetchone()
            is_locked = (lock_check.get("cnt") if isinstance(lock_check, dict) else lock_check[0]) > 0
            if is_locked:
                flash(f"Date {entry_date} is within a locked period. Unlock it first.", "danger")
                return redirect("/revenue-entry")

            for row in rows_to_save:
                cursor.execute(
                    """
                    INSERT INTO branch_revenue_entries
                    (entry_date, branch_name, sales_profit, service_charges, total_profit, zone, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        sales_profit=VALUES(sales_profit),
                        service_charges=VALUES(service_charges),
                        total_profit=VALUES(total_profit),
                        zone=VALUES(zone),
                        created_by=VALUES(created_by)
                    """,
                    (
                        entry_date,
                        row["branch_name"],
                        row["sales_profit"],
                        row["service_charges"],
                        row["total_profit"],
                        row.get("zone") or None,
                        session.get("username"),
                    ),
                )

            db.commit()
            flash(f"Revenue data saved: {len(rows_to_save)} row(s)", "success")
            return redirect(f"/revenue-entry?from_date={entry_date}&to_date={entry_date}")

        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=True,
            include_locked_periods=False,
            recent_limit=None,
        )
        return render_template("revenue_entry.html", can_manage_revenue=can_manage_revenue, **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue entry failed", e)
        return redirect("/dashboard")
    finally:
        _safe_close(cursor, db)


@app.route("/revenue-upload")
def revenue_upload_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    if not can_manage_revenue:
        flash("Access limited: Revenue Dashboard only", "warning")
        return redirect("/revenue-dashboard")
    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=False,
            recent_limit=20,
        )
        return render_template("revenue_upload.html", can_manage_revenue=can_manage_revenue, **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue upload page error", e)
        return redirect("/dashboard")


@app.route("/revenue-reports")
def revenue_reports_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    if not can_manage_revenue:
        flash("Access limited: Revenue Dashboard only", "warning")
        return redirect("/revenue-dashboard")
    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=False,
            recent_limit=None,
            allow_default_dates=False,
        )
        return render_template("revenue_reports.html", can_manage_revenue=can_manage_revenue, **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue reports error", e)
        return redirect("/dashboard")


@app.route("/revenue-reports/export")
def revenue_reports_export_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    if not can_manage_revenue:
        flash("Access limited: Revenue Dashboard only", "warning")
        return redirect("/revenue-dashboard")

    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=False,
            recent_limit=None,
            allow_default_dates=False,
        )

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Date", "Branch", "Sales Profit", "Service Charges", "Total Profit", "Zone", "Updated By", "Last Modified"])

        for row in context["rows"]:
            writer.writerow(
                [
                    row.get("entry_date", ""),
                    row.get("branch_name", ""),
                    row.get("sales_profit", ""),
                    row.get("service_charges", ""),
                    row.get("total_profit", ""),
                    row.get("zone", ""),
                    row.get("created_by", ""),
                    row.get("updated_on", ""),
                ]
            )

        filename = f"revenue_report_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            buffer.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue report export failed", e)
        return redirect("/revenue-reports")


@app.route("/admin-reports")
def admin_reports_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    if role not in ["super_admin", "admin"]:
        flash("Access denied", "danger")
        return redirect("/dashboard")

    try:
        context = _get_admin_reports_context(request.args, role, session_branch)
        return render_template("admin_reports.html", **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Admin reports error", e)
        return redirect("/dashboard")


@app.route("/admin-reports/export/<export_type>")
def admin_reports_export_page(export_type):
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    if role not in ["super_admin", "admin"]:
        flash("Access denied", "danger")
        return redirect("/dashboard")

    normalized_export_type = (export_type or "").strip().lower()
    if normalized_export_type not in {"snapshot", "branch", "engineer", "staff-closed", "staff-revenue"}:
        flash("Invalid report export option", "warning")
        return redirect("/admin-reports")

    try:
        context = _get_admin_reports_context(request.args, role, session_branch)

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        if normalized_export_type == "snapshot":
            writer.writerow(["Period", "Metric", "Count", "Amount"])

            snapshot_rows = [
                ("Today", "New Calls", context["today_jobs"]["new_calls"], ""),
                ("Today", "Closed Calls", context["today_jobs"]["closed_calls"], ""),
                ("Today", "Closed After 7+ Days", context["today_jobs"]["closed_after_7_days"], ""),
                ("Today", "SYSCARE", context["today_syscare"]["syscare_count"], context["today_syscare"]["syscare_amount"]),
                ("Today", "Service Charges", context["today_jobs"]["service_charge_count"], context["today_jobs"]["service_charge_total"]),
                ("Current Month", "New Calls", context["month_jobs"]["new_calls"], ""),
                ("Current Month", "Closed Calls", context["month_jobs"]["closed_calls"], ""),
                ("Current Month", "SYSCARE", context["month_syscare"]["syscare_count"], context["month_syscare"]["syscare_amount"]),
                ("Current Month", "Service Charges", context["month_jobs"]["service_charge_count"], context["month_jobs"]["service_charge_total"]),
                ("Selected Range", "New Calls", context["selected_jobs"]["new_calls"], ""),
                ("Selected Range", "Closed Calls", context["selected_jobs"]["closed_calls"], ""),
                ("Selected Range", "Closed After 7+ Days", context["selected_jobs"]["closed_after_7_days"], ""),
                ("Selected Range", "SYSCARE", context["selected_syscare"]["syscare_count"], context["selected_syscare"]["syscare_amount"]),
                ("Selected Range", "Service Charges", context["selected_jobs"]["service_charge_count"], context["selected_jobs"]["service_charge_total"]),
            ]

            for row in snapshot_rows:
                writer.writerow(row)

        elif normalized_export_type == "branch":
            writer.writerow(["Branch", "New Calls", "Closed Calls", "Closed 7+ Days", "SYSCARE Count", "SYSCARE Amount", "Service Charges"])
            for row in context["branch_summary_rows"]:
                writer.writerow(
                    [
                        row.get("branch_name", ""),
                        row.get("new_calls", 0),
                        row.get("closed_calls", 0),
                        row.get("closed_after_7_days", 0),
                        row.get("syscare_count", 0),
                        row.get("syscare_amount", 0),
                        row.get("service_charge_total", 0),
                    ]
                )
        elif normalized_export_type == "staff-closed":
            writer.writerow(["Engineer", "Employee Code", "Branch", "Total", "Open", "Closed Success", "Closed Failed", "Sales Revenue", "Service Charges", "Total Revenue"])
            for row in context.get("staff_closed_rank_rows", []):
                writer.writerow(
                    [
                        row.get("engineer_name", ""),
                        row.get("employee_code", ""),
                        row.get("branch_name", ""),
                        row.get("total_calls", 0),
                        row.get("open_calls", 0),
                        row.get("closed_success", 0),
                        row.get("closed_failed", 0),
                        row.get("sales_revenue", 0),
                        row.get("service_charges", 0),
                        row.get("total_revenue", 0),
                    ]
                )
        elif normalized_export_type == "staff-revenue":
            writer.writerow(["Engineer", "Employee Code", "Branch", "Total", "Open", "Closed Success", "Closed Failed", "Sales Revenue", "Service Charges", "Total Revenue"])
            for row in context.get("staff_revenue_rank_rows", []):
                writer.writerow(
                    [
                        row.get("engineer_name", ""),
                        row.get("employee_code", ""),
                        row.get("branch_name", ""),
                        row.get("total_calls", 0),
                        row.get("open_calls", 0),
                        row.get("closed_success", 0),
                        row.get("closed_failed", 0),
                        row.get("sales_revenue", 0),
                        row.get("service_charges", 0),
                        row.get("total_revenue", 0),
                    ]
                )
        else:
            writer.writerow(["Engineer", "New Calls", "Closed Calls", "Closed 7+ Days", "SYSCARE Count", "SYSCARE Amount", "Service Charges"])
            for row in context["engineer_summary_rows"]:
                writer.writerow(
                    [
                        row.get("engineer_name", ""),
                        row.get("new_calls", 0),
                        row.get("closed_calls", 0),
                        row.get("closed_after_7_days", 0),
                        row.get("syscare_count", 0),
                        row.get("syscare_amount", 0),
                        row.get("service_charge_total", 0),
                    ]
                )

        filename = f"admin_reports_{normalized_export_type}_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            buffer.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Admin report export failed", e)
        return redirect("/admin-reports")


@app.route("/staff-management")
def staff_management_page():
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    try:
        context = _get_staff_management_context(request.args)
        return render_template("staff_management.html", **context)
    except Error as e:
        _flash_internal_error("Staff management error", e)
        return redirect("/settings")


@app.route("/staff-management/export/<export_view>")
def staff_management_export_page(export_view):
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    ensure_staff_directory_table()
    normalized_export_view = (export_view or "").strip().lower()
    if normalized_export_view not in {"active", "resigned", "all"}:
        flash("Invalid staff export option", "warning")
        return redirect("/staff-management")

    try:
        context = _get_staff_management_context(request.args)

        if normalized_export_view == "active":
            export_rows = context["staff_rows"]
        elif normalized_export_view == "resigned":
            export_rows = context["resigned_staff_rows"]
        else:
            export_rows = list(context["staff_rows"]) + list(context["resigned_staff_rows"])

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "ID",
            "Name",
            "Contact Number",
            "Branch",
            "Salary",
            "ESI",
            "PF",
            "Room Rent",
            "Joined Date",
            "Resigned Date",
            "Total Payable",
            "Status",
        ])

        for row in export_rows:
            joined_date = row.get("joined_date")
            resigned_date = row.get("resigned_date")
            writer.writerow(
                [
                    row.get("id", ""),
                    row.get("staff_name", ""),
                    row.get("contact_number", ""),
                    row.get("branch_name", ""),
                    row.get("salary", 0),
                    row.get("esi", 0),
                    row.get("pf", 0),
                    row.get("room_rent", 0),
                    joined_date.strftime("%Y-%m-%d") if joined_date else "",
                    resigned_date.strftime("%Y-%m-%d") if resigned_date else "",
                    row.get("total_payable", 0),
                    "Resigned" if resigned_date else "Active",
                ]
            )

        filename = f"staff_management_{normalized_export_view}_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            buffer.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Error as e:
        _flash_internal_error("Staff export failed", e)
        return redirect("/staff-management")


@app.route("/staff-management/add", methods=["POST"])
def add_staff_member():
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    db = None
    cursor = None
    try:
        ensure_staff_directory_table()
        db = get_db()
        cursor = db.cursor(dictionary=True)
        payload, errors = _extract_staff_member_payload(cursor, request.form)
        if errors:
            for message in errors:
                flash(message, "danger")
            return redirect("/staff-management")

        cursor.execute(
            """
            INSERT INTO staff_directory (
                staff_name, contact_number, branch_name, salary, esi, pf, room, rent, joined_date, resigned_date
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                payload["staff_name"],
                payload["contact_number"],
                payload["branch_name"],
                payload["salary"],
                payload["esi"],
                payload["pf"],
                payload["room"],
                payload["rent"],
                payload["joined_date"],
                payload["resigned_date"],
            ),
        )
        db.commit()
        flash("Staff member added", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_staff_database_error("Failed to add staff member", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/staff-management")


@app.route("/staff-management/edit/<int:staff_id>", methods=["POST"])
def edit_staff_member(staff_id):
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    db = None
    cursor = None
    try:
        ensure_staff_directory_table()
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT id, resigned_date FROM staff_directory WHERE id=%s", (staff_id,))
        existing_row = cursor.fetchone()
        if not existing_row:
            flash("Staff member not found", "warning")
            return redirect("/staff-management")

        payload, errors = _extract_staff_member_payload(cursor, request.form)
        if errors:
            for message in errors:
                flash(message, "danger")
            return redirect("/staff-management")

        cursor.execute(
            """
            UPDATE staff_directory
            SET staff_name=%s,
                contact_number=%s,
                branch_name=%s,
                salary=%s,
                esi=%s,
                pf=%s,
                room=%s,
                rent=%s,
                joined_date=%s,
                resigned_date=%s
            WHERE id=%s
            """,
            (
                payload["staff_name"],
                payload["contact_number"],
                payload["branch_name"],
                payload["salary"],
                payload["esi"],
                payload["pf"],
                payload["room"],
                payload["rent"],
                payload["joined_date"],
                payload["resigned_date"],
                staff_id,
            ),
        )
        db.commit()
        was_resigned = bool(existing_row.get("resigned_date"))
        is_resigned = bool(payload.get("resigned_date"))
        if is_resigned and not was_resigned:
            flash("Staff member moved to resigned list", "success")
        elif was_resigned and not is_resigned:
            flash("Staff member moved back to active list", "success")
        else:
            flash("Staff member updated", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_staff_database_error("Failed to update staff member", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/staff-management")


@app.route("/staff-management/delete/<int:staff_id>", methods=["POST"])
def delete_staff_member(staff_id):
    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    db = None
    cursor = None
    try:
        ensure_staff_directory_table()
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("DELETE FROM staff_directory WHERE id=%s", (staff_id,))
        db.commit()
        flash("Staff member deleted", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_staff_database_error("Failed to delete staff member", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/staff-management")


@app.route("/cashflow", methods=["GET", "POST"])
def cashflow_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()

    db = None
    cursor = None
    try:
        if request.method == "POST":
            db = get_db()
            cursor = db.cursor(dictionary=True)
            form_action = (request.form.get("form_action") or "").strip().lower()
            return_to = (request.form.get("return_to") or "").strip()
            redirect_target = return_to if return_to.startswith("/cashflow") else "/cashflow"

            if form_action == "save_entry":
                flash("Manual cashflow entry is disabled. Closed job payments reflect here automatically.", "info")
                return redirect(redirect_target)

            if form_action == "request_transfer":
                if not _can_request_cash_transfer(role):
                    flash("Your role cannot create transfer requests", "danger")
                    return redirect(redirect_target)

                request_date = _normalize_date_input(request.form.get("request_date", "")) or business_now_naive().strftime("%Y-%m-%d")
                requested_branch = request.form.get("branch_name", "")
                branch_name = _resolve_branch_input(role, session_branch, requested_branch)
                branch_name = _resolve_known_branch(cursor, branch_name)
                amount = max(_parse_money(request.form.get("amount")), 0)
                transfer_to = (request.form.get("transfer_to") or "").strip()
                requested_notes = (request.form.get("requested_notes") or "").strip()

                if not branch_name or branch_name == "ALL":
                    flash("Valid branch is required for transfer request", "danger")
                    return redirect(redirect_target)
                if amount <= 0:
                    flash("Transfer amount must be greater than zero", "danger")
                    return redirect(redirect_target)

                available_cash = _get_branch_available_cash(cursor, branch_name, request_date)
                if amount > available_cash:
                    flash(f"Available cash for {branch_name} is only {available_cash:.2f}", "danger")
                    return redirect(redirect_target)

                cursor.execute(
                    """
                    INSERT INTO branch_cash_transfer_requests
                    (branch_name, request_date, amount, transfer_to, requested_notes, status, requested_by)
                    VALUES (%s, %s, %s, %s, %s, 'Pending', %s)
                    """,
                    (
                        branch_name,
                        request_date,
                        amount,
                        transfer_to or None,
                        requested_notes or None,
                        session.get("username"),
                    ),
                )
                db.commit()
                flash("Transfer request created", "success")
                return redirect(redirect_target)

            if form_action == "review_transfer":
                transfer_id = int(request.form.get("transfer_id") or 0)
                review_action = (request.form.get("review_action") or "").strip().lower()
                review_notes = (request.form.get("review_notes") or "").strip()
                transfer_reference = (request.form.get("transfer_reference") or "").strip()
                approval_amount_verification = _parse_money(request.form.get("approval_amount_verification"))

                if transfer_id <= 0 or review_action not in {"approve", "reject", "cancel"}:
                    flash("Invalid transfer review request", "danger")
                    return redirect(redirect_target)

                if review_action == "cancel":
                    if not _can_cancel_cash_transfer(role):
                        flash("Only admin and super admin can cancel transfer requests", "danger")
                        return redirect(redirect_target)
                elif not _can_approve_cash_transfer(role):
                    flash("Your role cannot review transfer requests", "danger")
                    return redirect(redirect_target)

                select_sql = """
                    SELECT id, branch_name, request_date, amount, status, requested_by
                    FROM branch_cash_transfer_requests
                    WHERE id=%s
                """
                select_params = [transfer_id]
                if not _user_has_all_branch_scope(role, session_branch):
                    select_sql += " AND branch_name=%s"
                    select_params.append(session_branch)
                cursor.execute(select_sql, tuple(select_params))
                transfer_row = cursor.fetchone()

                if not transfer_row:
                    flash("Transfer request not found", "danger")
                    return redirect(redirect_target)
                if str(transfer_row.get("status") or "").strip().lower() != "pending":
                    flash("Only pending transfer requests can be reviewed", "warning")
                    return redirect(redirect_target)

                if review_action != "cancel":
                    requested_by = str(transfer_row.get("requested_by") or "").strip().lower()
                    current_user = str(session.get("username") or "").strip().lower()
                    if requested_by and requested_by == current_user:
                        flash("Transfer requests must be reviewed by another user", "warning")
                        return redirect(redirect_target)

                status_map = {
                    "approve": "Approved",
                    "reject": "Rejected",
                    "cancel": "Cancelled",
                }
                next_status = status_map[review_action]
                if next_status == "Approved":
                    requested_amount = round(float(transfer_row.get("amount") or 0), 2)
                    if round(float(approval_amount_verification or 0), 2) != requested_amount:
                        flash(f"Type the exact transfer amount {requested_amount:.2f} to approve this request.", "danger")
                        return redirect(redirect_target)

                    request_date_text = (
                        transfer_row.get("request_date").strftime("%Y-%m-%d")
                        if hasattr(transfer_row.get("request_date"), "strftime")
                        else str(transfer_row.get("request_date") or "")
                    )
                    available_cash = _get_branch_available_cash(cursor, transfer_row.get("branch_name"), request_date_text)
                    if float(transfer_row.get("amount") or 0) > available_cash:
                        flash(
                            f"Available cash for {transfer_row.get('branch_name')} is only {available_cash:.2f}. Approval blocked.",
                            "danger",
                        )
                        return redirect(redirect_target)

                cursor.execute(
                    """
                    UPDATE branch_cash_transfer_requests
                    SET status=%s,
                        reviewed_by=%s,
                        reviewed_at=NOW(),
                        review_notes=%s,
                        transfer_reference=%s
                    WHERE id=%s
                    """,
                    (
                        next_status,
                        session.get("username"),
                        review_notes or None,
                        transfer_reference or None,
                        transfer_id,
                    ),
                )
                db.commit()
                flash(f"Transfer request {next_status.lower()}", "success")
                return redirect(redirect_target)

            flash("Unknown cashflow action", "warning")
            return redirect(redirect_target)

        context = _get_cashflow_view_context(request.args, role, session_branch)
        return render_template("cashflow.html", **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Cashflow operation failed", e)
        return redirect("/dashboard")
    finally:
        _safe_close(cursor, db)


@app.route("/cashflow/export")
def cashflow_export_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    try:
        context = _get_cashflow_view_context(request.args, role, session_branch)

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Branch Summary"])
        writer.writerow(["Branch", "Cash Balance", "Card", "UPI", "Collection", "Approved Transfer", "Pending Transfer", "Gross Cash"])
        for row in context["branch_summary_rows"]:
            writer.writerow(
                [
                    row.get("branch_name", ""),
                    row.get("cash_balance", ""),
                    row.get("card_total", ""),
                    row.get("upi_total", ""),
                    row.get("total_collected", ""),
                    row.get("approved_transfer_total", ""),
                    row.get("pending_transfer_total", ""),
                    row.get("gross_cash_total", ""),
                ]
            )

        writer.writerow([])
        writer.writerow(["Collections History"])
        writer.writerow(["Date", "Branch", "Cash", "Card", "UPI", "Total", "Remarks", "Updated By", "Last Modified"])
        for row in context["rows"]:
            writer.writerow(
                [
                    row.get("entry_date", ""),
                    row.get("branch_name", ""),
                    row.get("cash_amount", ""),
                    row.get("card_amount", ""),
                    row.get("upi_amount", ""),
                    row.get("total_amount", ""),
                    row.get("remarks", ""),
                    row.get("created_by", ""),
                    row.get("updated_on", ""),
                ]
            )

        filename = f"cashflow_report_{business_now_naive().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            buffer.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Cashflow export failed", e)
        return redirect("/cashflow")


@app.route("/revenue-locks")
def revenue_locks_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]
    if not can_manage_revenue:
        flash("Access limited: Revenue Dashboard only", "warning")
        return redirect("/revenue-dashboard")
    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=True,
            recent_limit=20,
        )
        return render_template("revenue_locks.html", can_manage_revenue=can_manage_revenue, **context)
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue lock page error", e)
        return redirect("/dashboard")


@app.route("/revenue-target-settings")
def revenue_target_settings_page():
    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = (session.get("branch") or "").strip()
    can_manage_revenue = role in ["super_admin", "admin"]

    if role != "super_admin":
        return "Access Denied"

    try:
        context = _get_revenue_view_context(
            request.args,
            can_manage_revenue,
            session_branch,
            include_edit_row=False,
            include_locked_periods=False,
            recent_limit=1,
        )

        db = get_db()
        cursor = db.cursor(dictionary=True)
        try:
            try:
                cursor.execute(
                    """
                    SELECT id, branch_name,
                           COALESCE(total_target, COALESCE(sales_target, 0) + COALESCE(service_target, 0)) AS effective_target
                    FROM branch_revenue_targets
                    ORDER BY branch_name ASC
                    """
                )
            except Error as query_err:
                if getattr(query_err, "errno", None) == 1054 or "total_target" in str(query_err).lower():
                    cursor.execute(
                        """
                        SELECT id, branch_name,
                               (COALESCE(sales_target, 0) + COALESCE(service_target, 0)) AS effective_target
                        FROM branch_revenue_targets
                        ORDER BY branch_name ASC
                        """
                    )
                else:
                    raise

            revenue_targets = cursor.fetchall()
        finally:
            _safe_close(cursor, db)

        return render_template(
            "revenue_target_settings.html",
            can_manage_revenue=can_manage_revenue,
            revenue_targets=revenue_targets,
            **context,
        )
    except PermissionError:
        return "Access Denied"
    except Error as e:
        _flash_internal_error("Revenue target settings error", e)
        return redirect("/dashboard")


@app.route("/upload-revenue-entry", methods=["POST"])
def upload_revenue_entry():
    if "username" not in session:
        return redirect("/login")
    if session.get("role") not in ["super_admin", "admin"]:
        return "Access Denied"

    session_branch = (session.get("branch") or "").strip()
    has_global_scope = session_branch.upper() == "ALL"

    from_date = _normalize_date_input(request.form.get("upload_from_date", ""))
    to_date = _normalize_date_input(request.form.get("upload_to_date", ""))
    if not from_date or not to_date:
        flash("From date and To date are required", "danger")
        return redirect("/revenue-upload")

    upload = request.files.get("revenue_excel_file")
    if not upload or not (upload.filename or "").strip():
        flash("Choose an Excel or CSV file", "danger")
        return redirect("/revenue-upload")

    # Check if period is locked
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(
            "SELECT COUNT(*) FROM revenue_entry_period_locks WHERE from_date <= %s AND to_date >= %s",
            (to_date, from_date),
        )
        if cursor.fetchone()[0] > 0:
            flash(f"Period {from_date} to {to_date} is locked. Unlock it first.", "danger")
            return redirect("/revenue-upload")
    finally:
        _safe_close(cursor, db)

    try:
        parsed_rows = list(_iter_revenue_entry_excel_rows(upload))
    except ValueError as e:
        _flash_internal_error("Could not read revenue upload file", e)
        return redirect("/revenue-upload")
    except Exception as e:
        _flash_internal_error("Could not read revenue upload file", e)
        return redirect("/revenue-upload")

    if not parsed_rows:
        flash("No valid data rows found in file", "warning")
        return redirect("/revenue-upload")

    db = get_db()
    cursor = db.cursor()
    saved = skipped = 0
    error_samples = []
    try:
        db2 = get_db()
        c2 = db2.cursor(dictionary=True)
        valid_branch_map = {b.upper(): b for b in _load_known_branches(c2)}
        c2.close()
        db2.close()

        for row_number, branch, sales, service, total, zone in parsed_rows:
            normalized = valid_branch_map.get(branch.strip().upper())
            if not normalized:
                skipped += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: unknown branch '{branch}'")
                continue
            if not has_global_scope and normalized != session_branch:
                skipped += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: branch '{branch}' is outside your scope")
                continue

            cursor.execute(
                """
                INSERT INTO branch_revenue_entries
                    (entry_date, branch_name, sales_profit, service_charges, total_profit, zone, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    sales_profit=VALUES(sales_profit),
                    service_charges=VALUES(service_charges),
                    total_profit=VALUES(total_profit),
                    zone=VALUES(zone),
                    created_by=VALUES(created_by)
                """,
                (from_date, normalized, round(sales, 2), round(service, 2), round(total, 2),
                 zone or None, session.get("username")),
            )
            saved += 1

        db.commit()
        flash(f"Upload complete: {saved} saved, {skipped} skipped", "success")
        for msg in error_samples:
            flash(msg, "warning")
    except Error as e:
        _flash_internal_error("Revenue upload failed", e)
    finally:
        _safe_close(cursor, db)

    return redirect(f"/revenue-reports?from_date={from_date}&to_date={to_date}")


@app.route("/lock-revenue-entry-period", methods=["POST"])
def lock_revenue_entry_period():
    if "username" not in session:
        return redirect("/login")
    if session.get("role") not in ["super_admin", "admin"]:
        return "Access Denied"

    action = (request.form.get("lock_action") or "lock").strip().lower()
    from_date = _normalize_date_input(request.form.get("lock_from_date", ""))
    to_date = _normalize_date_input(request.form.get("lock_to_date", ""))

    if not from_date or not to_date:
        flash("From date and To date are required to lock/unlock", "danger")
        return redirect("/revenue-locks")

    db = get_db()
    cursor = db.cursor()
    try:
        if action == "lock":
            cursor.execute(
                """
                INSERT INTO revenue_entry_period_locks (from_date, to_date, locked_by, locked_at)
                VALUES (%s, %s, %s, NOW())
                ON DUPLICATE KEY UPDATE locked_by=VALUES(locked_by), locked_at=NOW()
                """,
                (from_date, to_date, session.get("username")),
            )
            flash(f"Period locked: {from_date} to {to_date}", "success")
        else:
            cursor.execute(
                "DELETE FROM revenue_entry_period_locks WHERE from_date=%s AND to_date=%s",
                (from_date, to_date),
            )
            flash(f"Period unlocked: {from_date} to {to_date}", "success")
        db.commit()
    except Error as e:
        _flash_internal_error("Failed to update revenue entry lock", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/revenue-locks")


@app.route("/delete-revenue-entry/<int:entry_id>", methods=["POST"])
def delete_revenue_entry(entry_id):
    if "username" not in session:
        return redirect("/login")

    if session.get("role") not in ["super_admin", "admin"]:
        return "Access Denied"

    session_branch = (session.get("branch") or "").strip()
    has_global_scope = session_branch.upper() == "ALL"
    redirect_target = _get_revenue_reports_redirect_target()

    db = None
    cursor = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        result = _delete_revenue_entries(cursor, [entry_id], session_branch, has_global_scope)
        if result["missing"] > 0:
            flash("Revenue entry not found", "danger")
            db.rollback()
            return redirect(redirect_target)
        if result["locked"] > 0:
            flash("Cannot delete: this entry belongs to a locked period.", "danger")
            db.rollback()
            return redirect(redirect_target)
        db.commit()
        flash("Revenue entry deleted", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Failed to delete revenue entry", e)
    finally:
        _safe_close(cursor, db)

    return redirect(redirect_target)


@app.route("/revenue-reports/bulk-delete", methods=["POST"])
def bulk_delete_revenue_entries():
    if "username" not in session:
        return redirect("/login")

    if session.get("role") not in ["super_admin", "admin"]:
        return "Access Denied"

    selected_entry_ids = request.form.getlist("entry_ids")
    if not selected_entry_ids:
        flash("Select at least one revenue entry to delete.", "warning")
        return redirect(_get_revenue_reports_redirect_target())

    session_branch = (session.get("branch") or "").strip()
    has_global_scope = session_branch.upper() == "ALL"
    redirect_target = _get_revenue_reports_redirect_target()

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        result = _delete_revenue_entries(cursor, selected_entry_ids, session_branch, has_global_scope)
        db.commit()

        if result["deleted"] > 0:
            flash(f"{result['deleted']} revenue entr{'y' if result['deleted'] == 1 else 'ies'} deleted.", "success")
        if result["locked"] > 0:
            flash(
                f"{result['locked']} selected entr{'y was' if result['locked'] == 1 else 'ies were'} skipped because the period is locked.",
                "warning",
            )
        if result["missing"] > 0:
            flash(
                f"{result['missing']} selected entr{'y was' if result['missing'] == 1 else 'ies were'} not found or outside your branch scope.",
                "warning",
            )
        if result["deleted"] == 0 and result["locked"] == 0 and result["missing"] == 0:
            flash("No valid revenue entries were selected.", "warning")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Failed to bulk delete revenue entries", e)
    finally:
        _safe_close(cursor, db)

    return redirect(redirect_target)


@app.route("/quotations")
def quotations_page():

    if "username" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        role = session.get("role")
        session_branch = session.get("branch")

        from_date = _normalize_date_input(request.args.get("from_date", ""))
        to_date = _normalize_date_input(request.args.get("to_date", ""))
        search = (request.args.get("search") or "").strip()
        filter_branch = (request.args.get("filter_branch") or "").strip()

        where = ["1=1"]
        params = []

        if _user_has_all_branch_scope(role, session_branch):
            if filter_branch and filter_branch.upper() != "ALL":
                where.append("q.branch_name=%s")
                params.append(filter_branch)
        else:
            where.append("q.branch_name=%s")
            params.append(session_branch)

        if from_date:
            where.append("q.quote_date >= %s")
            params.append(from_date)
        if to_date:
            where.append("q.quote_date <= %s")
            params.append(to_date)

        if search:
            where.append("(q.quote_number LIKE %s OR q.customer_name LIKE %s OR q.customer_mobile LIKE %s)")
            token = f"%{search}%"
            params.extend([token, token, token])

        cursor.execute(
            f"""
            SELECT q.id, q.quote_number, q.quote_date, q.branch_name,
                   q.customer_name, q.customer_mobile, q.engineer_name,
                   q.grand_total, q.updated_at
            FROM quotations q
            WHERE {' AND '.join(where)}
            ORDER BY q.id DESC
            """,
            tuple(params),
        )
        rows = cursor.fetchall()

        for row in rows:
            row["updated_on"] = format_datetime_display(row.get("updated_at"))

        branch_options = _quotation_branch_options(cursor, role, session_branch)

        return render_template(
            "quotations.html",
            rows=rows,
            from_date=from_date,
            to_date=to_date,
            search=search,
            filter_branch=filter_branch,
            branch_options=branch_options,
            role=role,
            session_branch=session_branch,
        )

    except Error as e:
        _flash_internal_error("Error loading quotations", e)
        return redirect("/dashboard")

    finally:
        _safe_close(cursor, db)


def _quotation_form_context(cursor, quotation_id=None):
    role = session.get("role")
    session_branch = session.get("branch")

    quotation = {
        "id": None,
        "quote_number": _next_quote_number(cursor),
        "quote_date": business_now_naive().strftime("%Y-%m-%d"),
        "branch_name": session_branch if session_branch != "ALL" else "",
        "customer_name": "",
        "customer_mobile": "",
        "customer_address": "",
        "customer_gst_no": "",
        "engineer_name": "",
        "engineer_mobile": "",
        "terms_text": "",
        "grand_total": 0,
    }
    items = []

    if quotation_id:
        cursor.execute("SELECT * FROM quotations WHERE id=%s", (quotation_id,))
        found = cursor.fetchone()
        if not found:
            return None, None, None, None

        if not _branch_in_scope(role, session_branch, found.get("branch_name")):
            return None, None, None, None

        quotation = found
        if quotation.get("quote_date") and hasattr(quotation.get("quote_date"), "strftime"):
            quotation["quote_date"] = quotation.get("quote_date").strftime("%Y-%m-%d")

        cursor.execute(
            """
            SELECT line_no, item_name, narration, qty, amount, final_amount
            FROM quotation_items
            WHERE quotation_id=%s
            ORDER BY line_no ASC, id ASC
            """,
            (quotation_id,),
        )
        items = cursor.fetchall()

    branch_options = _quotation_branch_options(cursor, role, session_branch)

    default_branch = quotation.get("branch_name") or (branch_options[0] if branch_options else "ALL")
    profile = get_branch_print_profile(cursor, default_branch)

    if not quotation.get("terms_text"):
        quotation["terms_text"] = profile.get("quotation_terms") or profile.get("terms_text") or ""

    return quotation, items, branch_options, profile


def _format_print_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    try:
        return datetime.fromisoformat(str(value)).strftime("%d/%m/%Y")
    except Exception:
        return str(value)


@app.route("/quotation/new", methods=["GET", "POST"])
def quotation_new():

    if "username" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        role = session.get("role")
        session_branch = session.get("branch")

        if request.method == "POST":
            branch_name = (request.form.get("branch_name") or "").strip()
            quote_date = _normalize_date_input(request.form.get("quote_date", ""))
            customer_name = (request.form.get("customer_name") or "").strip()
            customer_mobile = (request.form.get("customer_mobile") or "").strip()
            customer_address = (request.form.get("customer_address") or "").strip()
            customer_gst_no = (request.form.get("customer_gst_no") or "").strip()
            engineer_name = (request.form.get("engineer_name") or "").strip()
            engineer_mobile = (request.form.get("engineer_mobile") or "").strip()
            terms_text = (request.form.get("terms_text") or "").strip()

            if not branch_name:
                flash("Branch is required", "danger")
                return redirect("/quotation/new")

            if not _branch_in_scope(role, session_branch, branch_name):
                flash("You do not have access to this branch", "danger")
                return redirect("/quotation/new")

            if not quote_date:
                quote_date = business_now_naive().strftime("%Y-%m-%d")

            if not customer_name:
                flash("Customer name is required", "danger")
                return redirect("/quotation/new")

            items = _extract_quotation_items_from_form(request.form)
            if not items:
                flash("Add at least one quotation item", "danger")
                return redirect("/quotation/new")

            quote_number = _next_quote_number(cursor)
            grand_total = round(sum(float(i.get("final_amount") or 0) for i in items), 2)

            cursor.execute(
                """
                INSERT INTO quotations
                (quote_number, quote_date, branch_name, customer_name, customer_mobile,
                 customer_address, customer_gst_no, engineer_name, engineer_mobile,
                 terms_text, grand_total, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    quote_number,
                    quote_date,
                    branch_name,
                    customer_name,
                    customer_mobile,
                    customer_address,
                    customer_gst_no,
                    engineer_name,
                    engineer_mobile,
                    terms_text,
                    grand_total,
                    session.get("username"),
                ),
            )

            quotation_id = cursor.lastrowid
            cursor.executemany(
                """
                INSERT INTO quotation_items
                (quotation_id, line_no, item_name, narration, qty, amount, final_amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        quotation_id,
                        idx + 1,
                        row["item_name"],
                        row["narration"],
                        row["qty"],
                        row["amount"],
                        row["final_amount"],
                    )
                    for idx, row in enumerate(items)
                ],
            )

            db.commit()
            flash(f"Quotation {quote_number} saved", "success")
            return redirect(f"/quotation/{quotation_id}/edit")

        quotation, items, branch_options, profile = _quotation_form_context(cursor)
        return render_template(
            "quick_quotation.html",
            quotation=quotation,
            items=items,
            branch_options=branch_options,
            profile=profile,
            mode="new",
        )

    except Error as e:
        _flash_internal_error("Quotation error", e)
        return redirect("/quotations")

    finally:
        _safe_close(cursor, db)


@app.route("/quotation/<int:quotation_id>/edit", methods=["GET", "POST"])
def quotation_edit(quotation_id):

    if "username" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    try:
        role = session.get("role")
        session_branch = session.get("branch")

        cursor.execute("SELECT id, branch_name, quote_number FROM quotations WHERE id=%s", (quotation_id,))
        row = cursor.fetchone()
        if not row:
            flash("Quotation not found", "danger")
            return redirect("/quotations")

        if not _branch_in_scope(role, session_branch, row.get("branch_name")):
            flash("Access denied for this quotation", "danger")
            return redirect("/quotations")

        if request.method == "POST":
            branch_name = (request.form.get("branch_name") or "").strip()
            quote_date = _normalize_date_input(request.form.get("quote_date", ""))
            customer_name = (request.form.get("customer_name") or "").strip()
            customer_mobile = (request.form.get("customer_mobile") or "").strip()
            customer_address = (request.form.get("customer_address") or "").strip()
            customer_gst_no = (request.form.get("customer_gst_no") or "").strip()
            engineer_name = (request.form.get("engineer_name") or "").strip()
            engineer_mobile = (request.form.get("engineer_mobile") or "").strip()
            terms_text = (request.form.get("terms_text") or "").strip()

            if not branch_name:
                flash("Branch is required", "danger")
                return redirect(f"/quotation/{quotation_id}/edit")

            if not _branch_in_scope(role, session_branch, branch_name):
                flash("You do not have access to this branch", "danger")
                return redirect(f"/quotation/{quotation_id}/edit")

            if not quote_date:
                quote_date = business_now_naive().strftime("%Y-%m-%d")

            if not customer_name:
                flash("Customer name is required", "danger")
                return redirect(f"/quotation/{quotation_id}/edit")

            items = _extract_quotation_items_from_form(request.form)
            if not items:
                flash("Add at least one quotation item", "danger")
                return redirect(f"/quotation/{quotation_id}/edit")

            grand_total = round(sum(float(i.get("final_amount") or 0) for i in items), 2)

            cursor.execute(
                """
                UPDATE quotations
                SET quote_date=%s,
                    branch_name=%s,
                    customer_name=%s,
                    customer_mobile=%s,
                    customer_address=%s,
                    customer_gst_no=%s,
                    engineer_name=%s,
                    engineer_mobile=%s,
                    terms_text=%s,
                    grand_total=%s
                WHERE id=%s
                """,
                (
                    quote_date,
                    branch_name,
                    customer_name,
                    customer_mobile,
                    customer_address,
                    customer_gst_no,
                    engineer_name,
                    engineer_mobile,
                    terms_text,
                    grand_total,
                    quotation_id,
                ),
            )

            cursor.execute("DELETE FROM quotation_items WHERE quotation_id=%s", (quotation_id,))
            cursor.executemany(
                """
                INSERT INTO quotation_items
                (quotation_id, line_no, item_name, narration, qty, amount, final_amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        quotation_id,
                        idx + 1,
                        row_item["item_name"],
                        row_item["narration"],
                        row_item["qty"],
                        row_item["amount"],
                        row_item["final_amount"],
                    )
                    for idx, row_item in enumerate(items)
                ],
            )

            db.commit()
            flash(f"Quotation {row.get('quote_number')} updated", "success")

        quotation, items, branch_options, profile = _quotation_form_context(cursor, quotation_id)
        if quotation is None:
            flash("Quotation not found or access denied", "danger")
            return redirect("/quotations")

        return render_template(
            "quick_quotation.html",
            quotation=quotation,
            items=items,
            branch_options=branch_options,
            profile=profile,
            mode="edit",
        )

    except Error as e:
        _flash_internal_error("Quotation error", e)
        return redirect("/quotations")

    finally:
        _safe_close(cursor, db)


@app.route("/quotation/<int:quotation_id>/delete", methods=["POST"])
def delete_quotation(quotation_id):

    if "username" not in session:
        return redirect("/login")

    return_to = (request.form.get("return_to") or "").strip()
    if not return_to.startswith("/quotations"):
        return_to = "/quotations"

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        role = session.get("role")
        session_branch = session.get("branch")

        cursor.execute("SELECT id, branch_name, quote_number FROM quotations WHERE id=%s", (quotation_id,))
        row = cursor.fetchone()
        if not row:
            flash("Quotation not found", "danger")
            return redirect(return_to)

        if not _branch_in_scope(role, session_branch, row.get("branch_name")):
            flash("Access denied for this quotation", "danger")
            return redirect(return_to)

        cursor.execute("DELETE FROM quotation_items WHERE quotation_id=%s", (quotation_id,))
        cursor.execute("DELETE FROM quotations WHERE id=%s", (quotation_id,))
        db.commit()
        flash(f"Quotation {row.get('quote_number')} deleted permanently", "success")

    except Error as e:
        if db is not None:
            db.rollback()
        _flash_internal_error("Failed to delete quotation", e)

    finally:
        _safe_close(cursor, db)

    return redirect(return_to)


@app.route("/quotation/<int:quotation_id>/print")
def quotation_print(quotation_id):

    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        quotation, items, _, profile = _quotation_form_context(cursor, quotation_id)
        if quotation is None:
            flash("Quotation not found or access denied", "danger")
            return redirect("/quotations")

        terms_text = (quotation.get("terms_text") or profile.get("quotation_terms") or profile.get("terms_text") or "").strip()

        return render_template(
            "quotation_print.html",
            quotation=quotation,
            items=items,
            profile=profile,
            terms_text=terms_text,
            quote_date_display=_format_print_date(quotation.get("quote_date")),
            print_date=business_now_naive().strftime("%d/%m/%Y %H:%M"),
        )

    except Error as e:
        _flash_internal_error("Could not load quotation print view", e)
        return redirect("/quotations")

    finally:
        _safe_close(cursor, db)


# ---------------- NEW JOB ---------------- #

@app.route("/new-job", methods=["GET", "POST"])
def new_job():

    if "username" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor(dictionary=True)

    # Load dropdown options
    cursor.execute("SELECT type,value FROM dropdown_options ORDER BY type, `order`, value")
    rows = cursor.fetchall()

    dropdowns = {}

    for row in rows:
        dropdowns.setdefault(row["type"], []).append(row["value"])

    # Load users for Received By and Assigned Engineer
    cursor.execute("SELECT username FROM users ORDER BY username")
    users = [row["username"] for row in cursor.fetchall()]
    engineer_users = _load_engineer_usernames(cursor)

    # Load allowed branches for the current user (for the branch select list)
    role = session.get("role")
    username = session.get("username")
    session_branch = session.get("branch")

    try:
        branch_scope = _get_branch_scope(role, session_branch)
    except PermissionError:
        _safe_close(cursor, db)
        flash("Invalid branch scope", "danger")
        return redirect("/dashboard")

    if _user_has_all_branch_scope(role, session_branch):
        branches = _load_known_branches(cursor)
    else:
        branches = [branch_scope]

    if request.method == "POST":

        customer_name = request.form["customer_name"]
        mobile = request.form["mobile"]
        alt_no = request.form.get("alt_no")
        email = request.form.get("email")
        location = request.form.get("location")
        address = request.form.get("address")
        pin_code = request.form.get("pin_code")
        device = request.form["device"]
        model = request.form["model"]
        serial = request.form["serial"]
        complaint = request.form["complaint"]
        accessories_received = request.form["accessories_received"]
        engineer_remarks = request.form["engineer_remarks"]

        priority = request.form.get("priority")
        call_type = request.form.get("call_type")
        complaint_type = request.form.get("complaint_type")
        warranty_status = request.form.get("warranty_status")
        backup_required = request.form.get("backup_required")
        received_by = (request.form.get("received_by") or session.get("username") or "").strip()
        assigned_engineer = (request.form.get("assigned_engineer") or "").strip()
        estimate_amount = request.form.get("estimate_amount")
        if not estimate_amount or str(estimate_amount).strip() == '':
            estimate_amount = None
        service_charges = _parse_money(request.form.get("service_charges"))

        # Validate phone numbers
        def digits_only(value):
            return re.sub(r"\D", "", value or "")

        if len(digits_only(mobile)) < 10:
            flash("Mobile number must contain at least 10 digits", "danger")
            cursor.close()
            db.close()
            return redirect("/new-job")

        if alt_no:
            if len(digits_only(alt_no)) < 10:
                flash("Alt number must contain at least 10 digits", "danger")
                cursor.close()
                db.close()
                return redirect("/new-job")

        if not received_by:
            flash("Received By is required", "danger")
            cursor.close()
            db.close()
            return redirect("/new-job")

        if not assigned_engineer:
            flash("Assigned Engineer is required", "danger")
            cursor.close()
            db.close()
            return redirect("/new-job")
        valid_assigned_engineer = _validate_assigned_engineer_name(assigned_engineer, engineer_users)
        if not valid_assigned_engineer:
            flash("Select Assigned Engineer from the engineers list", "danger")
            cursor.close()
            db.close()
            return redirect("/new-job")
        assigned_engineer = valid_assigned_engineer

        # -------- CONCURRENCY-SAFE JOB NUMBER -------- #
        saved_photo_filenames = []
        try:
            branch = _resolve_branch_input(role, session_branch, request.form.get("branch"))
            if not branch or branch.upper() == "ALL" or not _is_known_branch(cursor, branch):
                flash("Select a valid branch", "danger")
                return redirect("/new-job")

            job_number = _next_job_number(cursor)

            photo_filename = None
            upload_files = _get_job_upload_files(request.files)
            if len(upload_files) > MAX_JOB_ATTACHMENTS:
                raise ValueError(f"You can upload up to {MAX_JOB_ATTACHMENTS} images only")

            if upload_files:
                saved_photo_filenames = _save_job_images(upload_files, job_number)
                photo_filename = saved_photo_filenames[0]

            query = """
            INSERT INTO jobs
            (job_number, customer_name, mobile, alt_no, email, location, address, pin_code,
            device, model, serial_number, complaint, complaint_type, warranty_status, backup_required,
            priority, call_type, received_by, assigned_engineer, estimate_amount, accessories_received,
            engineer_remarks, photo, branch_name, status)

            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Open')
            """

            values = (
                job_number,
                customer_name,
                mobile,
                alt_no,
                email,
                location,
                address,
                pin_code,
                device,
                model,
                serial,
                complaint,
                complaint_type,
                warranty_status,
                backup_required,
                priority,
                call_type,
                received_by,
                assigned_engineer,
                estimate_amount,
                accessories_received,
                engineer_remarks,
                photo_filename,
                branch
            )

            cursor.execute(query, values)
            job_id = cursor.lastrowid
            _insert_job_attachments(cursor, job_id, saved_photo_filenames)
            used_spares = _extract_used_spares_from_form(request.form)
            if used_spares:
                cursor.executemany(
                    "INSERT INTO used_spares (job_id, spare_name, amount) VALUES (%s, %s, %s)",
                    [(job_id, item["spare_name"], item["amount"]) for item in used_spares],
                )
            _set_job_spares_billing(cursor, job_id, "Pending" if used_spares else "Not Required")
            db.commit()
            flash(f"Job {job_number} created successfully!", "success")
        except ValueError as upload_error:
            db.rollback()
            _remove_saved_images(saved_photo_filenames, app.config["JOB_PHOTO_FOLDER"])
            flash(str(upload_error), "danger")
            return redirect("/new-job")
        except PermissionError:
            db.rollback()
            _remove_saved_images(saved_photo_filenames, app.config["JOB_PHOTO_FOLDER"])
            flash("You do not have access to this branch", "danger")
            return redirect("/new-job")
        except Exception as e:
            db.rollback()
            _remove_saved_images(saved_photo_filenames, app.config["JOB_PHOTO_FOLDER"])
            _flash_internal_error("Job creation failed", e)
            return redirect("/new-job")
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception:
                    pass

        cursor.close()
        db.close()

        return redirect(f"/job-details/{job_id}")

    cursor.close()
    db.close()

    # For branch search/autocomplete
    search_query = request.args.get("branch_search", "")
    matching_branches = []
    if search_query:
        matching_branches = [branch_name for branch_name in branches if search_query.lower() in branch_name.lower()][:10]

    return render_template(
        "new_job.html",
        dropdowns=dropdowns,
        branches=branches,
        users=users,
        engineer_users=engineer_users,
        used_spares=[],
        branch=session.get("branch", "Unknown"),
        role=session.get("role"),
        matching_branches=matching_branches,
        search_query=search_query
    )


# ---------------- EDIT JOB ---------------- #

@app.route("/edit-job/<int:job_id>", methods=["GET", "POST"])
def edit_job(job_id):

    if "username" not in session:
        return redirect("/login")

    # Check role permissions
    allowed_roles = ["super_admin", "admin", "coordinator"]
    if session.get("role") not in allowed_roles:
        flash("Access denied", "danger")
        return redirect("/jobs")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    role = session.get("role")
    session_branch = session.get("branch")

    try:
        branch_scope = _get_branch_scope(role, session_branch)
        job = _fetch_scoped_job(cursor, job_id, role, session_branch)

        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_manage_job_core(role, session_branch, job):
            flash("Only the owning branch can edit this case", "danger")
            return redirect(f"/job-details/{job_id}")

        if job.get("closure_status") and role not in ["admin", "super_admin"]:
            flash("Cannot edit closed case", "danger")
            return redirect("/jobs")

        cursor.execute(
            "SELECT id, spare_name, amount FROM used_spares WHERE job_id=%s ORDER BY id ASC",
            (job_id,),
        )
        used_spares = cursor.fetchall()
        existing_spares_signature = _job_used_spares_signature(used_spares)

        cursor.execute("SELECT type,value FROM dropdown_options ORDER BY type, `order`, value")
        rows = cursor.fetchall()
        dropdowns = {}
        for row in rows:
            dropdowns.setdefault(row["type"], []).append(row["value"])

        cursor.execute("SELECT username FROM users ORDER BY username")
        users = [r["username"] for r in cursor.fetchall()]
        engineer_users = _load_engineer_usernames(cursor)

        if _user_has_all_branch_scope(role, session_branch):
            branches = _load_known_branches(cursor)
        else:
            branches = [branch_scope]

        attachments = _load_job_attachments(cursor, job_id, job.get("photo"))

        if request.method == "POST":
            customer_name = request.form["customer_name"]
            mobile = request.form["mobile"]
            alt_no = request.form.get("alt_no")
            email = request.form.get("email")
            location = request.form.get("location")
            address = request.form.get("address")
            pin_code = request.form.get("pin_code")
            device = request.form["device"]
            model = request.form["model"]
            serial = request.form["serial"]
            complaint = request.form["complaint"]
            accessories_received = request.form["accessories_received"]
            engineer_remarks = request.form["engineer_remarks"]

            priority = request.form.get("priority")
            call_type = request.form.get("call_type")
            complaint_type = request.form.get("complaint_type")
            warranty_status = request.form.get("warranty_status")
            backup_required = request.form.get("backup_required")
            received_by = (request.form.get("received_by") or "").strip()
            assigned_engineer = (request.form.get("assigned_engineer") or "").strip()
            estimate_amount = request.form.get("estimate_amount")
            if not estimate_amount or str(estimate_amount).strip() == "":
                estimate_amount = None
            service_charges = _parse_money(request.form.get("service_charges"))

            def digits_only(value):
                return re.sub(r"\D", "", value or "")

            if len(digits_only(mobile)) < 10:
                flash("Mobile number must contain at least 10 digits", "danger")
                return redirect(f"/edit-job/{job_id}")

            if alt_no and len(digits_only(alt_no)) < 10:
                flash("Alt number must contain at least 10 digits", "danger")
                return redirect(f"/edit-job/{job_id}")

            if not received_by:
                flash("Received By is required", "danger")
                return redirect(f"/edit-job/{job_id}")

            if not assigned_engineer:
                flash("Assigned Engineer is required", "danger")
                return redirect(f"/edit-job/{job_id}")
            valid_assigned_engineer = _validate_assigned_engineer_name(assigned_engineer, engineer_users)
            if not valid_assigned_engineer:
                flash("Select Assigned Engineer from the engineers list", "danger")
                return redirect(f"/edit-job/{job_id}")
            assigned_engineer = valid_assigned_engineer

            try:
                branch = _resolve_branch_input(role, session_branch, request.form.get("branch"))
            except PermissionError:
                flash("You do not have access to this branch", "danger")
                return redirect(f"/edit-job/{job_id}")

            if not branch or branch.upper() == "ALL" or not _is_known_branch(cursor, branch):
                flash("Select a valid branch", "danger")
                return redirect(f"/edit-job/{job_id}")

            updated_spares = _extract_used_spares_from_form(request.form)
            updated_spares_signature = _job_used_spares_signature(updated_spares)
            if role not in ["super_admin", "admin"] and not _saved_used_spares_preserved(used_spares, updated_spares):
                flash("Saved used spares can be removed only by admin or super admin.", "danger")
                return redirect(f"/edit-job/{job_id}")

            photo_filename = job.get("photo")
            upload_files = _get_job_upload_files(request.files)
            if len(attachments) + len(upload_files) > MAX_JOB_ATTACHMENTS:
                flash(f"Maximum {MAX_JOB_ATTACHMENTS} images are allowed per case", "danger")
                return redirect(f"/edit-job/{job_id}")

            saved_photo_filenames = []
            if upload_files:
                saved_photo_filenames = _save_job_images(upload_files, job_id)
                if not photo_filename:
                    photo_filename = saved_photo_filenames[0]

            try:
                cursor.execute(
                    """
                    UPDATE jobs SET
                        customer_name=%s,
                        mobile=%s,
                        alt_no=%s,
                        email=%s,
                        location=%s,
                        address=%s,
                        pin_code=%s,
                        device=%s,
                        model=%s,
                        serial_number=%s,
                        complaint=%s,
                        complaint_type=%s,
                        warranty_status=%s,
                        backup_required=%s,
                        priority=%s,
                        call_type=%s,
                        received_by=%s,
                        assigned_engineer=%s,
                        estimate_amount=%s,
                        service_charges=%s,
                        accessories_received=%s,
                        engineer_remarks=%s,
                        photo=%s,
                        branch_name=%s
                    WHERE id=%s
                    """,
                    (
                        customer_name,
                        mobile,
                        alt_no,
                        email,
                        location,
                        address,
                        pin_code,
                        device,
                        model,
                        serial,
                        complaint,
                        complaint_type,
                        warranty_status,
                        backup_required,
                        priority,
                        call_type,
                        received_by,
                        assigned_engineer,
                        estimate_amount,
                        service_charges,
                        accessories_received,
                        engineer_remarks,
                        photo_filename,
                        branch,
                        job_id,
                    )
                )

                if saved_photo_filenames:
                    _insert_job_attachments(cursor, job_id, saved_photo_filenames)

                cursor.execute("DELETE FROM used_spares WHERE job_id=%s", (job_id,))
                if updated_spares:
                    cursor.executemany(
                        "INSERT INTO used_spares (job_id, spare_name, amount) VALUES (%s, %s, %s)",
                        [(job_id, item["spare_name"], item["amount"]) for item in updated_spares],
                    )

                if updated_spares:
                    if updated_spares_signature != existing_spares_signature:
                        _set_job_spares_billing(cursor, job_id, "Pending", notes=job.get("spares_billing_notes"))
                    elif str(job.get("spares_invoice_no") or "").strip():
                        _set_job_spares_billing(
                            cursor,
                            job_id,
                            "Billed",
                            job.get("spares_invoice_no"),
                            job.get("spares_invoice_date"),
                            job.get("spares_billed_by"),
                            job.get("spares_billing_notes"),
                        )
                    else:
                        _set_job_spares_billing(cursor, job_id, "Pending", notes=job.get("spares_billing_notes"))
                else:
                    _set_job_spares_billing(cursor, job_id, "Not Required")

                db.commit()
            except Exception:
                db.rollback()
                _remove_saved_images(saved_photo_filenames, app.config["JOB_PHOTO_FOLDER"])
                raise

            flash("Job updated successfully", "success")
            return redirect(f"/job-details/{job_id}")

        return render_template(
            "edit_job.html",
            job=job,
            attachments=attachments,
            used_spares=used_spares,
            dropdowns=dropdowns,
            branches=branches,
            users=users,
            engineer_users=engineer_users,
            branch=session.get("branch", "Unknown"),
            role=session.get("role"),
        )
    except ValueError as upload_error:
        flash(str(upload_error), "danger")
        return redirect(f"/edit-job/{job_id}")
    except PermissionError:
        flash("Invalid branch scope", "danger")
        return redirect("/jobs")
    except Error as e:
        _flash_internal_error("Job update failed", e)
        return redirect("/jobs")
    finally:
        _safe_close(cursor, db)


# ---------------- JOB DETAILS ---------------- #

def _build_job_details_context(cursor, job_id, role, session_branch):
    job = _fetch_scoped_job(cursor, job_id, role, session_branch)

    if not job:
        return None

    now = business_now_naive()
    created_at = normalize_display_datetime(job.get("created_at"))
    closure_date = job.get("closure_date")
    age_days = None
    age_group = ""
    created_on = ""
    closed_on = ""

    if created_at:
        age_days = (now.date() - created_at.date()).days
        if age_days < 0:
            age_days = 0
        age_group = get_age_group(age_days)
        created_on = format_datetime_display(created_at)

    closed_on = format_datetime_display(closure_date)

    job["age_days"] = age_days
    job["age_group"] = age_group
    job["created_on"] = created_on
    job["closed_on"] = closed_on

    cursor.execute(
        "SELECT spare_name, amount FROM used_spares WHERE job_id=%s ORDER BY id ASC",
        (job_id,),
    )
    used_spares = cursor.fetchall()

    total_spares_amount = sum(float(s.get("amount") or 0) for s in used_spares)
    service_charges = float(job.get("service_charges") or 0)
    bill_total_amount = total_spares_amount + service_charges
    transfer_rows = _fetch_job_service_transfer_rows(cursor, job_id)
    transfer_summary = summarize_job_transfers(transfer_rows, service_charges)
    transfer_active = transfer_summary.get("active_transfer")
    can_manage_core = _user_can_manage_job_core(role, session_branch, job)
    can_send_transfer = bool(not job.get("closure_status") and _user_can_send_job_transfer(role, session_branch, job) and not transfer_summary.get("has_active_transfer"))
    can_update_transfer = bool(transfer_active and not job.get("closure_status") and _user_can_update_job_transfer(role, session_branch, job, transfer_active))

    cursor.execute("SELECT username FROM users ORDER BY username")
    transfer_user_options = [row.get("username") for row in cursor.fetchall() if row.get("username")]
    transfer_branch_options = []
    if can_send_transfer:
        transfer_branch_options = [
            branch_name
            for branch_name in _load_known_branches(cursor)
            if branch_name and branch_name.upper() != "ALL" and branch_name != job.get("branch_name")
        ]

    job["specialist_service_total"] = transfer_summary.get("specialist_service_total")
    job["closing_branch_service_margin"] = transfer_summary.get("closing_branch_service_margin")

    cursor.execute(
        "SELECT status, notes, updated_by, updated_at FROM job_status_logs WHERE job_id=%s ORDER BY updated_at ASC",
        (job_id,)
    )
    status_logs = cursor.fetchall()

    attachments = _load_job_attachments(cursor, job_id, job.get("photo"))
    spare_billing = _build_spares_billing_summary(job, used_spares)
    job["spares_billing_status_effective"] = spare_billing["status"]
    job["used_spares_count"] = spare_billing["used_spares_count"]
    job["spares_billing_pending"] = spare_billing["pending"]

    return {
        "job": job,
        "attachments": attachments,
        "used_spares": used_spares,
        "spare_billing": spare_billing,
        "total_spares_amount": total_spares_amount,
        "service_charges": service_charges,
        "bill_total_amount": bill_total_amount,
        "transfer_rows": transfer_rows,
        "transfer_summary": transfer_summary,
        "transfer_branch_options": transfer_branch_options,
        "transfer_user_options": transfer_user_options,
        "job_permissions": {
            "can_manage_core": can_manage_core,
            "can_send_transfer": can_send_transfer,
            "can_update_transfer": can_update_transfer,
            "can_view_specialist_history": bool(transfer_rows),
        },
        "transfer_allowed_next_statuses": allowed_next_transfer_statuses(transfer_active.get("status")) if transfer_active else (),
        "status_logs": status_logs,
    }

@app.route("/job-details/<int:job_id>")
def job_details(job_id):

    if "username" not in session:
        return redirect("/login")

    try:

        db = get_db()
        cursor = db.cursor(dictionary=True)
        role = session.get("role")
        session_branch = session.get("branch")

        context = _build_job_details_context(cursor, job_id, role, session_branch)
        if not context:
            flash("Job not found", "danger")
            return redirect("/jobs")

        open_close_case = request.args.get("open_close_case") == "1"
        open_spare_billing = request.args.get("open_spare_billing") == "1"
        continue_close_case = request.args.get("continue_close_case") == "1"

        return render_template(
            "job_details.html",
            open_close_case=open_close_case,
            open_spare_billing=open_spare_billing,
            continue_close_case=continue_close_case,
            **context,
        )

    except Error as e:
        _flash_internal_error("Could not load job details", e)
        return redirect("/jobs")

    finally:
        _safe_close(cursor, db)


@app.route("/job-print/<int:job_id>/<print_mode>")
def job_print(job_id, print_mode):

    if "username" not in session:
        return redirect("/login")

    normalized_mode = (print_mode or "").strip().lower()
    if normalized_mode not in {"create", "close"}:
        flash("Invalid print mode", "danger")
        return redirect(f"/job-details/{job_id}")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        role = session.get("role")
        session_branch = session.get("branch")

        context = _build_job_details_context(cursor, job_id, role, session_branch)
        if not context:
            flash("Job not found", "danger")
            return redirect("/jobs")

        job = context["job"]
        if normalized_mode == "close" and not job.get("closure_status"):
            flash("Close case print is available after the case is closed", "warning")
            return redirect(f"/job-details/{job_id}")

        profile = get_branch_print_profile(cursor, job.get("branch_name") or session_branch)
        return render_template(
            "job_print.html",
            print_mode=normalized_mode,
            print_title="Customer Job Copy" if normalized_mode == "create" else "Outward Challan",
            profile=profile,
            printed_at=business_now_naive().strftime("%d/%m/%Y %H:%M"),
            **context,
        )

    except Error as e:
        _flash_internal_error("Could not load job print view", e)
        return redirect(f"/job-details/{job_id}")

    finally:
        _safe_close(cursor, db)


@app.route("/job-spare-billing/<int:job_id>", methods=["POST"])
def job_spare_billing(job_id):

    if "username" not in session:
        return redirect("/login")

    allowed_roles = ["super_admin", "admin", "coordinator"]
    role = session.get("role")
    session_branch = session.get("branch")
    if role not in allowed_roles:
        flash("Access denied", "danger")
        return redirect("/jobs")

    requested_status = _normalize_spares_billing_status(request.form.get("spares_billing_status"), False)
    invoice_no = (request.form.get("spares_invoice_no") or "").strip()
    invoice_date = _normalize_date_input(request.form.get("spares_invoice_date", ""))
    billing_notes = (request.form.get("spares_billing_notes") or "").strip()
    continue_close_case = (request.form.get("next_action") or "").strip().lower() == "close_case"
    billing_redirect_url = f"/job-details/{job_id}?open_spare_billing=1"
    if continue_close_case:
        billing_redirect_url += "&continue_close_case=1"

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, role, session_branch)
        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_manage_job_core(role, session_branch, job):
            flash("Only the owning branch can update spare billing", "danger")
            return redirect(f"/job-details/{job_id}")

        if job.get("closure_status") and role not in ["super_admin", "admin"]:
            flash("Only admin can update spare billing after case closure", "danger")
            return redirect(f"/job-details/{job_id}")

        cursor.execute("SELECT spare_name, amount FROM used_spares WHERE job_id=%s ORDER BY id ASC", (job_id,))
        used_spares = cursor.fetchall()
        spare_billing = _build_spares_billing_summary(job, used_spares)

        if not spare_billing["has_used_spares"]:
            _set_job_spares_billing(cursor, job_id, "Not Required")
            db.commit()
            flash("No spare items found on this case. Billing marked as not required.", "info")
            if continue_close_case:
                return redirect(f"/job-details/{job_id}?open_close_case=1")
            return redirect(f"/job-details/{job_id}")

        if requested_status == "Not Required":
            flash("Used spare items exist for this case. Choose Pending or Billed.", "danger")
            return redirect(billing_redirect_url)

        if requested_status == "Billed":
            if not invoice_no or not invoice_date:
                flash("Invoice number and invoice date are required to mark spare billing as billed.", "danger")
                return redirect(billing_redirect_url)
            _set_job_spares_billing(cursor, job_id, "Billed", invoice_no, invoice_date, session.get("username"), billing_notes)
            db.commit()
            if continue_close_case:
                flash("Spare billing updated. Continue with close case.", "success")
                return redirect(f"/job-details/{job_id}?open_close_case=1")
            flash("Spare billing updated", "success")
            return redirect(f"/job-details/{job_id}")

        _set_job_spares_billing(cursor, job_id, "Pending", notes=billing_notes)
        db.commit()
        if continue_close_case:
            flash("Invoice number and date are required before closing this case.", "warning")
            return redirect(billing_redirect_url)
        flash("Spare billing marked as pending", "warning")
        return redirect(f"/job-details/{job_id}")

    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Could not update spare billing", e)
        return redirect(f"/job-details/{job_id}")

    finally:
        _safe_close(cursor, db)


@app.route("/job-transfer/send/<int:job_id>", methods=["POST"])
def send_job_transfer(job_id):

    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = session.get("branch")
    if role not in ["super_admin", "admin", "coordinator"]:
        flash("Access denied", "danger")
        return redirect(f"/job-details/{job_id}")

    to_branch_name = (request.form.get("to_branch_name") or "").strip()
    specialist_engineer = (request.form.get("specialist_engineer") or "").strip()
    service_type = (request.form.get("service_type") or "").strip()
    request_notes = (request.form.get("request_notes") or "").strip()

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, role, session_branch)
        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_send_job_transfer(role, session_branch, job):
            flash("Only the owning branch can send this case to a specialist branch", "danger")
            return redirect(f"/job-details/{job_id}")
        if job.get("closure_status"):
            flash("Closed cases cannot be transferred", "danger")
            return redirect(f"/job-details/{job_id}")

        resolved_branch = _resolve_known_branch(cursor, to_branch_name)
        if not resolved_branch or resolved_branch == "ALL":
            flash("Select a valid specialist branch", "danger")
            return redirect(f"/job-details/{job_id}")
        if resolved_branch == job.get("branch_name"):
            flash("Choose a different specialist branch", "danger")
            return redirect(f"/job-details/{job_id}")

        existing_transfers = _fetch_job_service_transfer_rows(cursor, job_id)
        transfer_summary = summarize_job_transfers(existing_transfers, float(job.get("service_charges") or 0))
        if transfer_summary.get("has_active_transfer"):
            flash("This case already has an active specialist transfer", "warning")
            return redirect(f"/job-details/{job_id}")

        cursor.execute(
            """
            INSERT INTO job_service_transfers (
                job_id,
                from_branch_name,
                to_branch_name,
                specialist_engineer,
                service_type,
                request_notes,
                status,
                sent_by,
                updated_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, 'Sent', %s, %s)
            """,
            (
                job_id,
                job.get("branch_name"),
                resolved_branch,
                specialist_engineer or None,
                service_type or None,
                request_notes or None,
                session.get("username"),
                session.get("username"),
            ),
        )
        db.commit()
        flash(f"Case sent to specialist branch {resolved_branch}", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Could not send case to specialist branch", e)
    finally:
        _safe_close(cursor, db)

    return redirect(f"/job-details/{job_id}")


@app.route("/job-transfer/update/<int:transfer_id>", methods=["POST"])
def update_job_transfer(transfer_id):

    if "username" not in session:
        return redirect("/login")

    role = session.get("role")
    session_branch = session.get("branch")
    if role not in ["super_admin", "admin", "coordinator", "engineer"]:
        flash("Access denied", "danger")
        return redirect("/jobs")

    requested_status = normalize_transfer_status(request.form.get("status"), default="")
    db = None
    cursor = None
    transfer_row = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        transfer_row = _fetch_job_service_transfer_by_id(cursor, transfer_id)
        if not transfer_row:
            flash("Specialist transfer not found", "danger")
            return redirect("/jobs")

        job = _fetch_scoped_job(cursor, transfer_row.get("job_id"), role, session_branch)
        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_update_job_transfer(role, session_branch, job, transfer_row):
            flash("You do not have permission to update this specialist transfer", "danger")
            return redirect(f"/job-details/{transfer_row.get('job_id')}")
        if not requested_status or requested_status not in TRANSFER_STATUSES:
            flash("Select a valid specialist transfer status", "danger")
            return redirect(f"/job-details/{transfer_row.get('job_id')}")
        if not can_transition_transfer_status(transfer_row.get("status"), requested_status):
            flash("Invalid specialist transfer status change", "danger")
            return redirect(f"/job-details/{transfer_row.get('job_id')}")
        if requested_status == "Cancelled" and not _user_can_send_job_transfer(role, session_branch, job) and not _user_has_all_branch_scope(role, session_branch):
            flash("Only the owning branch can cancel this specialist transfer", "danger")
            return redirect(f"/job-details/{transfer_row.get('job_id')}")

        specialist_engineer = (request.form.get("specialist_engineer") or "").strip() or transfer_row.get("specialist_engineer") or ""
        service_type = (request.form.get("service_type") or "").strip() or transfer_row.get("service_type") or ""
        status_notes = (request.form.get("status_notes") or "").strip() or transfer_row.get("status_notes") or ""
        raw_internal_service_charge = (request.form.get("internal_service_charge") or "").strip()
        internal_service_charge = float(transfer_row.get("internal_service_charge") or 0)
        if raw_internal_service_charge != "":
            internal_service_charge = max(_parse_money(raw_internal_service_charge), 0)

        if requested_status in ["Completed", "Returned"] and not specialist_engineer:
            flash("Specialist engineer is required before marking service completed or returned", "danger")
            return redirect(f"/job-details/{transfer_row.get('job_id')}")

        set_clauses = [
            "status=%s",
            "specialist_engineer=%s",
            "service_type=%s",
            "status_notes=%s",
            "internal_service_charge=%s",
            "updated_by=%s",
            "updated_at=NOW()",
        ]
        params = [
            requested_status,
            specialist_engineer or None,
            service_type or None,
            status_notes or None,
            internal_service_charge,
            session.get("username"),
        ]

        if requested_status in ["In Service", "Completed", "Returned"]:
            set_clauses.append("accepted_at=COALESCE(accepted_at, NOW())")
        if requested_status in ["Completed", "Returned"]:
            set_clauses.append("completed_at=COALESCE(completed_at, NOW())")
        if requested_status == "Returned":
            set_clauses.append("returned_at=COALESCE(returned_at, NOW())")

        params.append(transfer_id)
        cursor.execute(
            f"UPDATE job_service_transfers SET {', '.join(set_clauses)} WHERE id=%s",
            tuple(params),
        )
        db.commit()
        flash("Specialist transfer updated", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Could not update specialist transfer", e)
    finally:
        _safe_close(cursor, db)

    redirect_job_id = transfer_row.get("job_id") if transfer_row else 0
    return redirect(f"/job-details/{redirect_job_id}" if redirect_job_id else "/jobs")


# ---------------- UPDATE STATUS ---------------- #

@app.route("/close-case/<int:job_id>", methods=["POST"])
def close_case(job_id):

    if "username" not in session:
        return redirect("/login")

    allowed_roles = ["super_admin", "admin", "coordinator"]
    if session.get("role") not in allowed_roles:
        flash("Access denied", "danger")
        return redirect("/jobs")

    role = session.get("role")
    session_branch = session.get("branch")

    closure_main = (request.form.get("closure_main") or "").strip()
    closure_service_type = (request.form.get("closure_service_type") or "").strip().upper()
    closure_sub_reason = (request.form.get("closure_sub_reason") or "").strip()
    closure_notes = (request.form.get("closure_notes") or "").strip()
    payment_cash = _parse_money(request.form.get("payment_cash"))
    payment_upi = _parse_money(request.form.get("payment_upi"))
    payment_card = _parse_money(request.form.get("payment_card"))
    close_case_redirect_url = f"/job-details/{job_id}?open_close_case=1"

    if closure_main not in ["Closed Success", "Closed Failed"]:
        flash("Please select a valid closure status", "danger")
        return redirect(close_case_redirect_url)

    valid_failed_reasons = [
        "Non-repairable",
        "Spare not available",
        "Customer not approved",
    ]
    valid_closure_service_types = ["SYSCARE", "REGULAR", "WARRANTY"]

    if closure_main == "Closed Failed":
        if closure_sub_reason not in valid_failed_reasons:
            flash("Please select a valid failed reason", "danger")
            return redirect(close_case_redirect_url)
        final_closure_status = f"Closed Failed - {closure_sub_reason}"
        final_closure_service_type = None
    else:
        if closure_service_type not in valid_closure_service_types:
            flash("Please select a valid service type for Closed Success", "danger")
            return redirect(close_case_redirect_url)
        final_closure_status = "Closed Success"
        final_closure_service_type = closure_service_type

    if len(closure_notes) < 3:
        flash("Closure notes are required (minimum 3 characters)", "danger")
        return redirect(close_case_redirect_url)

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        job = _fetch_scoped_job(cursor, job_id, role, session_branch)

        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_manage_job_core(role, session_branch, job):
            flash("Only the owning branch can close this case", "danger")
            return redirect(f"/job-details/{job_id}")

        if job.get("closure_status"):
            flash("Case is already closed", "warning")
            return redirect("/jobs")

        transfer_summary = summarize_job_transfers(_fetch_job_service_transfer_rows(cursor, job_id), float(job.get("service_charges") or 0))
        if transfer_summary.get("has_active_transfer"):
            flash("Return the active specialist transfer before closing this case", "danger")
            return redirect(f"/job-details/{job_id}")

        cursor.execute(
            "SELECT COUNT(*) AS spare_count, COALESCE(SUM(amount), 0) AS total_spares FROM used_spares WHERE job_id=%s",
            (job_id,),
        )
        spare_total_row = cursor.fetchone() or {}
        spare_count = int(spare_total_row.get("spare_count") or 0)
        total_spares_amount = float(spare_total_row.get("total_spares") or 0)
        service_charges = float(job.get("service_charges") or 0)
        bill_total_amount = total_spares_amount + service_charges

        if spare_count > 0:
            invoice_no = str(job.get("spares_invoice_no") or "").strip()
            invoice_date = job.get("spares_invoice_date")
            if not invoice_no or not invoice_date:
                flash("Used spare items are pending billing. Enter invoice details in Spare Billing before closing this case.", "danger")
                return redirect(f"/job-details/{job_id}?open_spare_billing=1&continue_close_case=1")

        if bill_total_amount > 0:
            payment_total = payment_cash + payment_upi + payment_card
            if payment_total <= 0:
                flash("Mode of payment amounts are required for bill amount", "danger")
                return redirect(close_case_redirect_url)

            if abs(payment_total - bill_total_amount) > 0.01:
                flash("Payment split must match total bill amount exactly", "danger")
                return redirect(close_case_redirect_url)
        else:
            payment_cash = 0
            payment_upi = 0
            payment_card = 0

        cursor.execute(
            """
            UPDATE jobs
            SET status=%s,
                closure_status=%s,
                closure_service_type=%s,
                closure_notes=%s,
                closure_date=NOW(),
                closed_by=%s,
                payment_cash=%s,
                payment_upi=%s,
                payment_card=%s
            WHERE id=%s
            """,
            (
                "Closed",
                final_closure_status,
                final_closure_service_type,
                closure_notes,
                session.get("username"),
                payment_cash,
                payment_upi,
                payment_card,
                job_id,
            ),
        )
        db.commit()
        flash("Case closed successfully", "success")

    except Error as e:
        _flash_internal_error("Failed to close case", e)

    finally:
        _safe_close(cursor, db)

    return redirect(f"/job-details/{job_id}")


@app.route("/reopen-case/<int:job_id>", methods=["POST"])
def reopen_case(job_id):

    if "username" not in session:
        return redirect("/login")

    allowed_roles = ["super_admin", "admin"]
    if session.get("role") not in allowed_roles:
        flash("Access denied", "danger")
        return redirect("/jobs")

    role = session.get("role")
    session_branch = session.get("branch")
    username = session.get("username")
    db = None
    cursor = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        job = _fetch_scoped_job(cursor, job_id, role, session_branch)

        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _user_can_manage_job_core(role, session_branch, job):
            flash("Only the owning branch can reopen this case", "danger")
            return redirect(f"/job-details/{job_id}")

        if not job.get("closure_status"):
            flash("Only closed cases can be reopened", "warning")
            return redirect(f"/job-details/{job_id}")

        previous_closure_status = str(job.get("closure_status") or "Closed").strip()
        previous_service_type = str(job.get("closure_service_type") or "").strip()
        reopen_note = f"Case reopened by {username} from {previous_closure_status}"
        if previous_service_type:
            reopen_note = f"{reopen_note} ({previous_service_type})"

        cursor.execute(
            """
            UPDATE jobs
            SET status=%s,
                closure_status=NULL,
                closure_service_type=NULL,
                closure_notes=NULL,
                closure_date=NULL,
                closed_by=NULL,
                payment_cash=0,
                payment_upi=0,
                payment_card=0,
                status_update_notes=%s,
                status_updated_by=%s,
                status_updated_at=NOW()
            WHERE id=%s
            """,
            (
                "Open",
                reopen_note,
                username,
                job_id,
            ),
        )

        cursor.execute(
            """
            INSERT INTO job_status_logs (job_id, status, notes, updated_by, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            """,
            (job_id, "Open", reopen_note, username)
        )

        db.commit()
        flash("Case reopened successfully", "success")

    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Failed to reopen case", e)

    finally:
        _safe_close(cursor, db)

    return redirect(f"/job-details/{job_id}")


# ---------------- UPDATE STATUS ---------------- #

@app.route("/update-status", methods=["POST"])
def update_status():

    if "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    allowed_roles = ["super_admin", "admin", "coordinator"]
    if session.get("role") not in allowed_roles:
        flash("Access denied", "danger")
        return redirect("/jobs")

    job_id = request.form.get("job_id")
    status = request.form.get("status")
    status_note = (request.form.get("status_note") or "").strip()
    return_to = (request.form.get("return_to") or "").strip()
    if not return_to.startswith("/jobs"):
        return_to = "/jobs"

    if not job_id or not status:
        flash("Missing parameters", "danger")
        return redirect(return_to)

    if len(status_note) < 3:
        flash("Status note is required (minimum 3 characters)", "danger")
        return redirect(return_to)

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        role = session.get("role")
        session_branch = session.get("branch")

        try:
            job_id_int = int(job_id)
        except (TypeError, ValueError):
            flash("Invalid job", "danger")
            return redirect(return_to)

        job = _fetch_scoped_job(cursor, job_id_int, role, session_branch)
        if not job:
            flash("Job not found", "danger")
            return redirect(return_to)
        if not _user_can_manage_job_core(role, session_branch, job):
            flash("Only the owning branch can update the main case status", "danger")
            return redirect(return_to)

        # Update job status
        cursor.execute(
            """
            UPDATE jobs
            SET status=%s,
                status_update_notes=%s,
                status_updated_by=%s,
                status_updated_at=NOW()
            WHERE id=%s
            """,
            (status, status_note, session.get("username"), job_id_int)
        )

        # Insert status log
        cursor.execute(
            """
            INSERT INTO job_status_logs (job_id, status, notes, updated_by, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            """,
            (job_id_int, status, status_note, session.get("username"))
        )

        db.commit()

        flash("Status updated", "success")

    except Error as e:
        _flash_internal_error("Status update failed", e)

    finally:
        _safe_close(cursor, db)

    return redirect(return_to)


# ---------------- CUSTOMER HISTORY ---------------- #

@app.route("/customer-history/<mobile>")
def customer_history(mobile):

    if "username" not in session:
        return jsonify({"found": False}), 401

    role = session.get("role")
    session_branch = session.get("branch")

    try:

        db = get_db()
        cursor = db.cursor(dictionary=True)

        branch_scope = _get_branch_scope(role, session_branch)
        if branch_scope is None:
            cursor.execute(
                "SELECT * FROM jobs WHERE mobile=%s ORDER BY id DESC LIMIT 10",
                (mobile,),
            )
        else:
            cursor.execute(
                "SELECT * FROM jobs WHERE mobile=%s AND branch_name=%s ORDER BY id DESC LIMIT 10",
                (mobile, branch_scope),
            )

        jobs = cursor.fetchall()

        if not jobs:
            return jsonify({"found": False})

        latest = jobs[0]

        recent = [
            {
                "job_number": j["job_number"],
                "device": j["device"],
                "status": j["status"]
            }
            for j in jobs[:3]
        ]

        return jsonify({
            "found": True,
            "customer": latest["customer_name"],
            "device": latest["device"],
            "model": latest["model"],
            "jobs": len(jobs),
            "recent": recent
        })

    except Error as e:
        return _json_internal_error("Could not load customer history", exc=e, payload={"found": False})

    finally:
        _safe_close(cursor, db)


# ---------------- DROPDOWN OPTIONS ---------------- #

@app.route("/add-option", methods=["POST"])
def add_option():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    type_name = request.form.get("type")
    value = request.form.get("value")

    if not type_name or not value:
        flash("Type and value are required", "danger")
        return redirect("/settings")


    try:
        db = get_db()
        cursor = db.cursor()
        # Find max order for this type
        cursor.execute("SELECT MAX(`order`) FROM dropdown_options WHERE type=%s", (type_name.strip(),))
        max_order = cursor.fetchone()[0] or 0
        cursor.execute(
            "INSERT INTO dropdown_options (type, value, `order`) VALUES (%s, %s, %s)",
            (type_name.strip(), value.strip(), max_order + 1)
        )
        db.commit()
        flash("Option added successfully", "success")
    except Error as e:
        # Duplicate entry is common if the same value already exists
        _flash_internal_error("Failed to add option", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")

# --- Dropdown Option Reordering Route ---
@app.route("/move-option/<int:option_id>/<direction>", methods=["POST"])
def move_option(option_id, direction):
    if "username" not in session:
        return redirect("/login")
    if session.get("role") != "super_admin":
        return "Access Denied"
    db = get_db()
    cursor = db.cursor(dictionary=True)
    # Get current option
    cursor.execute("SELECT id, type, `order` FROM dropdown_options WHERE id=%s", (option_id,))
    option = cursor.fetchone()
    if not option:
        flash("Option not found", "danger")
        return redirect("/settings")
    type_name = option["type"]
    current_order = option["order"]
    # Find the neighbor to swap with
    if direction == "up":
        cursor.execute("SELECT id, `order` FROM dropdown_options WHERE type=%s AND `order` < %s ORDER BY `order` DESC LIMIT 1", (type_name, current_order))
    else:
        cursor.execute("SELECT id, `order` FROM dropdown_options WHERE type=%s AND `order` > %s ORDER BY `order` ASC LIMIT 1", (type_name, current_order))
    neighbor = cursor.fetchone()
    if not neighbor:
        flash("Cannot move option further", "warning")
        return redirect("/settings")
    # Swap orders
    cursor.execute("UPDATE dropdown_options SET `order`=%s WHERE id=%s", (neighbor["order"], option_id))
    cursor.execute("UPDATE dropdown_options SET `order`=%s WHERE id=%s", (current_order, neighbor["id"]))
    db.commit()
    flash("Option reordered", "success")
    return redirect("/settings")


@app.route("/delete-option/<int:option_id>", methods=["POST"])
def delete_option(option_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("DELETE FROM dropdown_options WHERE id=%s", (option_id,))
        db.commit()
        flash("Option deleted", "success")
    except Error as e:
        _flash_internal_error("Failed to delete option", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/save-branch-print-profile", methods=["POST"])
def save_branch_print_profile():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    payload = _extract_branch_print_profile_payload(request.form)

    if not payload["branch_name"]:
        flash("Branch name is required", "danger")
        return redirect("/settings")

    if payload["branch_name"] == "ALL":
        flash("Use a real branch name, not ALL", "danger")
        return redirect("/settings")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            """
            INSERT INTO branch_print_profiles (
                branch_name, company_name, address_line1, address_line2, gst_no,
                mobile1, mobile2, mobile3, terms_text, quotation_terms
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                company_name=VALUES(company_name),
                address_line1=VALUES(address_line1),
                address_line2=VALUES(address_line2),
                gst_no=VALUES(gst_no),
                mobile1=VALUES(mobile1),
                mobile2=VALUES(mobile2),
                mobile3=VALUES(mobile3),
                terms_text=VALUES(terms_text),
                quotation_terms=VALUES(quotation_terms)
            """,
            (
                payload["branch_name"],
                payload["company_name"],
                payload["address_line1"],
                payload["address_line2"],
                payload["gst_no"],
                payload["mobile1"],
                payload["mobile2"],
                payload["mobile3"],
                payload["terms_text"],
                payload["quotation_terms"],
            ),
        )

        cursor.execute(
            "SELECT id FROM dropdown_options WHERE type=%s AND UPPER(value)=UPPER(%s) LIMIT 1",
            ("branch", payload["branch_name"]),
        )
        if not cursor.fetchone():
            cursor.execute("SELECT MAX(`order`) FROM dropdown_options WHERE type=%s", ("branch",))
            max_order = cursor.fetchone()[0] or 0
            cursor.execute(
                "INSERT INTO dropdown_options (type, value, `order`) VALUES (%s, %s, %s)",
                ("branch", payload["branch_name"], max_order + 1),
            )

        db.commit()
        flash("Branch saved successfully", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Failed to save branch", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/delete-branch-print-profile/<int:profile_id>", methods=["POST"])
def delete_branch_print_profile(profile_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT branch_name FROM branch_print_profiles WHERE id=%s", (profile_id,))
        profile = cursor.fetchone()
        if not profile:
            flash("Branch not found", "warning")
            return redirect("/settings")

        branch_name = str(profile.get("branch_name") or "").strip()
        cursor.execute("DELETE FROM branch_print_profiles WHERE id=%s", (profile_id,))
        cursor.execute(
            "DELETE FROM dropdown_options WHERE type=%s AND UPPER(value)=UPPER(%s)",
            ("branch", branch_name),
        )
        db.commit()
        flash("Branch deleted from settings", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Failed to delete branch", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/add-user", methods=["POST"])
def add_user():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    role = (request.form.get("role") or "").strip().lower()

    if not username or not password or not role:
        flash("Username, password and role are required", "danger")
        return redirect("/settings")

    if len(password) < 8:
        flash("Password must be at least 8 characters", "danger")
        return redirect("/settings")

    if role not in ["super_admin", "admin", "coordinator", "engineer"]:
        flash("Invalid role selected", "danger")
        return redirect("/settings")

    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
            (username, _hash_password(password), role)
        )
        db.commit()
        flash("User added successfully", "success")
    except Error as e:
        _flash_internal_error("Failed to add user", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/edit-user/<int:user_id>", methods=["POST"])
def edit_user(user_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    role = (request.form.get("role") or "").strip().lower()

    if not username or not role:
        flash("Username and role are required", "danger")
        return redirect("/settings")

    if role not in ["super_admin", "admin", "coordinator", "engineer"]:
        flash("Invalid role selected", "danger")
        return redirect("/settings")

    if password and len(password) < 8:
        flash("Password must be at least 8 characters", "danger")
        return redirect("/settings")

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT id, username, role FROM users WHERE id=%s", (user_id,))
        existing_user = cursor.fetchone()

        if not existing_user:
            flash("User not found", "warning")
            return redirect("/settings")

        if existing_user.get("role") == "super_admin":
            flash("Cannot edit super admin user", "danger")
            return redirect("/settings")

        cursor.execute(
            "SELECT id FROM users WHERE LOWER(username)=LOWER(%s) AND id<>%s",
            (username, user_id),
        )
        duplicate_user = cursor.fetchone()
        if duplicate_user:
            flash("Username already exists", "danger")
            return redirect("/settings")

        old_username = str(existing_user.get("username") or "")

        if password:
            cursor.execute(
                "UPDATE users SET username=%s, password=%s, role=%s WHERE id=%s",
                (username, _hash_password(password), role, user_id),
            )
        else:
            cursor.execute(
                "UPDATE users SET username=%s, role=%s WHERE id=%s",
                (username, role, user_id),
            )

        if old_username and old_username != username:
            cursor.execute(
                "UPDATE user_branches SET username=%s WHERE username=%s",
                (username, old_username),
            )

        db.commit()
        flash("User updated successfully", "success")
    except Error as e:
        _flash_internal_error("Failed to update user", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/delete-user/<int:user_id>", methods=["POST"])
def delete_user(user_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT username, role FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()

        if not user:
            flash("User not found", "warning")
            return redirect("/settings")

        if user.get("role") == "super_admin":
            flash("Cannot delete super admin user", "danger")
            return redirect("/settings")

        cursor.execute("DELETE FROM user_branches WHERE username=%s", (user.get("username"),))
        cursor.execute("DELETE FROM users WHERE id=%s", (user_id,))
        db.commit()
        flash("User deleted", "success")
    except Error as e:
        _flash_internal_error("Failed to delete user", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/add-user-branch", methods=["POST"])
def add_user_branch():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    map_username = (request.form.get("map_username") or "").strip()
    branch_name = (request.form.get("branch_name") or "").strip()

    if not map_username or not branch_name:
        flash("Username and branch are required", "danger")
        return redirect("/settings")

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT role FROM users WHERE username=%s", (map_username,))
        user = cursor.fetchone()
        if not user:
            flash("Selected user does not exist", "danger")
            return redirect("/settings")

        if user.get("role") in ["super_admin", "admin"]:
            flash("No branch mapping needed for admin/super_admin", "warning")
            return redirect("/settings")

        cursor.execute(
            "SELECT 1 FROM user_branches WHERE username=%s AND branch_name=%s",
            (map_username, branch_name)
        )
        if cursor.fetchone():
            flash("This branch mapping already exists", "warning")
            return redirect("/settings")

        cursor.execute(
            "INSERT INTO user_branches (username, branch_name) VALUES (%s, %s)",
            (map_username, branch_name)
        )
        db.commit()
        flash("User branch mapping added", "success")
    except Error as e:
        _flash_internal_error("Failed to add mapping", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/bulk-upload-user-branches", methods=["POST"])
def bulk_upload_user_branches():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    upload = request.files.get("user_branch_file")
    if not upload or not (upload.filename or "").strip():
        flash("Choose an Excel or CSV file to upload", "danger")
        return redirect("/settings")

    try:
        parsed_rows = list(_iter_user_branch_upload_rows(upload))
    except ValueError as e:
        _flash_internal_error("Could not read user-branch upload file", e)
        return redirect("/settings")
    except Exception as e:
        _flash_internal_error("Could not read user-branch upload file", e)
        return redirect("/settings")

    if not parsed_rows:
        flash("The file does not contain any data rows", "warning")
        return redirect("/settings")

    cursor = None
    db = None

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT username, role FROM users")
        users = {str(row.get("username") or "").strip(): row.get("role") for row in cursor.fetchall()}

        cursor.execute("SELECT username, branch_name FROM user_branches")
        existing_pairs = {
            (
                str(row.get("username") or "").strip().casefold(),
                str(row.get("branch_name") or "").strip().upper(),
            )
            for row in cursor.fetchall()
        }

        valid_branch_map = {branch.upper(): branch for branch in ["ALL", *_load_known_branches(cursor)]}

        rows_to_insert = []
        added_count = 0
        duplicate_count = 0
        skipped_count = 0
        empty_count = 0
        error_samples = []

        for row_number, username, branch_name in parsed_rows:
            if not username and not branch_name:
                empty_count += 1
                continue

            if not username or not branch_name:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: username and branch are both required")
                continue

            role = users.get(username)
            if not role:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: user '{username}' does not exist")
                continue

            if role in ["super_admin", "admin"]:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: admin users do not need branch mapping")
                continue

            normalized_branch = valid_branch_map.get(branch_name.strip().upper())
            if not normalized_branch:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: unknown branch '{branch_name}'")
                continue

            pair_key = (username.casefold(), normalized_branch.upper())
            if pair_key in existing_pairs:
                duplicate_count += 1
                continue

            existing_pairs.add(pair_key)
            rows_to_insert.append((username, normalized_branch))

        if rows_to_insert:
            cursor.executemany(
                "INSERT INTO user_branches (username, branch_name) VALUES (%s, %s)",
                rows_to_insert,
            )
            db.commit()
            added_count = len(rows_to_insert)

        if added_count:
            flash(
                f"Bulk upload complete: {added_count} added, {duplicate_count} duplicates skipped, {skipped_count} invalid rows, {empty_count} empty rows.",
                "success",
            )
        else:
            flash(
                f"No mappings added. {duplicate_count} duplicates skipped, {skipped_count} invalid rows, {empty_count} empty rows.",
                "warning",
            )

        for message in error_samples:
            flash(message, "warning")

    except Error as e:
        _flash_internal_error("Bulk user-branch upload failed", e)
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

    return redirect("/settings")


@app.route("/bulk-upload-users", methods=["POST"])
def bulk_upload_users():
    if "username" not in session:
        return redirect("/login")
    if session.get("role") != "super_admin":
        return "Access Denied"
    upload = request.files.get("users_file")
    if not upload or not (upload.filename or "").strip():
        flash("Choose Excel or CSV file", "danger")
        return redirect("/settings")
    try:
        parsed_rows = list(_iter_users_upload_rows(upload))
    except ValueError as e:
        _flash_internal_error("Could not read user upload file", e)
        return redirect("/settings")
    except Exception as e:
        _flash_internal_error("Could not read user upload file", e)
        return redirect("/settings")
    if not parsed_rows:
        flash("No data rows", "warning")
        return redirect("/settings")
    cursor = None
    db = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT username FROM users")
        existing_usernames = {str(row.get("username") or "").strip().lower() for row in cursor.fetchall()}
        valid_roles = {"super_admin", "admin", "coordinator", "engineer"}
        rows_to_insert = []
        added_count = duplicate_count = skipped_count = empty_count = 0
        error_samples = []
        for row_number, username, password, role in parsed_rows:
            if not username and not password and not role:
                empty_count += 1
                continue
            if not username or not password or not role:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: all fields required")
                continue
            if role.lower() not in valid_roles:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: invalid role")
                continue
            if len(password) < 8:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: password must be at least 8 characters")
                continue
            if username.lower() in existing_usernames:
                duplicate_count += 1
                continue
            existing_usernames.add(username.lower())
            rows_to_insert.append((username, _hash_password(password), role.lower()))
        if rows_to_insert:
            cursor.executemany("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", rows_to_insert)
            db.commit()
            added_count = len(rows_to_insert)
        if added_count:
            flash(f"Users: +{added_count}, dups {duplicate_count}, invalid {skipped_count}, empty {empty_count}", "success")
        else:
            flash(f"No users added. Dups {duplicate_count}, invalid {skipped_count}, empty {empty_count}", "warning")
        for message in error_samples:
            flash(message, "warning")
    except Error as e:
        _flash_internal_error("User upload failed", e)
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()
    return redirect("/settings")


@app.route("/bulk-upload-revenue-targets", methods=["POST"])
def bulk_upload_revenue_targets():
    if "username" not in session:
        return redirect("/login")
    if session.get("role") != "super_admin":
        return "Access Denied"
    upload = request.files.get("revenue_targets_file")
    if not upload or not (upload.filename or "").strip():
        flash("Choose Excel or CSV file", "danger")
        return redirect("/revenue-target-settings")
    try:
        parsed_rows = list(_iter_revenue_targets_upload_rows(upload))
    except ValueError as e:
        _flash_internal_error("Could not read revenue target upload file", e)
        return redirect("/revenue-target-settings")
    except Exception as e:
        _flash_internal_error("Could not read revenue target upload file", e)
        return redirect("/revenue-target-settings")
    if not parsed_rows:
        flash("No data rows", "warning")
        return redirect("/revenue-target-settings")
    cursor = None
    db = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        valid_branch_map = {branch.upper(): branch for branch in ["ALL", *_load_known_branches(cursor)]}
        cursor.execute("SELECT branch_name FROM branch_revenue_targets")
        existing_branches = {str(row.get("branch_name") or "").strip().upper() for row in cursor.fetchall()}
        rows_to_insert = []
        rows_to_update = []
        added_count = updated_count = skipped_count = empty_count = 0
        error_samples = []
        for row_number, branch, target_str in parsed_rows:
            if not branch:
                empty_count += 1
                continue
            normalized_branch = valid_branch_map.get(branch.strip().upper())
            if not normalized_branch:
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: unknown branch")
                continue
            try:
                target_amount = float(target_str or 0)
            except (ValueError, TypeError):
                skipped_count += 1
                if len(error_samples) < 5:
                    error_samples.append(f"Row {row_number}: invalid amount")
                continue
            branch_upper = normalized_branch.upper()
            if branch_upper in existing_branches:
                rows_to_update.append((target_amount, normalized_branch))
                updated_count += 1
            else:
                rows_to_insert.append((normalized_branch, target_amount))
                existing_branches.add(branch_upper)
        if rows_to_insert:
            cursor.executemany("INSERT INTO branch_revenue_targets (branch_name, total_target) VALUES (%s, %s)", rows_to_insert)
            added_count = len(rows_to_insert)
        if rows_to_update:
            for row in rows_to_update:
                cursor.execute("UPDATE branch_revenue_targets SET total_target=%s WHERE branch_name=%s", row)
        if rows_to_insert or rows_to_update:
            db.commit()
        if added_count or updated_count:
            flash(f"Targets: +{added_count}, upd {updated_count}, invalid {skipped_count}, empty {empty_count}", "success")
        else:
            flash(f"No targets changed. Invalid {skipped_count}, empty {empty_count}", "warning")
        for message in error_samples:
            flash(message, "warning")
    except Error as e:
        _flash_internal_error("Revenue target upload failed", e)
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()
    return redirect("/revenue-target-settings")


@app.route("/delete-user-branch/<int:mapping_id>", methods=["POST"])
def delete_user_branch(mapping_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("DELETE FROM user_branches WHERE id=%s", (mapping_id,))
        db.commit()
        flash("User branch mapping deleted", "success")
    except Error as e:
        _flash_internal_error("Failed to delete mapping", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/settings")


@app.route("/save-revenue-target", methods=["POST"])
def save_revenue_target():

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    branch_name = (request.form.get("target_branch_name") or "").strip()

    try:
        total_target = float((request.form.get("total_target") or "0").strip() or 0)
    except ValueError:
        total_target = 0.0

    # Backward compatibility if old form fields are still posted.
    if total_target <= 0:
        try:
            legacy_sales = float((request.form.get("sales_target") or "0").strip() or 0)
        except ValueError:
            legacy_sales = 0.0
        try:
            legacy_service = float((request.form.get("service_target") or "0").strip() or 0)
        except ValueError:
            legacy_service = 0.0
        total_target = legacy_sales + legacy_service

    total_target = max(total_target, 0)

    if not branch_name:
        flash("Branch is required for revenue target", "danger")
        return redirect("/revenue-target-settings")

    try:
        db = get_db()
        cursor = db.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO branch_revenue_targets (branch_name, total_target)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE
                    total_target=VALUES(total_target)
                """,
                (branch_name, total_target),
            )
        except Error as write_err:
            # Backward-compatible path when total_target column is not yet present.
            if getattr(write_err, "errno", None) == 1054 or "Unknown column" in str(write_err):
                cursor.execute(
                    """
                    INSERT INTO branch_revenue_targets (branch_name, sales_target, service_target)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        sales_target=VALUES(sales_target),
                        service_target=VALUES(service_target)
                    """,
                    (branch_name, total_target, 0),
                )
            else:
                raise
        db.commit()
        flash("Revenue target saved", "success")
    except Error as e:
        _flash_internal_error("Failed to save revenue target", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/revenue-target-settings")


@app.route("/delete-revenue-target/<int:target_id>", methods=["POST"])
def delete_revenue_target(target_id):

    if "username" not in session:
        return redirect("/login")

    if session.get("role") != "super_admin":
        return "Access Denied"

    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("DELETE FROM branch_revenue_targets WHERE id=%s", (target_id,))
        db.commit()
        flash("Revenue target deleted", "success")
    except Error as e:
        _flash_internal_error("Failed to delete revenue target", e)
    finally:
        _safe_close(cursor, db)

    return redirect("/revenue-target-settings")


def _fetch_job_for_print(cursor, job_id, role, session_branch):
    job = _fetch_scoped_job(cursor, job_id, role, session_branch)
    if not job:
        return None, None

    created_at = normalize_display_datetime(job.get("created_at"))

    return job, created_at


@app.route("/job-photo/<int:job_id>")
def job_photo(job_id):
    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, session.get("role"), session.get("branch"))
        attachments = _load_job_attachments(cursor, job_id, job.get("photo")) if job else []
        if not job or not attachments:
            return redirect(f"/job-details/{job_id}")
        return send_from_directory(app.config["JOB_PHOTO_FOLDER"], attachments[0]["filename"])
    except Error as e:
        _flash_internal_error("Could not load job photo", e)
        return redirect(f"/job-details/{job_id}")
    finally:
        _safe_close(cursor, db)


@app.route("/job-attachment/<int:job_id>/<path:filename>")
def job_attachment(job_id, filename):
    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, session.get("role"), session.get("branch"))
        if not job or not _job_can_access_attachment(cursor, job, filename):
            return redirect(f"/job-details/{job_id}")
        return send_from_directory(app.config["JOB_PHOTO_FOLDER"], filename)
    except Error as e:
        _flash_internal_error("Could not load attachment", e)
        return redirect(f"/job-details/{job_id}")
    finally:
        _safe_close(cursor, db)


@app.route("/job-attachment/<int:job_id>/<path:filename>/delete", methods=["POST"])
def delete_job_attachment(job_id, filename):
    if "username" not in session:
        return redirect("/login")

    db = None
    cursor = None
    role = session.get("role")
    session_branch = session.get("branch")
    normalized_filename = _normalize_attachment_filename(filename)

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        job = _fetch_scoped_job(cursor, job_id, role, session_branch)
        if not job:
            flash("Job not found", "danger")
            return redirect("/jobs")
        if not _branch_in_scope(role, session_branch, job.get("branch_name")):
            flash("Only the owning branch can delete attachments for this case", "danger")
            return redirect(f"/job-details/{job_id}")

        if not _user_can_manage_job_attachments(role, job):
            flash("You do not have permission to delete attachments for this case", "danger")
            return redirect(f"/job-details/{job_id}")

        if not normalized_filename or not _job_can_access_attachment(cursor, job, normalized_filename):
            flash("Attachment not found", "warning")
            return redirect(f"/job-details/{job_id}")

        deleted_filename, _remaining_filenames = _delete_job_attachment_record(cursor, job, normalized_filename)
        if not deleted_filename:
            flash("Attachment not found", "warning")
            return redirect(f"/job-details/{job_id}")

        db.commit()
        _remove_saved_images([deleted_filename], app.config["JOB_PHOTO_FOLDER"])
        flash("Attachment deleted", "success")
    except Error as e:
        if db:
            db.rollback()
        _flash_internal_error("Could not delete attachment", e)
    finally:
        _safe_close(cursor, db)

    return redirect(_get_safe_referrer_path() or f"/edit-job/{job_id}")


# ---------------- RUN SERVER ---------------- #
register_dashboard_settings_routes(
    app,
    {
        "get_db": get_db,
        "Error": Error,
        "build_dashboard_filters": build_dashboard_filters,
        "_normalize_date_input": _normalize_date_input,
        "_get_multi_values": _get_multi_values,
        "_normalize_option_list": _normalize_option_list,
        "_has_legacy_password": _has_legacy_password,
        "_add_csrf_protected_endpoints": _add_csrf_protected_endpoints,
        "load_known_branches": _load_known_branches,
        "decorate_job_rows_with_transfer_summary": _decorate_job_rows_with_transfer_summary,
        "decorate_job_rows_with_transfer_summary": _decorate_job_rows_with_transfer_summary,
        "format_datetime_display": format_datetime_display,
        "get_age_group": get_age_group,
        "get_branch_revenue_target": get_branch_revenue_target,
        "get_cashflow_dashboard_snapshot": _get_cashflow_dashboard_snapshot,
        "DEFAULT_BRANCHES": DEFAULT_BRANCHES,
    },
)

register_syscare_routes(
    app,
    {
        "get_db": get_db,
        "Error": Error,
        "_normalize_date_input": _normalize_date_input,
        "_next_sequence_value": _next_sequence_value,
        "_add_csrf_protected_endpoints": _add_csrf_protected_endpoints,
        "load_workbook": load_workbook,
    },
)

app.register_blueprint(
    create_onsite_calls_blueprint(
        app,
        {
            "db_host": DB_HOST,
            "db_user": DB_USER,
            "db_password": DB_PASSWORD,
            "db_name": DB_NAME,
            "pool_size": DB_POOL_SIZE,
            "default_branches": DEFAULT_BRANCHES,
            "csrf_registrar": _add_csrf_protected_endpoints,
        },
    )
)

@app.route("/profile-picture/<username>")
def profile_picture(username):
    if "username" not in session:
        return redirect("/login")

    if session.get("username") != username and session.get("role") != "super_admin":
        return "Access Denied", 403

    db = None
    cursor = None
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT profile_picture FROM users WHERE username=%s", (username,))
        user_record = cursor.fetchone() or {}
        profile_filename = user_record.get("profile_picture")
        if not profile_filename:
            return redirect("/profile")
        return send_from_directory(app.config["PROFILE_PICTURE_FOLDER"], profile_filename)
    except Error as e:
        _flash_internal_error("Could not load profile picture", e)
        return redirect("/profile")
    finally:
        _safe_close(cursor, db)

@app.route("/profile", methods=["GET", "POST"])
def profile():
    if "username" not in session:
        return redirect("/login")
    
    username = session.get("username")
    db = None
    cursor = None
    
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        if request.method == "POST":
            action = request.form.get("action", "")
            
            if action == "upload_picture":
                profile_pic = request.files.get("profile_picture")
                if profile_pic and profile_pic.filename:
                    try:
                        pic_filename = _save_image_upload(
                            profile_pic,
                            app.config["PROFILE_PICTURE_FOLDER"],
                            re.sub(r"[^a-z0-9]+", "_", username.lower()).strip("_") or "profile",
                        )
                    except ValueError as upload_error:
                        flash(str(upload_error), "danger")
                    else:
                        cursor.execute(
                            "SELECT profile_picture FROM users WHERE username=%s",
                            (username,),
                        )
                        current_row = cursor.fetchone() or {}
                        previous_picture = current_row.get("profile_picture")
                        if previous_picture:
                            old_path = os.path.join(app.config["PROFILE_PICTURE_FOLDER"], previous_picture)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        cursor.execute("UPDATE users SET profile_picture=%s WHERE username=%s", (pic_filename, username))
                        db.commit()
                        flash("Profile picture updated", "success")
            
            elif action == "change_password":
                old_password = request.form.get("old_password", "")
                new_password = request.form.get("new_password", "")
                confirm_password = request.form.get("confirm_password", "")
                
                if not old_password or not new_password or not confirm_password:
                    flash("All password fields required", "danger")
                elif new_password != confirm_password:
                    flash("New passwords do not match", "danger")
                elif len(new_password) < 8:
                    flash("New password must be at least 8 characters", "danger")
                else:
                    # Verify old password
                    cursor.execute("SELECT password FROM users WHERE username=%s", (username,))
                    user_record = cursor.fetchone()
                    password_ok, password_reset_required = _verify_password(
                        user_record.get("password") if user_record else "",
                        old_password,
                    )
                    if password_reset_required:
                        flash("This account needs an admin password reset before it can be used again.", "danger")
                    elif password_ok:
                        new_password_hash = _hash_password(new_password)
                        cursor.execute("UPDATE users SET password=%s WHERE username=%s", (new_password_hash, username))
                        db.commit()
                        flash("Password changed successfully", "success")
                    else:
                        flash("Current password incorrect", "danger")
            
            return redirect("/profile")
        
        # GET request - display profile
        cursor.execute(
            "SELECT username, role, profile_picture FROM users WHERE username=%s",
            (username,)
        )
        user_data = cursor.fetchone()
        
        if not user_data:
            flash("User not found", "danger")
            return redirect("/dashboard")
        
        return render_template("profile.html", user=user_data)
    
    except Error as e:
        _flash_internal_error("Profile update failed", e)
        return redirect("/dashboard")
    
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@app.route("/change-password", methods=["POST"])
def change_password_api():
    """JSON API for password change (for future use)"""
    if "username" not in session:
        return {"success": False, "message": "Not authenticated"}, 401
    
    data = request.get_json() or {}
    username = session.get("username")
    old_password = data.get("old_password", "")
    new_password = data.get("new_password", "")
    
    db = None
    cursor = None
    try:
        if not old_password or not new_password:
            return {"success": False, "message": "Old and new passwords required"}, 400
        
        if len(new_password) < 8:
            return {"success": False, "message": "New password must be at least 8 chars"}, 400
        
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("SELECT password FROM users WHERE username=%s", (username,))
        user_record = cursor.fetchone()
        
        password_ok, password_reset_required = _verify_password(
            user_record.get("password") if user_record else "",
            old_password,
        )
        if password_reset_required:
            return {"success": False, "message": "This account needs an admin password reset before it can be used again."}, 409
        if not password_ok:
            return {"success": False, "message": "Current password incorrect"}, 401
        
        cursor.execute("UPDATE users SET password=%s WHERE username=%s", (_hash_password(new_password), username))
        db.commit()
        
        return {"success": True, "message": "Password changed"}
    
    except Error as e:
        return _json_internal_error("Password change failed", exc=e)
    
    finally:
        _safe_close(cursor, db)


if __name__ == "__main__":
    if APP_DEBUG:
        app.run(debug=True, host=APP_HOST, port=APP_PORT, threaded=True)
    else:
        try:
            from waitress import serve

            configured_waitress_threads = max(1, int(os.getenv("WAITRESS_THREADS", str(DEFAULT_WAITRESS_THREADS))))
            waitress_threads = max(1, min(configured_waitress_threads, DB_POOL_SIZE))
            serve(app, host=APP_HOST, port=APP_PORT, threads=waitress_threads)
        except ImportError:
            app.run(debug=False, host=APP_HOST, port=APP_PORT, threaded=True)


